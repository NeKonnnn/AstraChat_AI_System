"""
MemoAI Web Backend - FastAPI приложение
Современный веб-интерфейс для MemoAI с поддержкой всех функций
"""

# Настройка кодировки для Windows
import sys
import os

# Импортируем утилиту для исправления кодировки
try:
    from utils.encoding_fix import fix_windows_encoding, safe_print
    fix_windows_encoding()
except ImportError:
    # Если утилита недоступна, используем базовую настройку
    if sys.platform == "win32":
        os.system("chcp 65001 >nul 2>&1")
        if hasattr(sys.stdout, 'reconfigure'):
            sys.stdout.reconfigure(encoding='utf-8')
        if hasattr(sys.stderr, 'reconfigure'):
            sys.stderr.reconfigure(encoding='utf-8')

from fastapi import FastAPI, WebSocket, WebSocketDisconnect, HTTPException, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse
import uvicorn
import asyncio
import json
import os
import sys
import traceback
from typing import List, Dict, Any, Optional
from datetime import datetime
import logging
from socketio import AsyncServer, ASGIApp
from starlette.applications import Starlette

# Добавляем корневую директорию в путь для импорта модулей
current_dir = os.path.dirname(os.path.abspath(__file__))
root_dir = os.path.dirname(current_dir)

# Загружаем переменные окружения из .env файла (до инициализации логгера)
try:
    from dotenv import load_dotenv
    env_path = os.path.join(root_dir, '.env')
    if os.path.exists(env_path):
        load_dotenv(env_path)
        print(f"✅ .env файл загружен: {env_path}")
        # Проверяем MongoDB настройки
        mongodb_user = os.getenv("MONGODB_USER", "").strip()
        mongodb_password = os.getenv("MONGODB_PASSWORD", "").strip()
        if mongodb_user.startswith('#') or mongodb_password.startswith('#'):
            print(f"⚠️ ВНИМАНИЕ: MONGODB_USER или MONGODB_PASSWORD начинаются с '#', будут игнорироваться")
    else:
        print(f"⚠️ .env файл не найден: {env_path}")
except ImportError:
    print("⚠️ python-dotenv не установлен, переменные окружения не будут загружены из .env")

# В Docker: /app содержит main.py, agent.py и т.д.
# Для импортов "from backend.xxx" нужно чтобы /app был доступен как /backend
# Создаем временную структуру для импортов
if current_dir == '/app' and not os.path.exists('/app/backend'):
    # В Docker контейнере создаем symbolic link
    os.system('ln -sf /app /app/backend')

sys.path.insert(0, current_dir)  # Для доступа к /app/config
sys.path.insert(0, root_dir)      # Для доступа к / для импортов "from backend.xxx"

# Импортируем конфигурацию
from config import get_config, config

# Настройка логирования в самом начале
# Настройка логирования с поддержкой UTF-8
import logging
logging.basicConfig(
    level=logging.DEBUG,
    format='[%(asctime)s] %(levelname)s [Backend] %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S',
    handlers=[
        logging.StreamHandler(sys.stdout)
    ]
)

# Настройка кодировки для обработчиков логирования
for handler in logging.root.handlers:
    if hasattr(handler, 'stream') and hasattr(handler.stream, 'reconfigure'):
        handler.stream.reconfigure(encoding='utf-8')
logger = logging.getLogger(__name__)
logger.info("Логирование настроено")

# Проверяем, что .env файл был загружен и MongoDB настройки доступны
logger.info("Проверка переменных окружения MongoDB...")
mongodb_host = os.getenv("MONGODB_HOST", "localhost")
mongodb_port = os.getenv("MONGODB_PORT", "27017")
mongodb_user = os.getenv("MONGODB_USER", "").strip()
mongodb_password = os.getenv("MONGODB_PASSWORD", "").strip()
logger.info(f"  MONGODB_HOST: {mongodb_host}")
logger.info(f"  MONGODB_PORT: {mongodb_port}")
logger.info(f"  MONGODB_USER: '{mongodb_user}' (len={len(mongodb_user)})")
logger.info(f"  MONGODB_PASSWORD: {'*' * len(mongodb_password) if mongodb_password else ''} (len={len(mongodb_password)})")
if mongodb_user.startswith('#') or mongodb_password.startswith('#'):
    logger.warning("⚠️ MONGODB_USER или MONGODB_PASSWORD начинаются с '#', будут игнорироваться")

# Импорт authentication router
try:
    logger.info("Попытка импорта auth router...")
    from backend.auth.routes import router as auth_router
    logger.info("auth router импортирован успешно")
except ImportError as e:
    logger.warning(f"auth router недоступен: {e}")
    auth_router = None
except Exception as e:
    logger.warning(f"Ошибка при импорте auth router: {e}")
    auth_router = None

# Импорты из оригинального MemoAI
try:
    logger.info("Попытка импорта agent...")
    from backend.agent import ask_agent, model_settings, update_model_settings, reload_model_by_path, get_model_info, initialize_model
    from backend.context_prompts import context_prompt_manager
    logger.info("agent импортирован успешно")
    if ask_agent:
        logger.info("ask_agent функция доступна")
    else:
        logger.warning("ask_agent функция не доступна")
except ImportError as e:
    logger.error(f"Ошибка импорта agent: {e}")
    print(f"Ошибка импорта agent: {e}")
    print(f"Текущий путь: {os.getcwd()}")
    print(f"Python path: {sys.path}")
    ask_agent = None
    model_settings = None
    update_model_settings = None
    reload_model_by_path = None
    get_model_info = None
    initialize_model = None
except Exception as e:
    logger.error(f"Неожиданная ошибка при импорте agent: {e}")
    import traceback
    logger.error(f"Traceback: {traceback.format_exc()}")
    ask_agent = None
    model_settings = None
    update_model_settings = None
    reload_model_by_path = None
    get_model_info = None
    initialize_model = None

# Попытка импорта llm-svc версии (если доступна)
try:
    logger.info("Попытка импорта agent_llm_svc...")
    from backend.agent_llm_svc import ask_agent as ask_agent_llm_svc, model_settings as model_settings_llm_svc, update_model_settings as update_model_settings_llm_svc, reload_model_by_path as reload_model_by_path_llm_svc, get_model_info as get_model_info_llm_svc, initialize_model as initialize_model_llm_svc
    logger.info("agent_llm_svc импортирован успешно")
    
    # Проверяем, нужно ли использовать llm-svc
    use_llm_svc = os.getenv('USE_LLM_SVC', 'false').lower() == 'true'
    
    # Отладочный вывод режима работы
    print(f"Текущий режим: {'llm-svc' if use_llm_svc else 'оригинальный agent.py'}")
    logger.info(f"Режим работы: {'llm-svc' if use_llm_svc else 'оригинальный agent.py'}")
    
    if use_llm_svc:
        logger.info("Переключение на llm-svc версию agent")
        ask_agent = ask_agent_llm_svc
        model_settings = model_settings_llm_svc
        update_model_settings = update_model_settings_llm_svc
        reload_model_by_path = reload_model_by_path_llm_svc
        get_model_info = get_model_info_llm_svc
        initialize_model = initialize_model_llm_svc
        logger.info("Успешно переключено на llm-svc")
    else:
        logger.info("Используется оригинальная версия agent")
except ImportError as e:
    logger.warning(f"agent_llm_svc недоступен: {e}")
except Exception as e:
    logger.warning(f"Ошибка при импорте agent_llm_svc: {e}")

# Общая блокировка для загрузки моделей в режиме multi-llm
# Используется для предотвращения конфликтов при параллельной загрузке
import threading
model_load_lock = threading.Lock()
    
try:
    logger.info("Попытка импорта memory_service (MongoDB)...")
    from backend.database.memory_service import (
        save_dialog_entry, 
        load_dialog_history, 
        clear_dialog_history, 
        get_recent_dialog_history,
        reset_conversation,
        get_or_create_conversation_id
    )
    logger.info("memory_service (MongoDB) импортирован успешно")
    logger.info(f"save_dialog_entry импортирован: {save_dialog_entry is not None}, type: {type(save_dialog_entry)}")
    if save_dialog_entry:
        logger.info("save_dialog_entry функция доступна (MongoDB)")
    else:
        logger.error("save_dialog_entry функция не доступна (None или False)")

except ImportError as e:
    logger.error(f"Ошибка импорта memory_service: {e}")
    logger.error("Попытка использовать старый memory модуль (JSON)...")
    try:
        from backend.memory import save_dialog_entry, load_dialog_history, clear_dialog_history, get_recent_dialog_history
        logger.warning("Используется старый memory модуль (JSON)")
        reset_conversation = None
        get_or_create_conversation_id = None
    except:
        logger.error("Ни один модуль памяти не доступен!")
        save_dialog_entry = None
        load_dialog_entry = None
        load_dialog_history = None
        clear_dialog_history = None
        get_recent_dialog_history = None
        reset_conversation = None
        get_or_create_conversation_id = None
except Exception as e:
    logger.error(f"Неожиданная ошибка при импорте memory: {e}")
    import traceback
    logger.error(f"Traceback: {traceback.format_exc()}")
    save_dialog_entry = None
    load_dialog_entry = None
    load_dialog_history = None
    clear_dialog_history = None
    get_recent_dialog_history = None
    reset_conversation = None
    get_or_create_conversation_id = None
    
try:
    logger.info("Попытка импорта voice...")
    from backend.voice import speak_text, recognize_speech, recognize_speech_from_file, check_vosk_model
    logger.info("voice импортирован успешно")

except ImportError as e:
    logger.error(f"Ошибка импорта voice: {e}")
    print(f"Ошибка импорта voice: {e}")
    speak_text = None
    recognize_speech = None
    recognize_speech_from_file = None
    check_vosk_model = None
except Exception as e:
    logger.error(f"Неожиданная ошибка при импорте voice: {e}")
    import traceback
    logger.error(f"Traceback: {traceback.format_exc()}")
    speak_text = None
    recognize_speech = None
    recognize_speech_from_file = None
    check_vosk_model = None

# Импорт MinIO клиента для хранения временных файлов
try:
    logger.info("Попытка импорта MinIO клиента...")
    from backend.database.minio import get_minio_client
    minio_client = get_minio_client()
    if minio_client:
        logger.info("MinIO клиент успешно инициализирован")
    else:
        logger.warning("MinIO клиент недоступен, будут использоваться локальные временные файлы")
except ImportError as e:
    logger.warning(f"MinIO клиент недоступен: {e}. Будут использоваться локальные временные файлы")
    minio_client = None
except Exception as e:
    logger.warning(f"Ошибка при инициализации MinIO клиента: {e}. Будут использоваться локальные временные файлы")
    import traceback
    logger.error(f"Traceback: {traceback.format_exc()}")
    minio_client = None

try:
    logger.info("Попытка импорта document_processor...")
    from backend.document_processor import DocumentProcessor
    logger.info("document_processor импортирован успешно")
except ImportError as e:
    logger.error(f"Ошибка импорта document_processor: {e}")
    print("Предупреждение: модуль document_processor не найден")
    DocumentProcessor = None
except Exception as e:
    logger.error(f"Неожиданная ошибка при импорте document_processor: {e}")
    import traceback
    logger.error(f"Traceback: {traceback.format_exc()}")
    DocumentProcessor = None
    
try:
    logger.info("Попытка импорта universal_transcriber...")
    from backend.universal_transcriber import UniversalTranscriber
    logger.info("universal_transcriber импортирован успешно")
except ImportError as e:
    logger.error(f"Ошибка импорта universal_transcriber: {e}")
    import traceback
    logger.error(f"Traceback: {traceback.format_exc()}")
    print("Предупреждение: модуль universal_transcriber не найден")
    UniversalTranscriber = None
except Exception as e:
    logger.error(f"Неожиданная ошибка при импорте universal_transcriber: {e}")
    import traceback
    logger.error(f"Traceback: {traceback.format_exc()}")
    UniversalTranscriber = None
    
try:
    logger.info("Попытка импорта online_transcription...")
    from backend.online_transcription import OnlineTranscriber
    logger.info("online_transcription импортирован успешно")
    if OnlineTranscriber:
        logger.info("OnlineTranscriber класс доступен")
    else:
        logger.warning("OnlineTranscriber класс не доступен")
except ImportError as e:
    logger.error(f"Ошибка импорта online_transcription: {e}")
    logger.error(f"Traceback: {traceback.format_exc()}")
    print("Предупреждение: модуль online_transcription не найден")
    OnlineTranscriber = None
except Exception as e:
    logger.error(f"Неожиданная ошибка при импорте online_transcription: {e}")
    logger.error(f"Traceback: {traceback.format_exc()}")
    OnlineTranscriber = None

# Импорт агентной архитектуры
try:
    logger.info("Попытка импорта агентной архитектуры...")
    from backend.orchestrator import initialize_agent_orchestrator, get_agent_orchestrator
    logger.info("Агентная архитектура импортирована успешно")
except ImportError as e:
    logger.error(f"Ошибка импорта агентной архитектуры: {e}")
    print("Предупреждение: модуль агентной архитектуры не найден")
    initialize_agent_orchestrator = None
    get_agent_orchestrator = None
except Exception as e:
    logger.error(f"Неожиданная ошибка при импорте агентной архитектуры: {e}")
    logger.error(f"Traceback: {traceback.format_exc()}")
    initialize_agent_orchestrator = None
    get_agent_orchestrator = None

# Импорт модуля баз данных
try:
    logger.info("Попытка импорта database модуля...")
    from backend.database.init_db import (
        init_databases, 
        close_databases,
        get_conversation_repository,
        get_document_repository,
        get_vector_repository
    )
    logger.info("Database модуль импортирован успешно")
    database_available = True
except ImportError as e:
    logger.warning(f"Database модуль недоступен: {e}")
    logger.warning("Приложение будет работать без баз данных (файловый режим)")
    init_databases = None
    close_databases = None
    get_conversation_repository = None
    get_document_repository = None
    get_vector_repository = None
    database_available = False
except Exception as e:
    logger.error(f"Неожиданная ошибка при импорте database модуля: {e}")
    logger.error(f"Traceback: {traceback.format_exc()}")
    init_databases = None
    close_databases = None
    get_conversation_repository = None
    get_document_repository = None
    get_vector_repository = None
    database_available = False

# Глобальный словарь для хранения флагов остановки генерации
stop_generation_flags = {}

# Глобальный флаг для остановки голосового чата
voice_chat_stop_flag = False

# Создание Socket.IO сервера
sio = AsyncServer(
    async_mode='asgi',
    cors_allowed_origins=[
        "http://localhost:3000", 
        "http://127.0.0.1:3000",
        "http://localhost:8000",
        "http://127.0.0.1:8000",
        "http://memoai-frontend:3000",
        "http://memoai-backend:8000",
    ],
    ping_timeout=120,  # ping timeout до 2 минут
    ping_interval=25,  # Отправляем ping каждые 25 секунд
    logger=True,  # Включаем логирование для отладки
    engineio_logger=True  # Включаем логирование engine.io
)

# Создание FastAPI приложения с конфигурацией
app_config = config.get("app", {})
app = FastAPI(
    title=app_config.get("name", "MemoAI Web API"),
    description=app_config.get("description", "Веб-интерфейс для персонального AI-ассистента MemoAI"),
    version=app_config.get("version", "1.0.0"),
    debug=app_config.get("debug", False)
)

# Настройка CORS из конфигурации
cors_config = config.get("cors", {})
app.add_middleware(
    CORSMiddleware,
    allow_origins=cors_config.get("allowed_origins", [
        "http://localhost:3000", 
        "http://127.0.0.1:3000",
        "http://localhost:3001",
        "http://127.0.0.1:3001",
        "http://localhost:5173",  # Vite dev server
        "http://127.0.0.1:5173",
        "http://localhost:8000",
        "http://127.0.0.1:8000",
        "http://memoai-frontend:3000",
        "http://memoai-backend:8000",
    ]),
    allow_credentials=cors_config.get("allow_credentials", True),
    allow_methods=cors_config.get("allow_methods", ["*"]),
    allow_headers=cors_config.get("allow_headers", ["*"]),
)

# Подключаем authentication routes
if auth_router:
    app.include_router(auth_router)
    logger.info("✅ Auth routes подключены (/api/auth/*)")
else:
    logger.warning("⚠️ Auth routes не подключены (auth_router недоступен)")

# Startup событие для инициализации агентной архитектуры и баз данных
@app.on_event("startup")
async def startup_event():
    """Инициализация при запуске приложения"""
    logger.info("Запуск приложения...")
    
    # Инициализируем базы данных
    logger.info(f"Проверка доступности баз данных: init_databases={init_databases is not None}, database_available={database_available}")
    if init_databases and database_available:
        try:
            logger.info("Инициализация баз данных...")
            success = await init_databases()
            if success:
                logger.info("Базы данных успешно инициализированы")
                logger.info("  - MongoDB: готов для хранения диалогов")
                logger.info("  - PostgreSQL + pgvector: готов для RAG системы")
                # Проверяем статус MinIO
                if minio_client:
                    logger.info(f"  - MinIO: готов для хранения файлов (endpoint: {minio_client.endpoint})")
                else:
                    logger.warning("  - MinIO: не инициализирован, используется локальное хранение")
            else:
                logger.warning("Не удалось инициализировать некоторые базы данных")
                logger.warning("Приложение продолжит работу в файловом режиме")
        except Exception as e:
            logger.error(f"Ошибка инициализации баз данных: {e}")
            logger.error(f"Traceback: {traceback.format_exc()}")
            logger.warning("Приложение продолжит работу в файловом режиме")
    else:
        if not init_databases:
            logger.warning("⚠️ init_databases не импортирован или недоступен")
        if not database_available:
            logger.warning("⚠️ database_available = False")
        logger.warning("⚠️ Базы данных не настроены, используется файловый режим")
    
    # Инициализируем агентную архитектуру
    if initialize_agent_orchestrator:
        try:
            success = await initialize_agent_orchestrator()
            if success:
                logger.info("Агентная архитектура успешно инициализирована")
            else:
                logger.warning("Не удалось инициализировать агентную архитектуру")
        except Exception as e:
            logger.error(f"Ошибка инициализации агентной архитектуры: {e}")
    
    logger.info("Приложение запущено")

# Shutdown событие для корректного закрытия подключений
@app.on_event("shutdown")
async def shutdown_event():
    """Корректное закрытие при остановке приложения"""
    logger.info("Остановка приложения...")
    
    # Закрываем подключения к базам данных
    if close_databases and database_available:
        try:
            logger.info("Закрытие подключений к базам данных...")
            await close_databases()
            logger.info("Подключения к базам данных закрыты")
        except Exception as e:
            logger.error(f"Ошибка при закрытии баз данных: {e}")
    
    logger.info("Приложение остановлено")

# Создание Starlette приложения для Socket.IO
starlette_app = Starlette()
socket_app = ASGIApp(sio, starlette_app)

# Монтирование Socket.IO
app.mount("/socket.io", socket_app)

# Инициализация сервисов
logger.info("=== Инициализация сервисов ===")

try:
    logger.info("Импортируем DocumentProcessor...")
    doc_processor = DocumentProcessor() if DocumentProcessor else None
    if doc_processor:
        logger.info("DocumentProcessor инициализирован успешно")
        # Проверяем состояние
        doc_list = doc_processor.get_document_list()
        logger.info(f"Начальное состояние документов: {doc_list}")
        logger.info(f"Количество документов: {len(doc_list) if doc_list else 0}")
        
        # Проверяем атрибуты
        logger.info(f"Vectorstore доступен: {hasattr(doc_processor, 'vectorstore')}")
        logger.info(f"Documents доступен: {hasattr(doc_processor, 'documents')}")
        logger.info(f"Doc_names доступен: {hasattr(doc_processor, 'doc_names')}")
        logger.info(f"Embeddings доступен: {hasattr(doc_processor, 'embeddings')}")
        
        if hasattr(doc_processor, 'vectorstore'):
            logger.info(f"Vectorstore значение: {doc_processor.vectorstore is not None}")
            if doc_processor.vectorstore:
                logger.info("Vectorstore инициализирован успешно")
            else:
                logger.warning("Vectorstore не инициализирован")
        if hasattr(doc_processor, 'documents'):
            logger.info(f"Documents значение: {len(doc_processor.documents) if doc_processor.documents else 0}")
            if doc_processor.documents:
                logger.info("Documents коллекция содержит документы")
            else:
                logger.info("Documents коллекция пуста")
        if hasattr(doc_processor, 'doc_names'):
            logger.info(f"Doc_names значение: {len(doc_processor.doc_names) if doc_processor.doc_names else 0}")
            if doc_processor.doc_names:
                logger.info("Doc_names содержит имена документов")
            else:
                logger.info("Doc_names пуст")
        if hasattr(doc_processor, 'embeddings'):
            logger.info(f"Embeddings значение: {doc_processor.embeddings is not None}")
            if doc_processor.embeddings:
                logger.info("Embeddings модель загружена успешно")
            else:
                logger.warning("Embeddings модель не загружена")
    else:
        logger.warning("DocumentProcessor не доступен")
except Exception as e:
    logger.error(f"Ошибка инициализации DocumentProcessor: {e}")
    logger.error(f"Traceback: {traceback.format_exc()}")
    doc_processor = None

try:
    if UniversalTranscriber:
        logger.info("Инициализация UniversalTranscriber с движком whisperx...")
        transcriber = UniversalTranscriber(engine="whisperx")
        if transcriber:
            logger.info("UniversalTranscriber инициализирован успешно")
        else:
            logger.warning("UniversalTranscriber не удалось создать")
    else:
        logger.warning("UniversalTranscriber не доступен")
        transcriber = None
except Exception as e:
    logger.error(f"Ошибка инициализации UniversalTranscriber: {e}")
    logger.error(f"Traceback: {traceback.format_exc()}")
    transcriber = None

try:
    if OnlineTranscriber:
        logger.info("Инициализация OnlineTranscriber...")
        online_transcriber = OnlineTranscriber()
        if online_transcriber:
            logger.info("OnlineTranscriber инициализирован успешно")
        else:
            logger.warning("OnlineTranscriber не удалось создать")
    else:
        logger.warning("OnlineTranscriber класс не доступен")
        online_transcriber = None
except Exception as e:
    logger.error(f"Ошибка инициализации OnlineTranscriber: {e}")
    logger.error(f"Traceback: {traceback.format_exc()}")
    online_transcriber = None

logger.info("=== Инициализация сервисов завершена ===")

# Глобальные настройки транскрибации
current_transcription_engine = "whisperx"
current_transcription_language = "ru"

# Глобальные настройки памяти
memory_max_messages = 20
memory_include_system_prompts = True
memory_clear_on_restart = False

# Путь к файлу настроек
SETTINGS_FILE = os.path.join(os.path.dirname(__file__), "..", "settings.json")

def load_app_settings():
    """Загрузить настройки приложения из файла"""
    global current_transcription_engine, current_transcription_language, memory_max_messages, memory_include_system_prompts, memory_clear_on_restart
    
    try:
        if os.path.exists(SETTINGS_FILE):
            with open(SETTINGS_FILE, 'r', encoding='utf-8') as f:
                settings = json.load(f)
            
            current_transcription_engine = settings.get('transcription_engine', 'whisperx')
            current_transcription_language = settings.get('transcription_language', 'ru')
            
            # Загружаем настройки памяти
            memory_max_messages = settings.get('memory_max_messages', 20)
            memory_include_system_prompts = settings.get('memory_include_system_prompts', True)
            memory_clear_on_restart = settings.get('memory_clear_on_restart', False)
            
            logger.info(f"Настройки загружены: engine={current_transcription_engine}, language={current_transcription_language}, memory_max_messages={memory_max_messages}")
            return settings
    except Exception as e:
        logger.error(f"Ошибка загрузки настроек: {e}")
    
    # Возвращаем дефолтные настройки
    return {
        'transcription_engine': current_transcription_engine,
        'transcription_language': current_transcription_language,
        'memory_max_messages': memory_max_messages,
        'memory_include_system_prompts': memory_include_system_prompts,
        'memory_clear_on_restart': memory_clear_on_restart,
        'current_model_path': None
    }

def save_app_settings(settings_to_save):
    """Сохранить настройки приложения в файл"""
    try:
        # Загружаем существующие настройки
        existing_settings = {}
        if os.path.exists(SETTINGS_FILE):
            with open(SETTINGS_FILE, 'r', encoding='utf-8') as f:
                existing_settings = json.load(f)
        
        # Обновляем настройки
        existing_settings.update(settings_to_save)
        
        # Сохраняем в файл
        with open(SETTINGS_FILE, 'w', encoding='utf-8') as f:
            json.dump(existing_settings, f, ensure_ascii=False, indent=2)
        
        logger.info(f"Настройки сохранены: {settings_to_save}")
        return True
    except Exception as e:
        logger.error(f"Ошибка сохранения настроек: {e}")
        return False

# Загружаем настройки при старте
loaded_settings = load_app_settings()

# Очищаем память при перезапуске, если это настроено
if memory_clear_on_restart and clear_dialog_history:
    try:
        logger.info("Очистка памяти при перезапуске (настройка включена)")
        clear_dialog_history()
        logger.info("Память очищена при перезапуске")
    except Exception as e:
        logger.warning(f"Не удалось очистить память при перезапуске: {e}")

# WebSocket менеджер для управления соединениями
class ConnectionManager:
    def __init__(self):
        self.active_connections: List[WebSocket] = []

    async def connect(self, websocket: WebSocket):
        await websocket.accept()
        self.active_connections.append(websocket)
        logger.info(f"WebSocket connection established. Total connections: {len(self.active_connections)}")

    def disconnect(self, websocket: WebSocket):
        self.active_connections.remove(websocket)
        logger.info(f"WebSocket connection closed. Total connections: {len(self.active_connections)}")

    async def send_personal_message(self, message: str, websocket: WebSocket):
        await websocket.send_text(message)

    async def broadcast(self, message: str):
        for connection in self.active_connections:
            try:
                await connection.send_text(message)
            except:
                pass

manager = ConnectionManager()

# Socket.IO события
@sio.event
async def connect(sid, environ):
    logger.info(f"Socket.IO client connected: {sid}")
    # Очищаем флаг остановки при подключении
    stop_generation_flags[sid] = False
    await sio.emit('connected', {'data': 'Connected to MemoAI'}, room=sid)

@sio.event
async def disconnect(sid):
    logger.info(f"Socket.IO client disconnected: {sid}")
    # Удаляем флаг остановки при отключении
    if sid in stop_generation_flags:
        del stop_generation_flags[sid]

@sio.event
async def ping(sid, data):
    """Обработка heartbeat ping от клиента"""
    try:
        # Отвечаем pong для подтверждения что сервер жив
        await sio.emit('pong', {
            'timestamp': data.get('timestamp', 0),
            'server_time': datetime.now().isoformat()
        }, room=sid)
    except Exception as e:
        logger.error(f"Ошибка обработки ping: {e}")

@sio.event
async def stop_generation(sid, data):
    """Обработка команды остановки генерации через Socket.IO"""
    logger.info(f"Socket.IO: получена команда остановки генерации от {sid}")
    
    # Устанавливаем флаг остановки для этого пользователя
    stop_generation_flags[sid] = True  # True = остановить генерацию
    
    # Отправляем подтверждение остановки
    await sio.emit('generation_stopped', {
        'content': 'Генерация остановлена',
        'timestamp': datetime.now().isoformat()
    }, room=sid)
    
    logger.info(f"Socket.IO: установлен флаг остановки для {sid}")

@sio.event
async def chat_message(sid, data):
    """Обработка сообщений чата через Socket.IO"""
    if not ask_agent or not save_dialog_entry:
        logger.error(f"AI services недоступны: ask_agent={ask_agent is not None}, save_dialog_entry={save_dialog_entry is not None}")
        if not ask_agent:
            logger.error("ask_agent функция не доступна - проверьте импорт agent модуля")
        if not save_dialog_entry:
            logger.error("save_dialog_entry функция не доступна - проверьте импорт memory_service модуля")
        await sio.emit('chat_error', {
            'error': 'AI services not available'
        }, room=sid)
        return
        
    try:
        user_message = data.get("message", "")
        streaming = data.get("streaming", True)
        
        logger.info(f"Socket.IO chat: {user_message[:50]}...")
        
        # Сбрасываем флаг остановки для нового сообщения
        stop_generation_flags[sid] = False
        
        # Получаем историю
        logger.info(f"DEBUG: get_recent_dialog_history = {get_recent_dialog_history}")
        logger.info(f"DEBUG: type = {type(get_recent_dialog_history)}")
        if get_recent_dialog_history:
            logger.info("DEBUG: Вызываем get_recent_dialog_history...")
            history = await get_recent_dialog_history(max_entries=memory_max_messages)
            logger.info(f"DEBUG: История получена, длина = {len(history)}")
        else:
            logger.info("DEBUG: get_recent_dialog_history недоступен, используем пустую историю")
            history = []
        
        # Сохраняем сообщение пользователя
        try:
            # Получаем message_id и conversation_id из данных, если они переданы
            user_message_id = data.get("message_id", None)
            conversation_id = data.get("conversation_id", None)
            await save_dialog_entry("user", user_message, None, user_message_id, conversation_id)
        except RuntimeError as e:
            # Ошибка MongoDB - продолжаем работу, но логируем
            error_msg = str(e)
            if "MongoDB" in error_msg:
                logger.error(f"MongoDB недоступен. Сообщение не будет сохранено: {e}")
                await sio.emit('chat_error', {
                    'error': 'MongoDB недоступен. Невозможно сохранить сообщение.'
                }, room=sid)
                return
            else:
                # Другие ошибки - пробрасываем дальше
                raise
        
        # Проверяем, доступна ли агентная архитектура
        orchestrator = get_agent_orchestrator()
        use_agent_mode = orchestrator and orchestrator.get_mode() == "agent"
        use_multi_llm_mode = orchestrator and orchestrator.get_mode() == "multi-llm"
        
        logger.info(f"Socket.IO DEBUG: orchestrator = {orchestrator is not None}")
        if orchestrator:
            logger.info(f"Socket.IO DEBUG: orchestrator.get_mode() = '{orchestrator.get_mode()}'")
        logger.info(f"Socket.IO DEBUG: use_agent_mode = {use_agent_mode}, use_multi_llm_mode = {use_multi_llm_mode}")
        
        # Функция для отправки частей ответа
        async def async_stream_callback(chunk: str, accumulated_text: str):
            try:
                logger.info(f"Отправляем chunk: '{chunk[:50]}...', накоплено: {len(accumulated_text)} символов")
                await sio.emit('chat_chunk', {
                    'chunk': chunk,
                    'accumulated': accumulated_text
                }, room=sid)
                logger.info(f"Chunk отправлен успешно")
            except Exception as e:
                logger.error(f"Ошибка отправки chunk: {e}")
                pass
        
        # Переменная для хранения event loop
        loop = asyncio.get_event_loop()
        
        # Синхронная обертка для потокового callback
        def sync_stream_callback(chunk: str, accumulated_text: str):
            try:
                # Проверяем флаг остановки
                if stop_generation_flags.get(sid, False):
                    logger.info(f"Socket.IO: генерация остановлена для {sid}, возвращаем False")
                    return False
                
                # Планируем выполнение в основном event loop
                asyncio.run_coroutine_threadsafe(
                    async_stream_callback(chunk, accumulated_text), 
                    loop
                )
                
                return True
            except Exception as e:
                logger.error(f"Ошибка планирования задачи для chunk: {e}")
                return True
        
        try:
            # Если включен режим multi-llm, генерируем ответы от нескольких моделей параллельно
            if use_multi_llm_mode:
                logger.info("Socket.IO: РЕЖИМ MULTI-LLM: Параллельная генерация от нескольких моделей")
                logger.info(f"Socket.IO: Запрос пользователя: '{user_message[:100]}{'...' if len(user_message) > 100 else ''}'")
                
                multi_llm_models = orchestrator.get_multi_llm_models()
                if not multi_llm_models:
                    logger.warning("Socket.IO: Режим multi-llm активирован, но модели не выбраны")
                    await sio.emit('chat_error', {
                        'error': 'Режим multi-llm активирован, но модели не выбраны'
                    }, room=sid)
                    return
                
                logger.info(f"Socket.IO: Генерируем ответы от моделей: {multi_llm_models}")
                
                # Проверяем, есть ли загруженные документы
                doc_context = None
                if doc_processor:
                    doc_list = doc_processor.get_document_list()
                    if doc_list and len(doc_list) > 0:
                        logger.info(f"Socket.IO: Найдены документы в режиме multi-llm: {doc_list}")
                        try:
                            doc_context = doc_processor.get_document_context(user_message)
                            logger.info(f"Socket.IO: Получен контекст документов для multi-llm, длина: {len(doc_context) if doc_context else 0} символов")
                        except Exception as e:
                            logger.error(f"Socket.IO: Ошибка при получении контекста документов: {e}")
                
                # Формируем финальное сообщение с контекстом документов, если есть
                final_user_message = user_message
                if doc_context:
                    final_user_message = f"""Контекст из загруженных документов:
{doc_context}

Вопрос пользователя: {user_message}

Пожалуйста, ответьте на вопрос пользователя, используя информацию из предоставленных документов. Если в документах нет информации для ответа, честно скажите об этом."""
                
                # Получаем event loop для отправки чанков из текущего async контекста
                loop = asyncio.get_running_loop()
                
                # Функция для генерации ответа от одной модели
                async def generate_single_model_response(model_name: str):
                    try:
                        logger.info(f"Socket.IO: Начинаем генерацию от модели: {model_name}")
                        
                        # Отправляем событие начала генерации для этой модели
                        await sio.emit('multi_llm_start', {
                            'model': model_name,
                            'total_models': len(multi_llm_models),
                            'models': multi_llm_models
                        }, room=sid)
                        
                        # Логируем сообщение с контекстом для отладки
                        logger.info(f"Socket.IO: Модель {model_name} будет использовать сообщение длиной {len(final_user_message)} символов")
                        if doc_context:
                            logger.info(f"Socket.IO: Модель {model_name} получила контекст документов, длина: {len(doc_context)} символов")
                        else:
                            logger.info(f"Socket.IO: Модель {model_name} не получила контекст документов (документы отсутствуют)")
                        
                        # Определяем путь к модели
                        if model_name.startswith("llm-svc://"):
                            # Модель из llm-svc - не требует загрузки
                            model_path = model_name
                        else:
                            # Локальная модель - нужно загрузить её перед генерацией
                            model_path = os.path.join("models", model_name) if not os.path.isabs(model_name) else model_name
                            
                            # Загружаем модель для этого запроса
                            # Используем общую блокировку для предотвращения конфликтов при параллельной загрузке
                            with model_load_lock:
                                logger.info(f"Socket.IO: Загрузка модели {model_name} для генерации...")
                                if reload_model_by_path:
                                    # Перезагружаем модель с блокировкой
                                    success = reload_model_by_path(model_path)
                                    if not success:
                                        logger.error(f"Socket.IO: Не удалось загрузить модель {model_name}")
                                        return {"model": model_name, "response": f"Ошибка: Не удалось загрузить модель {model_name}", "error": True}
                                    logger.info(f"Socket.IO: Модель {model_name} успешно загружена")
                                    # Небольшая задержка после загрузки для стабилизации
                                    import time
                                    time.sleep(0.5)
                                else:
                                    logger.warning(f"Socket.IO: Функция reload_model_by_path недоступна, используем текущую модель")
                        
                        # Функция для отправки чанков от конкретной модели
                        def model_stream_callback(chunk: str, acc_text: str):
                            try:
                                # Планируем отправку чанка в event loop
                                asyncio.run_coroutine_threadsafe(
                                    sio.emit('multi_llm_chunk', {
                                        'model': model_name,
                                        'chunk': chunk,
                                        'accumulated': acc_text
                                    }, room=sid),
                                    loop
                                )
                            except Exception as e:
                                logger.error(f"Socket.IO: Ошибка отправки чанка от модели {model_name}: {e}")
                            return True
                        
                        # Для режима multi-llm используем пустую историю, чтобы избежать смешивания контекстов
                        # Это гарантирует, что каждая модель отвечает только на текущий вопрос
                        multi_llm_history = []
                        
                        # Генерируем ответ
                        response = None
                        if streaming:
                            # Потоковая генерация для каждой модели
                            import concurrent.futures
                            with concurrent.futures.ThreadPoolExecutor() as executor:
                                response = await asyncio.get_event_loop().run_in_executor(
                                    executor,
                                    ask_agent,
                                    final_user_message,  # Используем сообщение с контекстом документов
                                    multi_llm_history,  # Используем пустую историю для режима multi-llm
                                    None,  # max_tokens
                                    True,  # streaming
                                    model_stream_callback,
                                    model_path,  # model_path
                                    None   # custom_prompt_id
                                )
                        else:
                            # Обычная генерация
                            import concurrent.futures
                            with concurrent.futures.ThreadPoolExecutor() as executor:
                                response = await asyncio.get_event_loop().run_in_executor(
                                    executor,
                                    ask_agent,
                                    final_user_message,  # Используем сообщение с контекстом документов
                                    multi_llm_history,  # Используем пустую историю для режима multi-llm
                                    None,  # max_tokens
                                    False,  # streaming
                                    None,   # stream_callback
                                    model_path,  # model_path
                                    None    # custom_prompt_id
                                )
                        
                        # Проверяем, является ли ответ ошибкой
                        # ask_agent возвращает строку с ошибкой, если что-то пошло не так
                        has_error = False
                        if response:
                            error_indicators = [
                                "Извините, произошла ошибка",
                                "llama_decode returned -1",
                                "Ошибка",
                                "error",
                                "Error"
                            ]
                            has_error = any(indicator.lower() in response.lower() for indicator in error_indicators)
                        
                        if has_error:
                            logger.warning(f"Socket.IO: Модель {model_name} вернула ошибку: {response[:100]}")
                            return {"model": model_name, "response": response, "error": True}
                        else:
                            return {"model": model_name, "response": response}
                    except Exception as e:
                        logger.error(f"Socket.IO: Исключение при генерации от модели {model_name}: {e}")
                        import traceback
                        logger.error(traceback.format_exc())
                        return {"model": model_name, "response": f"Ошибка: {str(e)}", "error": True}
                
                # Запускаем параллельную генерацию от всех моделей
                # Примечание: событие multi_llm_start теперь отправляется для каждой модели отдельно внутри generate_single_model_response
                tasks = [generate_single_model_response(model) for model in multi_llm_models]
                results = await asyncio.gather(*tasks, return_exceptions=True)
                
                # Отправляем результаты для каждой модели
                for i, result in enumerate(results):
                    if isinstance(result, Exception):
                        logger.error(f"Socket.IO: Исключение при генерации: {result}")
                        await sio.emit('multi_llm_complete', {
                            'model': 'unknown',
                            'response': f'Ошибка: {str(result)}',
                            'error': True,
                            'index': i,
                            'total': len(multi_llm_models)
                        }, room=sid)
                    else:
                        await sio.emit('multi_llm_complete', {
                            'model': result.get('model', 'unknown'),
                            'response': result.get('response', ''),
                            'error': result.get('error', False),
                            'index': i,
                            'total': len(multi_llm_models)
                        }, room=sid)
                
                logger.info("Socket.IO: Все ответы от моделей сгенерированы")
                return
            
            # Если включен агентный режим, используем агентную архитектуру
            if use_agent_mode:
                logger.info("Socket.IO: АГЕНТНАЯ АРХИТЕКТУРА: Переключение на агентный режим обработки")
                logger.info(f"Socket.IO: Запрос пользователя: '{user_message[:100]}{'...' if len(user_message) > 100 else ''}'")
                
                # Используем агентную архитектуру
                context = {
                    "history": history,
                    "user_message": user_message,
                    "doc_processor": doc_processor  # Передаем doc_processor для DocumentAgent
                }
                logger.info(f"[Socket.IO] doc_processor ID в контексте: {id(doc_processor)}")
                logger.info(f"[Socket.IO] doc_processor doc_names: {doc_processor.doc_names if doc_processor else 'None'}")
                response = await orchestrator.process_message(user_message, context)
                logger.info(f"Socket.IO: АГЕНТНАЯ АРХИТЕКТУРА: Получен ответ, длина: {len(response)} символов")
                
                # Отправляем ответ через Socket.IO
                await sio.emit('chat_complete', {
                    'response': response,
                    'timestamp': datetime.now().isoformat()
                }, room=sid)
                
                # Сохраняем ответ в память
                try:
                    conversation_id = data.get("conversation_id", None)
                    await save_dialog_entry("assistant", response, None, None, conversation_id)
                except RuntimeError as e:
                    # Ошибка MongoDB - логируем, но продолжаем работу
                    error_msg = str(e)
                    if "MongoDB" in error_msg:
                        logger.warning(f"MongoDB недоступен. Ответ не будет сохранен: {e}")
                    else:
                        logger.error(f"Ошибка при сохранении ответа: {e}")
                return
            
            # Иначе используем прямой режим
            logger.info("Socket.IO: ПРЯМОЙ РЕЖИМ: Переключение на прямое общение с LLM")
            logger.info(f"Socket.IO: Запрос пользователя: '{user_message[:100]}{'...' if len(user_message) > 100 else ''}'")
            
            # =============================================
            # ЛОГИКА ОБРАБОТКИ С ДОКУМЕНТАМИ (как в WebSocket)
            # =============================================
            final_message = user_message
            

            
            # Проверяем наличие документов и используем их контекст
            images = None  # Пути к изображениям для мультимодальной модели
            if doc_processor:
                logger.info("Socket.IO: doc_processor доступен")
                doc_list = doc_processor.get_document_list()
                logger.info(f"Socket.IO: список документов: {doc_list}")
                logger.info(f"Socket.IO: количество документов: {len(doc_list) if doc_list else 0}")
                # Дополнительная диагностика
                if hasattr(doc_processor, 'doc_names'):
                    logger.info(f"Socket.IO: doc_processor.doc_names = {doc_processor.doc_names}")
                if hasattr(doc_processor, 'vectorstore'):
                    logger.info(f"Socket.IO: vectorstore доступен: {doc_processor.vectorstore is not None}")
                    if doc_processor.vectorstore:
                        logger.info(f"Socket.IO: количество документов в vectorstore: {len(doc_processor.documents) if hasattr(doc_processor, 'documents') and doc_processor.documents else 0}")
                
                # Получаем пути к изображениям для мультимодальной модели
                # get_image_paths создает временные файлы из данных в памяти только при необходимости
                image_paths = doc_processor.get_image_paths()
                if image_paths and len(image_paths) > 0:
                    # Проверяем, что временные файлы существуют (они создаются в get_image_paths)
                    available_images = []
                    for img_path in image_paths:
                        if img_path and os.path.exists(img_path):
                            available_images.append(img_path)
                    images = available_images if available_images else None
                    if images:
                        logger.info(f"Socket.IO: найдены изображения для мультимодальной модели: {len(images)} файлов")
                        # Временные файлы будут удалены после использования в LLM клиенте
                
                if doc_list and len(doc_list) > 0:
                    logger.info(f"Socket.IO: найдены документы: {doc_list}")
                    # Используем document processor для ответа с контекстом документов
                    logger.info("Socket.IO: используем document processor для ответа с контекстом документов")
                    
                    # Получаем контекст из документов
                    try:
                        doc_context = doc_processor.get_document_context(user_message)
                        logger.info(f"Socket.IO: получен контекст документов, длина: {len(doc_context) if doc_context else 0} символов")
                        
                        if doc_context and not images:
                            # Формируем промпт с контекстом документов только если нет изображений
                            # (для изображений используем мультимодальный формат)
                            final_message = f"""Документы: {doc_context}

Вопрос: {user_message}

Ответь на основе документов."""
                            
                            logger.info("Socket.IO: отправляем промпт с контекстом в AI agent")
                        else:
                            logger.warning("Socket.IO: контекст документов пуст или есть изображения, используем исходное сообщение")
                            
                    except Exception as e:
                        logger.error(f"Socket.IO: ошибка при получении контекста документов: {e}")
                        # Fallback к обычному сообщению
                        logger.info("Socket.IO: используем fallback к исходному сообщению")
                else:
                    logger.info("Socket.IO: список документов пуст, используем исходное сообщение")
            else:
                logger.info("Socket.IO: doc_processor не доступен, используем исходное сообщение")
            
            # Генерация ответа
            current_model_path = get_current_model_path()
            if streaming:
                # Потоковая генерация в отдельном потоке
                import concurrent.futures
                with concurrent.futures.ThreadPoolExecutor() as executor:
                    response = await asyncio.get_event_loop().run_in_executor(
                        executor,
                        ask_agent,
                        final_message,
                        history,
                        None,  # max_tokens
                        True,  # streaming
                        sync_stream_callback,
                        current_model_path,  # model_path
                        None,   # custom_prompt_id
                        images  # images для мультимодальной модели
                    )
                logger.info(f"Socket.IO: получен потоковый ответ, длина: {len(response)} символов")
                
                # Проверяем, не была ли генерация остановлена
                if response is None:
                    logger.info(f"Socket.IO: потоковая генерация была остановлена для {sid}")
                    return
            else:
                # Обычная генерация в отдельном потоке
                import concurrent.futures
                with concurrent.futures.ThreadPoolExecutor() as executor:
                    response = await asyncio.get_event_loop().run_in_executor(
                        executor,
                        ask_agent,
                        final_message,
                        history,
                        None,  # max_tokens
                        False,  # streaming
                        None,   # stream_callback
                        current_model_path,  # model_path
                        None,    # custom_prompt_id
                        images   # images для мультимодальной модели
                    )
                logger.info(f"Socket.IO: получен ответ, длина: {len(response)} символов")
            
            # Проверяем, не была ли запрошена остановка
            if stop_generation_flags.get(sid, False):
                logger.info(f"Socket.IO: генерация была остановлена для {sid}, не отправляем финальное сообщение")
                # Очищаем флаг остановки
                stop_generation_flags[sid] = False
                return
            
            # Сохраняем ответ
            try:
                conversation_id = data.get("conversation_id", None)
                await save_dialog_entry("assistant", response, None, None, conversation_id)
            except RuntimeError as e:
                # Ошибка MongoDB - логируем, но продолжаем работу
                error_msg = str(e)
                if "MongoDB" in error_msg:
                    logger.warning(f"MongoDB недоступен. Ответ не будет сохранен: {e}")
                else:
                    logger.error(f"Ошибка при сохранении ответа: {e}")
            
            # Очищаем флаг остановки после завершения генерации
            if sid in stop_generation_flags:
                stop_generation_flags[sid] = False
            
            # Отправляем финальное сообщение
            await sio.emit('chat_complete', {
                'response': response,
                'timestamp': datetime.now().isoformat()
            }, room=sid)
            logger.info("Socket.IO: финальное сообщение отправлено")
            
        except Exception as e:
            logger.error(f"Ошибка генерации: {e}")
            await sio.emit('chat_error', {
                'error': str(e)
            }, room=sid)
            
    except Exception as e:
        logger.error(f"Socket.IO chat error: {e}")
        try:
            await sio.emit('chat_error', {
                'error': str(e)
            }, room=sid)
        except:
            logger.error("Не удалось отправить сообщение об ошибке клиенту")

# Модели данных
from pydantic import BaseModel

class ChatMessage(BaseModel):
    message: str
    streaming: bool = True

class ModelSettings(BaseModel):
    context_size: int = 2048
    output_tokens: int = 512
    temperature: float = 0.7
    top_p: float = 0.95
    repeat_penalty: float = 1.05
    use_gpu: bool = False
    streaming: bool = True

class VoiceSettings(BaseModel):
    voice_id: str = "ru"
    speech_rate: float = 1.0
    voice_speaker: str = "baya"

class MemorySettings(BaseModel):
    max_messages: int = 20
    include_system_prompts: bool = True
    clear_on_restart: bool = False

class ModelLoadRequest(BaseModel):
    model_path: str

class ModelLoadResponse(BaseModel):
    message: str
    success: bool

# ================================
# ОСНОВНЫЕ API ENDPOINTS
# ================================

@app.get("/")
async def root():
    """Главная страница API"""
    return {"message": "MemoAI Web API", "status": "active", "version": "1.0.0"}

@app.get("/socket-test")
async def socket_test():
    """Тестовый endpoint для проверки Socket.IO"""
    return {
        "socketio_status": "active",
        "endpoint": "/socket.io/",
        "cors_origins": ["http://localhost:3000", "http://127.0.0.1:3000"],
        "ping_timeout": 120,
        "ping_interval": 25
    }

@app.get("/health")
async def health_check():
    """Проверка состояния системы"""
    try:
        model_info = get_model_info() if get_model_info else {"loaded": False}
        vosk_status = check_vosk_model() if check_vosk_model else False
        
        return {
            "status": "healthy",
            "timestamp": datetime.now().isoformat(),
            "services": {
                "llm_model": model_info.get("loaded", False),
                "vosk_model": vosk_status,
                "document_processor": DocumentProcessor is not None,
                "transcriber": UniversalTranscriber is not None
            }
        }
    except Exception as e:
        return {
            "status": "unhealthy",
            "error": str(e),
            "timestamp": datetime.now().isoformat()
        }

# ================================
# ЧАТ API
# ================================

@app.post("/api/chat")
async def chat_with_ai(message: ChatMessage):
    """Отправить сообщение AI и получить ответ"""
    if not ask_agent:
        raise HTTPException(status_code=503, detail="AI agent не доступен")
    if not save_dialog_entry:
        raise HTTPException(status_code=503, detail="Memory module не доступен")
        
    try:
        logger.info(f"Chat request: {message.message[:50]}...")
        
        # Получаем историю диалога
        if get_recent_dialog_history:
            history = await get_recent_dialog_history(max_entries=memory_max_messages)
        else:
            history = []
        
        # Проверяем, доступна ли агентная архитектура
        orchestrator = get_agent_orchestrator()
        use_agent_mode = orchestrator and orchestrator.get_mode() == "agent"
        
        logger.info(f"DEBUG: orchestrator = {orchestrator is not None}")
        if orchestrator:
            logger.info(f"DEBUG: orchestrator.get_mode() = '{orchestrator.get_mode()}'")
        logger.info(f"DEBUG: use_agent_mode = {use_agent_mode}")
        
        if use_agent_mode:
            logger.info("АГЕНТНАЯ АРХИТЕКТУРА: Переключение на агентный режим обработки")
            logger.info(f"Запрос пользователя: '{message.message[:100]}{'...' if len(message.message) > 100 else ''}'")
            # Используем агентную архитектуру
            context = {
                "history": history,
                "user_message": message.message,
                "doc_processor": doc_processor  # Передаем doc_processor для DocumentAgent
            }
            response = await orchestrator.process_message(message.message, context)
            logger.info(f"АГЕНТНАЯ АРХИТЕКТУРА: Получен ответ, длина: {len(response)} символов")
        else:
            logger.info("ПРЯМОЙ РЕЖИМ: Переключение на прямое общение с LLM")
            logger.info(f"Запрос пользователя: '{message.message[:100]}{'...' if len(message.message) > 100 else ''}'")
            # Проверяем, есть ли загруженные документы
            logger.info(f"doc_processor доступен: {doc_processor is not None}")
            if doc_processor:
                doc_list = doc_processor.get_document_list()
                logger.info(f"Список документов: {doc_list}")
                logger.info(f"Количество документов: {len(doc_list) if doc_list else 0}")
                
                if doc_list and len(doc_list) > 0:
                    logger.info(f"ПРЯМОЙ РЕЖИМ: Найдены документы: {doc_list}")
                    # Используем document processor для ответа с контекстом документов
                    logger.info("ПРЯМОЙ РЕЖИМ: Используем document processor для ответа с контекстом документов")
                    response = doc_processor.process_query(message.message, ask_agent)
                    logger.info(f"ПРЯМОЙ РЕЖИМ: Получен ответ от document processor, длина: {len(response)} символов")
                else:
                    logger.info("ПРЯМОЙ РЕЖИМ: Список документов пуст, используем обычный AI agent")
                    # Отправляем запрос к модели без контекста документов
                    current_model_path = get_current_model_path()
                    response = ask_agent(
                        message.message,
                        history=history,
                        streaming=False,  # Для REST API используем обычный режим
                        model_path=current_model_path
                    )
                    logger.info(f"ПРЯМОЙ РЕЖИМ: Получен ответ от AI agent, длина: {len(response)} символов")
            else:
                logger.info("ПРЯМОЙ РЕЖИМ: doc_processor не доступен, используем обычный AI agent")
                # Отправляем запрос к модели без контекста документов
                current_model_path = get_current_model_path()
                response = ask_agent(
                    message.message,
                    history=history,
                    streaming=False,  # Для REST API используем обычный режим
                    model_path=current_model_path
                )
                logger.info(f"ПРЯМОЙ РЕЖИМ: Получен ответ от AI agent, длина: {len(response)} символов")
        
        # Добавляем отладочную информацию в ответ для прямого режима
        # if not use_agent_mode:
        #     debug_info = f"\n\n--- AstraChatЦИЯ ---\n"
        #     debug_info += f"Режим: Прямое общение с LLM\n"
        #     debug_info += f"Модель: {get_current_model_path()}\n"
        #     debug_info += f"История диалога: {len(history)} сообщений\n"
        #     debug_info += f"--- КОНЕЦ ОТЛАДОЧНОЙ ИНФОРМАЦИИ ---"
        #     response = response + debug_info
        
        # Сохраняем в память
        await save_dialog_entry("user", message.message)
        await save_dialog_entry("assistant", response)
        
        return {
            "response": response,
            "timestamp": datetime.now().isoformat(),
            "success": True
        }
        
    except Exception as e:
        logger.error(f"Chat error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.put("/api/messages/{conversation_id}/{message_id}")
async def update_message(conversation_id: str, message_id: str, request: Dict[str, str]):
    """Обновить содержимое сообщения в MongoDB"""
    try:
        if not get_conversation_repository:
            raise HTTPException(status_code=503, detail="MongoDB repository не доступен")
        
        conversation_repo = get_conversation_repository()
        content = request.get("content", "")
        old_content = request.get("old_content", None)  # Старое содержимое для поиска
        
        if not content:
            raise HTTPException(status_code=400, detail="Поле 'content' обязательно")
        
        success = await conversation_repo.update_message(conversation_id, message_id, content, old_content)
        
        if success:
            return {
                "message": "Сообщение обновлено",
                "success": True,
                "timestamp": datetime.now().isoformat()
            }
        else:
            raise HTTPException(status_code=404, detail="Сообщение не найдено")
            
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Ошибка при обновлении сообщения: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.websocket("/ws/chat")
async def websocket_chat(websocket: WebSocket):
    """WebSocket для потокового чата с AI"""
    if not ask_agent or not save_dialog_entry:
        await websocket.close(code=1008, reason="AI services not available")
        return
        
    await manager.connect(websocket)
    try:
        while True:
            # Получаем сообщение от клиента
            data = await websocket.receive_text()
            message_data = json.loads(data)
            
            # Получаем сообщение чата
            user_message = message_data.get("message", "")
            streaming = message_data.get("streaming", True)
            
            logger.info(f"WebSocket chat: {user_message[:50]}...")
            
            # Получаем историю
            if get_recent_dialog_history:
                history = await get_recent_dialog_history(max_entries=memory_max_messages)
            else:
                history = []
            
            # Сохраняем сообщение пользователя
            await save_dialog_entry("user", user_message)
            
            # Проверяем, доступна ли агентная архитектура
            orchestrator = get_agent_orchestrator()
            use_agent_mode = orchestrator and orchestrator.get_mode() == "agent"
            use_multi_llm_mode = orchestrator and orchestrator.get_mode() == "multi-llm"
            
            logger.info(f"WebSocket DEBUG: orchestrator = {orchestrator is not None}")
            if orchestrator:
                logger.info(f"WebSocket DEBUG: orchestrator.get_mode() = '{orchestrator.get_mode()}'")
            logger.info(f"WebSocket DEBUG: use_agent_mode = {use_agent_mode}, use_multi_llm_mode = {use_multi_llm_mode}")
            
            # Функция для отправки частей ответа
            def stream_callback(chunk: str, accumulated_text: str):
                try:
                    logger.info(f"WebSocket: отправляем чанк, длина: {len(chunk)} символов, накоплено: {len(accumulated_text)} символов")
                    asyncio.create_task(websocket.send_text(json.dumps({
                        "type": "chunk",
                        "chunk": chunk,
                        "accumulated": accumulated_text
                    })))
                    logger.info("WebSocket: чанк успешно отправлен")
                    return True  # Возвращаем True для продолжения
                except Exception as e:
                    logger.error(f"WebSocket: ошибка отправки чанка: {e}")
                    return False
            
            try:
                # Если включен режим multi-llm, генерируем ответы от нескольких моделей параллельно
                if use_multi_llm_mode:
                    logger.info("WebSocket: РЕЖИМ MULTI-LLM: Параллельная генерация от нескольких моделей")
                    logger.info(f"WebSocket: Запрос пользователя: '{user_message[:100]}{'...' if len(user_message) > 100 else ''}'")
                    
                    multi_llm_models = orchestrator.get_multi_llm_models()
                    if not multi_llm_models:
                        logger.warning("WebSocket: Режим multi-llm активирован, но модели не выбраны")
                        await websocket.send_text(json.dumps({
                            "type": "error",
                            "error": "Режим multi-llm активирован, но модели не выбраны"
                        }))
                        continue
                    
                    logger.info(f"WebSocket: Генерируем ответы от моделей: {multi_llm_models}")
                    
                    # Проверяем, есть ли загруженные документы
                    doc_context = None
                    if doc_processor:
                        doc_list = doc_processor.get_document_list()
                        if doc_list and len(doc_list) > 0:
                            logger.info(f"WebSocket: Найдены документы в режиме multi-llm: {doc_list}")
                            try:
                                doc_context = doc_processor.get_document_context(user_message)
                                logger.info(f"WebSocket: Получен контекст документов для multi-llm, длина: {len(doc_context) if doc_context else 0} символов")
                            except Exception as e:
                                logger.error(f"WebSocket: Ошибка при получении контекста документов: {e}")
                    
                    # Формируем финальное сообщение с контекстом документов, если есть
                    final_user_message = user_message
                    if doc_context:
                        final_user_message = f"""Контекст из загруженных документов:
{doc_context}

Вопрос пользователя: {user_message}

Пожалуйста, ответьте на вопрос пользователя, используя информацию из предоставленных документов. Если в документах нет информации для ответа, честно скажите об этом."""
                    
                    # Функция для генерации ответа от одной модели
                    async def generate_single_model_response(model_name: str):
                        try:
                            logger.info(f"WebSocket: Начинаем генерацию от модели: {model_name}")
                            
                            # Отправляем событие начала генерации для этой модели
                            await websocket.send_text(json.dumps({
                                "type": "multi_llm_start",
                                "model": model_name,
                                "total_models": len(multi_llm_models),
                                "models": multi_llm_models
                            }))
                            
                            # Логируем сообщение с контекстом для отладки
                            logger.info(f"WebSocket: Модель {model_name} будет использовать сообщение длиной {len(final_user_message)} символов")
                            if doc_context:
                                logger.info(f"WebSocket: Модель {model_name} получила контекст документов, длина: {len(doc_context)} символов")
                            else:
                                logger.info(f"WebSocket: Модель {model_name} не получила контекст документов (документы отсутствуют)")
                            
                            # Определяем путь к модели
                            if model_name.startswith("llm-svc://"):
                                # Модель из llm-svc
                                model_path = model_name
                            else:
                                # Локальная модель
                                model_path = os.path.join("models", model_name) if not os.path.isabs(model_name) else model_name
                            
                            # Для режима multi-llm используем пустую историю, чтобы избежать смешивания контекстов
                            multi_llm_history = []
                            
                            # Генерируем ответ
                            if streaming:
                                # Потоковая генерация для каждой модели
                                accumulated_text = ""
                                def model_stream_callback(chunk: str, acc_text: str):
                                    nonlocal accumulated_text
                                    accumulated_text = acc_text
                                    try:
                                        asyncio.create_task(websocket.send_text(json.dumps({
                                            "type": "multi_llm_chunk",
                                            "model": model_name,
                                            "chunk": chunk,
                                            "accumulated": acc_text
                                        })))
                                    except Exception as e:
                                        logger.error(f"WebSocket: Ошибка отправки чанка от модели {model_name}: {e}")
                                    return True
                                
                                response = ask_agent(
                                    final_user_message,  # Используем сообщение с контекстом документов
                                    history=multi_llm_history,  # Используем пустую историю для режима multi-llm
                                    streaming=True,
                                    stream_callback=model_stream_callback,
                                    model_path=model_path
                                )
                                return {"model": model_name, "response": accumulated_text}
                            else:
                                # Обычная генерация
                                response = ask_agent(
                                    final_user_message,  # Используем сообщение с контекстом документов
                                    history=multi_llm_history,  # Используем пустую историю для режима multi-llm
                                    streaming=False,
                                    model_path=model_path
                                )
                                return {"model": model_name, "response": response}
                        except Exception as e:
                            logger.error(f"WebSocket: Исключение при генерации от модели {model_name}: {e}")
                            import traceback
                            logger.error(traceback.format_exc())
                            return {"model": model_name, "response": f"Ошибка: {str(e)}", "error": True}
                    
                    # Запускаем параллельную генерацию от всех моделей
                    import concurrent.futures
                    loop = asyncio.get_event_loop()
                    
                    # Создаем задачи для каждой модели
                    tasks = [generate_single_model_response(model) for model in multi_llm_models]
                    
                    # Запускаем все задачи параллельно
                    results = await asyncio.gather(*tasks, return_exceptions=True)
                    
                    # Отправляем результаты для каждой модели
                    for result in results:
                        if isinstance(result, Exception):
                            logger.error(f"WebSocket: Исключение при генерации: {result}")
                            await websocket.send_text(json.dumps({
                                "type": "multi_llm_complete",
                                "model": "unknown",
                                "response": f"Ошибка: {str(result)}",
                                "error": True
                            }))
                        else:
                            await websocket.send_text(json.dumps({
                                "type": "multi_llm_complete",
                                "model": result.get("model", "unknown"),
                                "response": result.get("response", ""),
                                "error": result.get("error", False)
                            }))
                    
                    logger.info("WebSocket: Все ответы от моделей сгенерированы")
                    continue
                
                # Если включен агентный режим, используем агентную архитектуру
                if use_agent_mode:
                    logger.info("WebSocket: АГЕНТНАЯ АРХИТЕКТУРА: Переключение на агентный режим обработки")
                    logger.info(f"WebSocket: Запрос пользователя: '{user_message[:100]}{'...' if len(user_message) > 100 else ''}'")
                    
                    # Используем агентную архитектуру
                    context = {
                        "history": history,
                        "user_message": user_message,
                        "doc_processor": doc_processor  # Передаем doc_processor для DocumentAgent
                    }
                    response = await orchestrator.process_message(user_message, context)
                    logger.info(f"WebSocket: АГЕНТНАЯ АРХИТЕКТУРА: Получен ответ, длина: {len(response)} символов")
                    
                    # Отправляем ответ через WebSocket
                    await websocket.send_text(json.dumps({
                        "type": "complete",
                        "response": response
                    }))
                    
                    # Сохраняем ответ в память
                    await save_dialog_entry("assistant", response)
                    continue
                
                # Иначе используем прямой режим
                logger.info("WebSocket: ПРЯМОЙ РЕЖИМ: Переключение на прямое общение с LLM")
                logger.info(f"WebSocket: Запрос пользователя: '{user_message[:100]}{'...' if len(user_message) > 100 else ''}'")
                # Проверяем, есть ли загруженные документы
                logger.info(f"WebSocket: doc_processor доступен: {doc_processor is not None}")
                if doc_processor:
                    doc_list = doc_processor.get_document_list()
                    logger.info(f"WebSocket: список документов: {doc_list}")
                    logger.info(f"WebSocket: количество документов: {len(doc_list) if doc_list else 0}")
                    logger.info(f"WebSocket: doc_list is None: {doc_list is None}")
                    logger.info(f"WebSocket: doc_list == []: {doc_list == []}")
                    logger.info(f"WebSocket: bool(doc_list): {bool(doc_list)}")
                    
                    if doc_list and len(doc_list) > 0:
                        logger.info(f"WebSocket: найдены документы: {doc_list}")
                        # Используем document processor для ответа с контекстом документов
                        logger.info("WebSocket: используем document processor для ответа с контекстом документов")
                        
                        # Получаем контекст из документов
                        try:
                            doc_context = doc_processor.get_document_context(user_message)
                            logger.info(f"WebSocket: получен контекст документов, длина: {len(doc_context) if doc_context else 0} символов")
                            
                            # Формируем промпт с контекстом документов
                            enhanced_prompt = f"""Контекст из загруженных документов:
{doc_context}

Вопрос пользователя: {user_message}

Пожалуйста, ответьте на вопрос пользователя, используя информацию из предоставленных документов. Если в документах нет информации для ответа, честно скажите об этом."""
                            
                            logger.info("WebSocket: отправляем промпт с контекстом в AI agent")
                            
                            current_model_path = get_current_model_path()
                            if streaming:
                                response = ask_agent(
                                    enhanced_prompt,
                                    history=history,
                                    streaming=True,
                                    stream_callback=stream_callback,
                                    model_path=current_model_path
                                )
                            else:
                                response = ask_agent(
                                    enhanced_prompt,
                                    history=history,
                                    streaming=False,
                                    model_path=current_model_path
                                )
                            
                            logger.info(f"WebSocket: получен ответ от AI agent с контекстом документов, длина: {len(response)} символов")
                            
                        except Exception as e:
                            logger.error(f"WebSocket: ошибка при получении контекста документов: {e}")
                            # Fallback к обычному AI agent
                            current_model_path = get_current_model_path()
                            if streaming:
                                response = ask_agent(
                                    user_message,
                                    history=history,
                                    streaming=True,
                                    stream_callback=stream_callback,
                                    model_path=current_model_path
                                )
                            else:
                                response = ask_agent(
                                    user_message,
                                    history=history,
                                    streaming=False,
                                    model_path=current_model_path
                                )
                            logger.info(f"WebSocket: использован fallback к обычному AI agent")
                    else:
                        logger.info("WebSocket: список документов пуст, используем обычный AI agent")
                        current_model_path = get_current_model_path()
                        if streaming:
                            # Потоковая генерация
                            response = ask_agent(
                                user_message,
                                history=history,
                                streaming=True,
                                stream_callback=stream_callback,
                                model_path=current_model_path
                            )
                            logger.info(f"WebSocket: получен потоковый ответ от AI agent, длина: {len(response)} символов")
                        else:
                            # Обычная генерация
                            response = ask_agent(
                                user_message,
                                history=history,
                                streaming=False,
                                model_path=current_model_path
                            )
                            logger.info(f"WebSocket: получен потоковый ответ от AI agent, длина: {len(response)} символов")
                else:
                    logger.info("WebSocket: doc_processor не доступен, используем обычный AI agent")
                    current_model_path = get_current_model_path()
                    if streaming:
                        # Потоковая генерация
                        response = ask_agent(
                            user_message,
                            history=history,
                            streaming=True,
                            stream_callback=stream_callback,
                            model_path=current_model_path
                        )
                        logger.info(f"WebSocket: получен потоковый ответ от AI agent, длина: {len(response)} символов")
                    else:
                        # Обычная генерация
                        response = ask_agent(
                            user_message,
                            history=history,
                            streaming=False,
                            model_path=current_model_path
                        )
                        logger.info(f"WebSocket: получен ответ от AI agent, длина: {len(response)} символов")
                
                # Сохраняем ответ
                await save_dialog_entry("assistant", response)
                
                # Отправляем финальное сообщение
                await websocket.send_text(json.dumps({
                    "type": "complete",
                    "response": response,
                    "timestamp": datetime.now().isoformat()
                }))
                
            except Exception as e:
                await websocket.send_text(json.dumps({
                    "type": "error",
                    "error": str(e)
                }))
                
    except WebSocketDisconnect:
        logger.info("WebSocket отключен клиентом - нормальное отключение")
        try:
            manager.disconnect(websocket)
        except Exception as e:
            logger.warning(f"Ошибка при отключении WebSocket в менеджере: {e}")
    except Exception as e:
        logger.error(f"WebSocket error: {e}")
        manager.disconnect(websocket)

async def process_audio_data(websocket: WebSocket, data: bytes):
    """Обработка аудио данных от WebSocket клиента"""
    import tempfile
    temp_dir = tempfile.gettempdir()
    audio_object_name = None
    audio_file = None
    
    # Проверяем флаг остановки голосового чата
    if globals().get('voice_chat_stop_flag', False):
        logger.info("Обработка аудио данных остановлена - установлен флаг остановки")
        return
    
    logger.info(f"Начинаю обработку аудио данных размером {len(data)} байт")
    
    try:
        # Проверяем, что получили действительно аудио данные
        if len(data) < 100:  # Слишком маленький размер для аудио
            logger.warning(f"Получены данные слишком маленького размера: {len(data)} байт")
            await websocket.send_text(json.dumps({
                "type": "error",
                "error": "Получены некорректные аудио данные"
            }))
            return
        
        # Сохраняем файл в MinIO или локально
        if minio_client:
            try:
                audio_object_name = minio_client.generate_object_name(prefix="voice_", extension=".wav")
                minio_client.upload_file(data, audio_object_name, content_type="audio/wav")
                # Получаем локальный путь для обработки (функция распознавания требует файл)
                audio_file = minio_client.get_file_path(audio_object_name)
                logger.info(f"Аудио файл загружен в MinIO: {audio_object_name}")
            except Exception as e:
                logger.warning(f"Ошибка загрузки в MinIO, используем локальный файл: {e}")
                audio_file = os.path.join(temp_dir, f"voice_{datetime.now().timestamp()}.wav")
                with open(audio_file, "wb") as f:
                    f.write(data)
        else:
            audio_file = os.path.join(temp_dir, f"voice_{datetime.now().timestamp()}.wav")
            with open(audio_file, "wb") as f:
                f.write(data)
        
        # Распознаем речь
        logger.info(f"Обрабатываю аудио файл: {audio_file}")
        
        if not recognize_speech_from_file:
            logger.warning("recognize_speech_from_file функция не доступна")
            await websocket.send_text(json.dumps({
                "type": "speech_error",
                "error": "Модуль распознавания речи недоступен. Проверьте установку Vosk."
            }))
            return
            
        recognized_text = recognize_speech_from_file(audio_file)
        logger.info(f"РАСПОЗНАННЫЙ ТЕКСТ: '{recognized_text}'")
        
        if recognized_text and recognized_text.strip():
            # Отправляем распознанный текст клиенту
            await websocket.send_text(json.dumps({
                "type": "speech_recognized",
                "text": recognized_text,
                "timestamp": datetime.now().isoformat()
            }))
            
            # Получаем ответ от AI
            if not ask_agent:
                logger.warning("ask_agent функция не доступна")
                await websocket.send_text(json.dumps({
                    "type": "speech_error", 
                    "error": "AI модуль недоступен. Проверьте загрузку модели."
                }))
                return
                
            if get_recent_dialog_history:
                history = await get_recent_dialog_history(max_entries=memory_max_messages)
            else:
                history = []
            logger.info(f"ОТПРАВЛЯЮ В LLM: текст='{recognized_text}', история={len(history)} записей")
            
            try:
                current_model_path = get_current_model_path()
                ai_response = ask_agent(recognized_text, history=history, streaming=False, model_path=current_model_path)
                logger.info(f"ОТВЕТ ОТ LLM: '{ai_response[:100]}{'...' if len(ai_response) > 100 else ''}')")
            except Exception as ai_error:
                logger.error(f"Ошибка обращения к AI: {ai_error}")
                await websocket.send_text(json.dumps({
                    "type": "speech_error",
                    "error": f"Ошибка AI модуля: {str(ai_error)}"
                }))
                return
            
            # Сохраняем в память
            await save_dialog_entry("user", recognized_text)
            await save_dialog_entry("assistant", ai_response)
            
            # Отправляем ответ AI клиенту
            await websocket.send_text(json.dumps({
                "type": "ai_response",
                "text": ai_response,
                "timestamp": datetime.now().isoformat()
            }))
            
            # Синтезируем речь
            speech_file = os.path.join(temp_dir, f"speech_{datetime.now().timestamp()}.wav")
            speech_object_name = None
            
            if not speak_text:
                logger.warning("speak_text функция не доступна")
                await websocket.send_text(json.dumps({
                    "type": "tts_error",
                    "error": "Модуль синтеза речи недоступен. Проверьте установку TTS библиотек."
                }))
                return
            
            if speak_text(ai_response, speaker='baya', voice_id='ru', save_to_file=speech_file):
                # Проверяем, что файл создался и не пустой
                if os.path.exists(speech_file) and os.path.getsize(speech_file) > 44:  # Минимальный размер WAV заголовка
                    with open(speech_file, "rb") as f:
                        audio_data = f.read()
                    
                    # Сохраняем в MinIO, если доступен
                    if minio_client:
                        try:
                            speech_object_name = minio_client.generate_object_name(prefix="speech_", extension=".wav")
                            minio_client.upload_file(audio_data, speech_object_name, content_type="audio/wav")
                            logger.debug(f"Синтезированная речь сохранена в MinIO: {speech_object_name}")
                        except Exception as e:
                            logger.warning(f"Ошибка сохранения синтезированной речи в MinIO: {e}")
                    
                    await websocket.send_bytes(audio_data)
                    
                    # Удаляем локальный файл
                    try:
                        if os.path.exists(speech_file):
                            os.remove(speech_file)
                    except Exception as e:
                        logger.warning(f"Ошибка удаления локального файла речи: {e}")
                else:
                    # Файл не создался или поврежден
                    await websocket.send_text(json.dumps({
                        "type": "tts_error",
                        "error": "Не удалось создать аудиофайл"
                    }))
                    if os.path.exists(speech_file):
                        os.remove(speech_file)
            else:
                # Синтез не удался
                await websocket.send_text(json.dumps({
                    "type": "tts_error",
                    "error": "Ошибка синтеза речи"
                }))
        else:
            logger.warning("Речь не распознана или пустой текст")
            await websocket.send_text(json.dumps({
                "type": "speech_error",
                "error": "Речь не распознана, попробуйте еще раз"
            }))
            
    except Exception as e:
        logger.error(f"Ошибка обработки аудио данных: {e}")
        logger.error(f"Тип ошибки: {type(e).__name__}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
    finally:
        # Очистка временных файлов
        try:
            if audio_file and os.path.exists(audio_file):
                os.remove(audio_file)
            # Удаляем из MinIO, если был загружен
            if minio_client and audio_object_name:
                try:
                    minio_client.delete_file(audio_object_name)
                except Exception as e:
                    logger.warning(f"Ошибка удаления файла из MinIO: {e}")
        except Exception as e:
            logger.warning(f"Ошибка очистки временных файлов: {e}")

@app.websocket("/ws/voice")
async def websocket_voice(websocket: WebSocket):
    """WebSocket для голосового чата в реальном времени"""
    
    # Принимаем соединение в любом случае
    await manager.connect(websocket)
    
    # Проверяем доступность сервисов после подключения
    if not ask_agent or not save_dialog_entry:
        logger.warning("AI services недоступны для WebSocket /ws/voice")
        try:
            await websocket.send_text(json.dumps({
                "type": "error",
                "error": "AI сервисы недоступны. Проверьте настройки модели."
            }))
        except Exception as e:
            logger.warning(f"Не удалось отправить сообщение об ошибке: {e}")
        # Не закрываем соединение, просто отправляем ошибку
        
    try:
        while True:
            # Получаем сообщение (может быть JSON команда или аудио байты)
            try:
                # Пытаемся получить текстовое сообщение сначала
                message = await websocket.receive_text()
                logger.info(f"Получено текстовое сообщение: {message[:100]}...")  # Логируем первые 100 символов
                
                try:
                    data = json.loads(message)
                    logger.debug(f"Распарсенные данные: {data}")
                    
                    if data.get("type") == "start_listening":
                        # Команда начать прослушивание
                        logger.info("Получена команда start_listening")
                        await websocket.send_text(json.dumps({
                            "type": "listening_started",
                            "message": "Готов к приему голоса"
                        }))
                        continue
                    elif data.get("type") == "stop_processing":
                        # Команда остановить обработку (новое)
                        logger.info("Получена команда stop_processing")
                        # Используем globals() для доступа к глобальной переменной
                        globals()['voice_chat_stop_flag'] = True
                        await websocket.send_text(json.dumps({
                            "type": "processing_stopped",
                            "message": "Обработка остановлена"
                        }))
                        logger.info("Флаг остановки голосового чата установлен")
                        continue
                    elif data.get("type") == "reset_processing":
                        # Команда сбросить флаг остановки
                        logger.info("Получена команда reset_processing")
                        # Используем globals() для доступа к глобальной переменной
                        globals()['voice_chat_stop_flag'] = False
                        await websocket.send_text(json.dumps({
                            "type": "processing_reset",
                            "message": "Обработка возобновлена"
                        }))
                        logger.info("Флаг остановки голосового чата сброшен")
                        continue
                    else:
                        logger.warning(f"Неизвестный тип сообщения: {data.get('type', 'unknown')}")
                        logger.debug(f"Полные данные неизвестного сообщения: {data}")
                        continue
                        
                except json.JSONDecodeError as e:
                    logger.error(f"Ошибка парсинга JSON: {e}")
                    logger.error(f"Проблемное сообщение: {message}")
                    continue
                    
            except UnicodeDecodeError:
                # Если не можем декодировать как текст, пробуем получить как байты
                try:
                    data = await websocket.receive_bytes()
                    logger.info(f"Получены аудио данные размером: {len(data)} байт")
                    
                    # Обрабатываем аудио данные с дополнительной защитой
                    try:
                        await process_audio_data(websocket, data)
                    except Exception as process_error:
                        logger.error(f"Ошибка обработки аудио данных: {process_error}")
                        logger.error(f"Тип ошибки: {type(process_error).__name__}")
                        import traceback
                        logger.error(f"Traceback: {traceback.format_exc()}")
                        
                        # Отправляем ошибку клиенту, но не закрываем соединение
                        try:
                            await websocket.send_text(json.dumps({
                                "type": "error",
                                "error": f"Ошибка обработки аудио: {str(process_error)}"
                            }))
                        except Exception as send_error:
                            logger.error(f"Не удалось отправить сообщение об ошибке: {send_error}")
                        
                        # Продолжаем работу WebSocket
                        continue
                    
                except Exception as e:
                    logger.error(f"Ошибка получения аудио данных: {e}")
                    logger.error(f"Тип ошибки: {type(e).__name__}")
                    continue
                
                # Убираем дублирующийся код - теперь обрабатываем через process_audio_data
                continue
                    
    except WebSocketDisconnect:
        logger.info("WebSocket отключен клиентом - нормальное отключение")
        try:
            manager.disconnect(websocket)
        except Exception as e:
            logger.warning(f"Ошибка при отключении WebSocket в менеджере: {e}")
    except Exception as e:
        logger.error(f"Voice WebSocket error: {e}")
        logger.error(f"Тип ошибки: {type(e).__name__}")
        import traceback
        logger.error(f"Full traceback: {traceback.format_exc()}")
        
        # Не закрываем соединение при ошибках, только логируем их
        # Это позволит WebSocket оставаться активным
        try:
            await websocket.send_text(json.dumps({
                "type": "error",
                "error": f"Временная ошибка: {str(e)}"
            }))
        except Exception as send_error:
            logger.error(f"Не удалось отправить сообщение об ошибке: {send_error}")
            # Не закрываем соединение даже при ошибке отправки
        # Убираем finally блок, который закрывал соединение

# ================================
# ИСТОРИЯ ДИАЛОГОВ
# ================================

@app.get("/api/history")
async def get_chat_history(limit: int = None):
    """Получить историю диалогов"""
    # Если лимит не указан, используем настройку памяти
    if limit is None:
        limit = memory_max_messages if 'memory_max_messages' in globals() else 20
    """Получить историю диалогов"""
    if not get_recent_dialog_history:
        # Попытка прямого чтения файла если модуль memory недоступен
        try:
            import json
            import os
            from backend.config.config import MEMORY_PATH
            
            dialog_file = os.path.join(MEMORY_PATH, "dialog_history_dialog.json")
            
            if os.path.exists(dialog_file):
                with open(dialog_file, "r", encoding="utf-8") as f:
                    history = json.load(f)
                    # Ограничиваем количество записей настройкой памяти
                    max_entries = memory_max_messages if 'memory_max_messages' in globals() else 20
                    limited_history = history[-max_entries:] if len(history) > max_entries else history
                    logger.info(f"Загружено {len(limited_history)} записей истории из файла (модуль memory недоступен, лимит: {max_entries})")
                    return {
                        "history": limited_history,
                        "count": len(limited_history),
                        "max_messages": max_entries,
                        "timestamp": datetime.now().isoformat(),
                        "source": "file_fallback"
                    }
            else:
                logger.warning(f"Файл истории не найден: {dialog_file}")
                return {
                    "history": [],
                    "count": 0,
                    "max_messages": memory_max_messages if 'memory_max_messages' in globals() else 20,
                    "timestamp": datetime.now().isoformat(),
                    "source": "file_fallback",
                    "message": "Файл истории не найден"
                }
        except Exception as e:
            logger.error(f"Ошибка чтения истории из файла: {e}")
            return {
                "history": [],
                "count": 0,
                "max_messages": memory_max_messages if 'memory_max_messages' in globals() else 20,
                "timestamp": datetime.now().isoformat(),
                "source": "fallback_error",
                "error": str(e)
            }
    
    try:
        history = await get_recent_dialog_history(max_entries=limit)
        logger.info(f"Загружено {len(history)} записей истории через модуль memory")
        return {
            "history": history,
            "count": len(history),
            "max_messages": memory_max_messages,
            "timestamp": datetime.now().isoformat(),
            "source": "memory_module"
        }
    except Exception as e:
        logger.error(f"Ошибка получения истории через модуль memory: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/api/history")
async def clear_chat_history():
    """Очистить историю диалогов"""
    if not clear_dialog_history:
        # Попытка прямого удаления файлов если модуль memory недоступен
        try:
            import os
            from backend.config.config import MEMORY_PATH
            
            dialog_file = os.path.join(MEMORY_PATH, "dialog_history_dialog.json")
            memory_file = os.path.join(MEMORY_PATH, "dialog_history.txt")
            
            files_removed = []
            if os.path.exists(dialog_file):
                os.remove(dialog_file)
                files_removed.append("dialog_history_dialog.json")
            if os.path.exists(memory_file):
                os.remove(memory_file)
                files_removed.append("dialog_history.txt")
            
            logger.info(f"Удалены файлы истории: {files_removed} (модуль memory недоступен)")
            return {
                "message": f"История очищена (удалено файлов: {len(files_removed)})",
                "success": True,
                "files_removed": files_removed,
                "source": "file_fallback"
            }
        except Exception as e:
            logger.error(f"Ошибка удаления файлов истории: {e}")
            raise HTTPException(status_code=500, detail=f"Ошибка очистки истории: {str(e)}")
    
    try:
        result = await clear_dialog_history()
        logger.info(f"История очищена через модуль memory: {result}")
        return {
            "message": "История очищена", 
            "success": True,
            "source": "memory_module"
        }
    except Exception as e:
        logger.error(f"Ошибка очистки истории через модуль memory: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# ================================
# УПРАВЛЕНИЕ МОДЕЛЯМИ
# ================================

@app.get("/api/models/current")
async def get_current_model():
    """Получить информацию о текущей модели"""
    # Пытаемся получить информацию от модуля AI
    if get_model_info:
        try:
            result = get_model_info()
            logger.info(f"Информация о текущей модели от AI модуля: {result}")
            
            # Сохраняем информацию о текущей модели
            if result and 'path' in result:
                save_app_settings({
                    'current_model_path': result['path'],
                    'current_model_name': result.get('name', 'Unknown'),
                    'current_model_status': result.get('status', 'loaded')
                })
            
            return result
        except Exception as e:
            logger.error(f"Ошибка получения информации о модели от AI модуля: {e}")
    
    # Если AI модуль недоступен, проверяем сохраненные настройки
    try:
        settings = load_app_settings()
        current_model_path = settings.get('current_model_path')
        
        if current_model_path and os.path.exists(current_model_path):
            file_size = os.path.getsize(current_model_path)
            return {
                "name": settings.get('current_model_name', os.path.basename(current_model_path)),
                "path": current_model_path,
                "status": "loaded_from_settings",
                "size": file_size,
                "size_mb": round(file_size / (1024 * 1024), 2),
                "type": "gguf"
            }
    except Exception as e:
        logger.error(f"Ошибка проверки сохраненных настроек модели: {e}")
    
    # Если ничего не найдено, возвращаем заглушку
    logger.warning("get_model_info функция не доступна и нет сохраненной модели")
    return {
        "name": "Модель не загружена",
        "path": "",
        "status": "not_loaded",
        "size": 0,
        "type": "unknown"
    }

@app.get("/api/models")
async def get_models():
    """Получить список доступных моделей (алиас для /api/models/available)"""
    return await get_available_models()

@app.get("/api/models/available")
async def get_available_models():
    """Получить список доступных моделей"""
    try:
        # Проверяем, используется ли llm-svc
        use_llm_svc = os.getenv('USE_LLM_SVC', 'false').lower() == 'true'
        
        if use_llm_svc:
            # Получаем модели через llm-svc
            logger.info("Запрос списка моделей через llm-svc")
            try:
                from backend.llm_client import get_llm_service
                service = await get_llm_service()
                models_data = await service.client.get_models()
                
                # Преобразуем формат ответа llm-svc в наш формат
                models = []
                for model_data in models_data:
                    models.append({
                        "name": model_data.get("id", "Unknown"),
                        "path": f"llm-svc://{model_data.get('id', 'unknown')}",
                        "size": 0,  # llm-svc не предоставляет размер
                        "size_mb": 0,
                        "object": model_data.get("object", "model"),
                        "owned_by": model_data.get("owned_by", "llm-svc")
                    })
                
                logger.info(f"Получено моделей через llm-svc: {len(models)}")
                return {"models": models}
            except Exception as e:
                logger.error(f"Ошибка получения моделей через llm-svc: {e}")
                # Fallback к пустому списку
                return {"models": [], "error": str(e)}
        else:
            # Используем локальные .gguf модели
            logger.info("Запрос списка локальных .gguf моделей")
            models_dir = "models"
            if not os.path.exists(models_dir):
                return {"models": []}
            
            models = []
            for file in os.listdir(models_dir):
                if file.endswith('.gguf'):
                    file_path = os.path.join(models_dir, file)
                    size = os.path.getsize(file_path)
                    models.append({
                        "name": file,
                        "path": file_path,
                        "size": size,
                        "size_mb": round(size / (1024 * 1024), 2)
                    })
            
            return {"models": models}
    except Exception as e:
        logger.error(f"Ошибка получения списка моделей: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/models/load")
async def load_model(request: ModelLoadRequest):
    """Загрузить модель по указанному пути"""
    if not reload_model_by_path:
        logger.warning("reload_model_by_path функция не доступна")
        return ModelLoadResponse(
            message="Функция загрузки модели недоступна. Проверьте инициализацию AI agent.", 
            success=False
        )
    
    try:
        logger.info(f"Загружаю модель: {request.model_path}")
        success = reload_model_by_path(request.model_path)
        if success:
            logger.info(f"Модель успешно загружена: {request.model_path}")
            
            # Сохраняем информацию о загруженной модели
            model_name = os.path.basename(request.model_path)
            save_app_settings({
                'current_model_path': request.model_path,
                'current_model_name': model_name,
                'current_model_status': 'loaded'
            })
            
            return ModelLoadResponse(message="Модель успешно загружена", success=True)
        else:
            logger.error(f"Не удалось загрузить модель: {request.model_path}")
            return ModelLoadResponse(message="Не удалось загрузить модель", success=False)
    except Exception as e:
        logger.error(f"Ошибка загрузки модели: {e}")
        return ModelLoadResponse(message=f"Ошибка загрузки модели: {str(e)}", success=False)

@app.get("/api/models/settings")
async def get_model_settings():
    """Получить настройки модели"""
    if not model_settings:
        logger.warning("model_settings не доступен, возвращаю дефолтные настройки")
        # Возвращаем дефолтные настройки вместо 503 ошибки
        return {
            "context_size": 2048,
            "output_tokens": 512,
            "temperature": 0.7,
            "top_p": 0.95,
            "repeat_penalty": 1.05,
            "use_gpu": False,
            "streaming": True,
            "streaming_speed": 50
        }
    try:
        result = model_settings.get_all()
        logger.info(f"Настройки модели: {result}")
        return result
    except Exception as e:
        logger.error(f"Ошибка получения настроек модели: {e}")
        # Возвращаем дефолтные настройки в случае ошибки
        return {
            "context_size": 2048,
            "output_tokens": 512,
            "temperature": 0.7,
            "top_p": 0.95,
            "repeat_penalty": 1.05,
            "use_gpu": False,
            "streaming": True,
            "streaming_speed": 50
        }

@app.put("/api/models/settings")
async def update_model_settings_api(settings: ModelSettings):
    """Обновить настройки модели"""
    if not update_model_settings:
        raise HTTPException(status_code=503, detail="AI agent не доступен")
    try:
        success = update_model_settings(settings.dict())
        if success:
            return {"message": "Настройки обновлены", "success": True}
        else:
            raise HTTPException(status_code=400, detail="Не удалось обновить настройки")
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/models/settings/reset")
async def reset_model_settings():
    """Сбросить настройки модели к рекомендуемым значениям по умолчанию"""
    if not model_settings:
        raise HTTPException(status_code=503, detail="AI agent не доступен")
    try:
        model_settings.reset_to_defaults()
        return {
            "message": "Настройки сброшены к рекомендуемым значениям", 
            "success": True,
            "settings": model_settings.get_all()
        }
    except Exception as e:
        logger.error(f"Ошибка сброса настроек модели: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/models/settings/recommended")
async def get_recommended_settings():
    """Получить рекомендуемые настройки модели"""
    if not model_settings:
        raise HTTPException(status_code=503, detail="AI agent не доступен")
    try:
        return {
            "recommended": model_settings.get_recommended_settings(),
            "max_values": model_settings.get_max_values()
        }
    except Exception as e:
        logger.error(f"Ошибка получения рекомендуемых настроек: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# ================================
# ГОЛОСОВЫЕ ФУНКЦИИ
# ================================

class VoiceSynthesizeRequest(BaseModel):
    text: str
    voice_id: str = "ru"
    voice_speaker: str = "baya"
    speech_rate: float = 1.0

class TranscriptionSettings(BaseModel):
    engine: str = "whisperx"  # whisperx или vosk
    language: str = "ru"
    auto_detect: bool = True

class YouTubeTranscribeRequest(BaseModel):
    url: str

class DocumentQueryRequest(BaseModel):
    query: str

@app.post("/api/voice/synthesize")
async def synthesize_speech(request: VoiceSynthesizeRequest):
    """Синтезировать речь из текста"""
    if not speak_text:
        logger.warning("speak_text функция не доступна")
        raise HTTPException(status_code=503, detail="Модуль синтеза речи недоступен. Проверьте установку библиотек для TTS (pyttsx3, sounddevice, torch).")
    
    import tempfile
    temp_dir = tempfile.gettempdir()
    audio_file = os.path.join(temp_dir, f"speech_{datetime.now().timestamp()}.wav")
    audio_object_name = None
    
    try:
        # Логируем отладочную информацию
        logger.info(f"Синтезирую речь: '{request.text[:100]}{'...' if len(request.text) > 100 else ''}'")
        logger.info(f"Параметры: voice_id={request.voice_id}, voice_speaker={request.voice_speaker}, speech_rate={request.speech_rate}")
        
        # Синтезируем речь с правильными параметрами
        success = speak_text(
            text=request.text, 
            speaker=request.voice_speaker, 
            voice_id=request.voice_id, 
            speech_rate=request.speech_rate,
            save_to_file=audio_file
        )
        
        if success and os.path.exists(audio_file):
            logger.info(f"Аудиофайл создан: {audio_file}, размер: {os.path.getsize(audio_file)} байт")
            
            # Сохраняем в MinIO, если доступен
            if minio_client:
                try:
                    with open(audio_file, "rb") as f:
                        audio_data = f.read()
                    audio_object_name = minio_client.generate_object_name(prefix="speech_", extension=".wav")
                    minio_client.upload_file(audio_data, audio_object_name, content_type="audio/wav")
                    logger.debug(f"Синтезированная речь сохранена в MinIO: {audio_object_name}")
                except Exception as e:
                    logger.warning(f"Ошибка сохранения в MinIO: {e}")
            
            # Создаем временную копию для возврата, оригинал удалится автоматически
            temp_copy = os.path.join(temp_dir, f"speech_copy_{datetime.now().timestamp()}.wav")
            import shutil
            shutil.copy2(audio_file, temp_copy)
            
            # Возвращаем копию, которая удалится после отправки
            async def cleanup_temp_file():
                try:
                    if os.path.exists(temp_copy):
                        os.remove(temp_copy)
                        logger.info(f"Временный файл удален: {temp_copy}")
                except Exception as e:
                    logger.error(f"Ошибка при удалении временного файла: {e}")
            
            return FileResponse(
                temp_copy,
                media_type="audio/wav",
                filename="speech.wav",
                background=cleanup_temp_file
            )
        else:
            logger.error(f"Не удалось создать аудиофайл: success={success}, exists={os.path.exists(audio_file)}")
            raise HTTPException(status_code=500, detail="Не удалось создать аудиофайл")
            
    except Exception as e:
        logger.error(f"Ошибка синтеза речи: {e}")
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        # Всегда очищаем временные файлы
        try:
            if os.path.exists(audio_file):
                os.remove(audio_file)
                logger.info(f"Оригинальный временный файл удален: {audio_file}")
        except Exception as e:
            logger.error(f"Ошибка при удалении оригинального временного файла: {e}")

@app.post("/api/voice/recognize")
async def recognize_speech_api(audio_file: UploadFile = File(...)):
    """Распознать речь из аудиофайла"""
    if not recognize_speech_from_file:
        logger.warning("recognize_speech_from_file функция не доступна")
        return {
            "text": "",
            "success": False,
            "error": "Модуль распознавания речи недоступен. Проверьте настройки Vosk.",
            "timestamp": datetime.now().isoformat()
        }
    
    import tempfile
    temp_dir = tempfile.gettempdir()
    file_path = None
    file_object_name = None
    
    try:
        # Сохраняем загруженный файл
        content = await audio_file.read()
        logger.info(f"Получен аудиофайл: {audio_file.filename}, размер: {len(content)} байт")
        
        # Сохраняем в MinIO или локально
        if minio_client:
            try:
                logger.info("MinIO клиент доступен, загружаю файл в MinIO...")
                file_object_name = minio_client.generate_object_name(prefix="audio_", extension=".wav")
                logger.debug(f"Сгенерировано имя объекта: {file_object_name}")
                minio_client.upload_file(content, file_object_name, content_type="audio/wav")
                # Получаем локальный путь для обработки
                file_path = minio_client.get_file_path(file_object_name)
                logger.info(f"✅ Аудиофайл загружен в MinIO: {file_object_name}")
            except Exception as e:
                logger.warning(f"⚠️ Ошибка загрузки в MinIO, используем локальный файл: {e}")
                import traceback
                logger.debug(f"Traceback: {traceback.format_exc()}")
                file_path = os.path.join(temp_dir, f"audio_{datetime.now().timestamp()}.wav")
                with open(file_path, "wb") as f:
                    f.write(content)
        else:
            logger.warning("⚠️ MinIO клиент недоступен (minio_client is None), используем локальное хранение")
            logger.info("Проверьте:")
            logger.info("  1. Запущен ли MinIO: docker-compose ps minio (или локально)")
            logger.info("  2. Правильно ли настроен .env файл (MINIO_ENDPOINT, MINIO_PORT и т.д.)")
            logger.info("  3. Установлена ли библиотека: pip install minio")
            file_path = os.path.join(temp_dir, f"audio_{datetime.now().timestamp()}.wav")
            with open(file_path, "wb") as f:
                f.write(content)
        
        logger.info(f"Аудиофайл сохранен: {file_path}")
        
        # Распознаем речь используя правильную функцию
        text = recognize_speech_from_file(file_path)
        
        # Логируем результат распознавания
        logger.info(f"Распознанный текст: '{text}'")
        
        return {
            "text": text,
            "success": True,
            "timestamp": datetime.now().isoformat()
        }
        
    except Exception as e:
        logger.error(f"Ошибка распознавания речи: {e}")
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        # Всегда удаляем временный файл
        try:
            if file_path and os.path.exists(file_path):
                os.remove(file_path)
                logger.info(f"Временный файл удален: {file_path}")
            # Удаляем из MinIO, если был загружен
            if minio_client and file_object_name:
                try:
                    minio_client.delete_file(file_object_name)
                except Exception as e:
                    logger.warning(f"Ошибка удаления файла из MinIO: {e}")
        except Exception as e:
            logger.warning(f"Ошибка очистки временных файлов: {e}")

@app.get("/api/voice/settings")
async def get_voice_settings():
    """Получить настройки голоса"""
    # Возвращаем дефолтные настройки, можно расширить для сохранения в файл
    return {
        "voice_id": "ru",
        "speech_rate": 1.0,
        "voice_speaker": "baya"
    }

@app.put("/api/voice/settings")
async def update_voice_settings(settings: VoiceSettings):
    """Обновить настройки голоса"""
    # В реальной реализации можно сохранять настройки в файл
    return {
        "message": "Настройки голоса обновлены",
        "success": True,
        "settings": settings.dict()
    }

@app.get("/api/transcription/settings")
async def get_transcription_settings():
    """Получить настройки транскрибации"""
    global current_transcription_engine, current_transcription_language
    return {
        "engine": current_transcription_engine,
        "language": current_transcription_language,
        "auto_detect": True
    }

@app.put("/api/transcription/settings")
async def update_transcription_settings(settings: TranscriptionSettings):
    """Обновить настройки транскрибации"""
    global current_transcription_engine, current_transcription_language, transcriber
    
    try:
        # Обновляем глобальные настройки
        if settings.engine:
            current_transcription_engine = settings.engine.lower()
            logger.info(f"Переключение движка транскрибации на: {current_transcription_engine}")
            
            # Переключаем движок в UniversalTranscriber
            if transcriber and hasattr(transcriber, 'switch_engine'):
                success = transcriber.switch_engine(current_transcription_engine)
                if success:
                    logger.info(f"Движок успешно переключен на {current_transcription_engine}")
                else:
                    logger.error(f"Ошибка переключения движка на {current_transcription_engine}")
                    # Возвращаем ошибку если переключение не удалось
                    raise HTTPException(status_code=400, detail=f"Не удалось переключить движок на {current_transcription_engine}")
            else:
                logger.warning("Transcriber не поддерживает переключение движков")
        
        if settings.language:
            current_transcription_language = settings.language
            logger.info(f"Язык транскрибации изменен на: {current_transcription_language}")
            
            # Устанавливаем язык в текущем транскрайбере
            if transcriber and hasattr(transcriber, 'set_language'):
                transcriber.set_language(current_transcription_language)
        
        # Сохраняем настройки транскрибации в файл
        save_app_settings({
            'transcription_engine': current_transcription_engine,
            'transcription_language': current_transcription_language
        })
        
        return {
            "message": "Настройки транскрибации обновлены",
            "success": True,
            "settings": {
                "engine": current_transcription_engine,
                "language": current_transcription_language,
                "auto_detect": settings.auto_detect if hasattr(settings, 'auto_detect') else True
            }
        }
        
    except Exception as e:
        logger.error(f"Ошибка обновления настроек транскрибации: {e}")
        raise HTTPException(status_code=500, detail=f"Ошибка обновления настроек: {str(e)}")

# ================================
# НАСТРОЙКИ ПАМЯТИ
# ================================

@app.get("/api/memory/settings")
async def get_memory_settings():
    """Получить настройки памяти"""
    global memory_max_messages, memory_include_system_prompts, memory_clear_on_restart
    
    return {
        "max_messages": memory_max_messages,
        "include_system_prompts": memory_include_system_prompts,
        "clear_on_restart": memory_clear_on_restart
    }

@app.put("/api/memory/settings")
async def update_memory_settings(settings: MemorySettings):
    """Обновить настройки памяти"""
    global memory_max_messages, memory_include_system_prompts, memory_clear_on_restart
    
    try:
        # Обновляем глобальные настройки
        memory_max_messages = settings.max_messages
        memory_include_system_prompts = settings.include_system_prompts
        memory_clear_on_restart = settings.clear_on_restart
        
        logger.info(f"Настройки памяти обновлены: max_messages={memory_max_messages}, include_system_prompts={memory_include_system_prompts}, clear_on_restart={memory_clear_on_restart}")
        
        # Сохраняем настройки в файл
        save_app_settings({
            'memory_max_messages': memory_max_messages,
            'memory_include_system_prompts': memory_include_system_prompts,
            'memory_clear_on_restart': memory_clear_on_restart
        })
        
        return {
            "message": "Настройки памяти обновлены",
            "success": True,
            "settings": {
                "max_messages": memory_max_messages,
                "include_system_prompts": memory_include_system_prompts,
                "clear_on_restart": memory_clear_on_restart
            }
        }
        
    except Exception as e:
        logger.error(f"Ошибка обновления настроек памяти: {e}")
        raise HTTPException(status_code=500, detail=f"Ошибка обновления настроек памяти: {str(e)}")

@app.get("/api/memory/status")
async def get_memory_status():
    """Получить статус памяти"""
    try:
        if not get_recent_dialog_history:
            raise HTTPException(status_code=503, detail="Memory module не доступен")
        
        # Получаем текущую историю
        history = await get_recent_dialog_history(max_entries=memory_max_messages)
        
        return {
            "message_count": len(history),
            "max_messages": memory_max_messages,
            "include_system_prompts": memory_include_system_prompts,
            "clear_on_restart": memory_clear_on_restart,
            "success": True
        }
        
    except Exception as e:
        logger.error(f"Ошибка получения статуса памяти: {e}")
        raise HTTPException(status_code=500, detail=f"Ошибка получения статуса памяти: {str(e)}")

@app.post("/api/memory/clear")
async def clear_memory():
    """Очистить память"""
    try:
        if not clear_dialog_history:
            raise HTTPException(status_code=503, detail="Memory module не доступен")
        
        result = await clear_dialog_history()
        logger.info(f"Память очищена: {result}")
        
        return {
            "message": "Память успешно очищена",
            "success": True,
            "result": result
        }
        
    except Exception as e:
        logger.error(f"Ошибка очистки памяти: {e}")
        raise HTTPException(status_code=500, detail=f"Ошибка очистки памяти: {str(e)}")

# ================================
# РАБОТА С ДОКУМЕНТАМИ
# ================================

@app.post("/api/documents/upload")
async def upload_document(file: UploadFile = File(...)):
    """Загрузить и обработать документ"""
    logger.info(f"=== Загрузка документа: {file.filename} ===")
    
    if not doc_processor:
        logger.error("Document processor не доступен")
        raise HTTPException(status_code=503, detail="Document processor не доступен")
    
    file_object_name = None
    documents_bucket = os.getenv('MINIO_DOCUMENTS_BUCKET_NAME', 'memoai-documents')
        
    try:
        # Читаем содержимое файла в память
        content = await file.read()
        logger.info(f"Файл получен, размер: {len(content)} байт")
        
        # Определяем тип файла
        file_extension = os.path.splitext(file.filename)[1].lower() if file.filename else ""
        is_image = file_extension in ['.jpg', '.jpeg', '.png', '.webp']
        
        # Определяем content_type
        content_type_map = {
            '.pdf': 'application/pdf',
            '.docx': 'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
            '.doc': 'application/msword',
            '.txt': 'text/plain',
            '.xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            '.xls': 'application/vnd.ms-excel',
            '.jpg': 'image/jpeg',
            '.jpeg': 'image/jpeg',
            '.png': 'image/png',
            '.webp': 'image/webp'
        }
        content_type = content_type_map.get(file_extension, 'application/octet-stream')
        
        # Сохраняем в MinIO или локально
        if minio_client:
            try:
                # Генерируем имя объекта
                file_object_name = minio_client.generate_object_name(
                    prefix="doc_" if not is_image else "img_",
                    extension=file_extension
                )
                
                # Загружаем в MinIO в bucket для документов
                minio_client.upload_file(
                    content, 
                    file_object_name, 
                    content_type=content_type,
                    bucket_name=documents_bucket
                )
                logger.info(f"Документ загружен в MinIO: {documents_bucket}/{file_object_name}")
            except Exception as e:
                logger.warning(f"Ошибка загрузки в MinIO: {e}")
                # Продолжаем обработку даже если не удалось загрузить в MinIO
                file_object_name = None
        
        # Обрабатываем документ напрямую из памяти (bytes)
        logger.info("Начинаем обработку документа из памяти...")
        success, message = doc_processor.process_document(
            file_data=content,
            filename=file.filename or file_object_name or "unknown",
            file_extension=file_extension,
            minio_object_name=file_object_name,
            minio_bucket=documents_bucket if minio_client and file_object_name else None
        )
        logger.info(f"Результат обработки: success={success}, message={message}")
        
        if success:
            # Получаем список документов после обработки
            doc_list = doc_processor.get_document_list()
            logger.info(f"Список документов после обработки: {doc_list}")
            logger.info(f"Количество документов: {len(doc_list) if doc_list else 0}")
            
            # Проверяем состояние vectorstore
            if hasattr(doc_processor, 'vectorstore'):
                logger.info(f"Vectorstore доступен: {doc_processor.vectorstore is not None}")
                if hasattr(doc_processor, 'documents'):
                    logger.info(f"Количество документов в коллекции: {len(doc_processor.documents) if doc_processor.documents else 0}")
            
            # Файлы обрабатываются из памяти, локальные файлы не создаются
            logger.info("Обработка завершена, файл хранится в MinIO")
            
            # Возвращаем информацию о файле
            result = {
                "message": "Документ успешно загружен и обработан",
                "filename": file.filename,
                "success": True
            }
            
            # Для изображений возвращаем информацию о MinIO объекте
            if is_image and minio_client and file_object_name:
                result["minio_object"] = file_object_name
                result["minio_bucket"] = documents_bucket
            
            return result
        else:
            # В случае ошибки удаляем файл из MinIO, если он был загружен
            if minio_client and file_object_name:
                try:
                    minio_client.delete_file(file_object_name, bucket_name=documents_bucket)
                    logger.info(f"Файл удален из MinIO после ошибки: {documents_bucket}/{file_object_name}")
                except Exception as e:
                    logger.warning(f"Ошибка удаления файла из MinIO: {e}")
            
            raise HTTPException(status_code=400, detail=message)
            
    except Exception as e:
        logger.error(f"Ошибка при загрузке документа: {str(e)}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/documents/query")
async def query_document(request: DocumentQueryRequest):
    """Задать вопрос по загруженному документу"""
    logger.info(f"=== Запрос к документам: {request.query[:50]}... ===")
    
    if not doc_processor:
        logger.error("Document processor не доступен")
        raise HTTPException(status_code=503, detail="Document processor не доступен")
        
    try:
        if not ask_agent:
            logger.error("AI agent не доступен")
            raise HTTPException(status_code=503, detail="AI agent не доступен")
        
        # Получаем список документов
        doc_list = doc_processor.get_document_list()
        logger.info(f"Доступные документы: {doc_list}")
        logger.info(f"Количество документов: {len(doc_list) if doc_list else 0}")
        
        # Проверяем состояние vectorstore
        if hasattr(doc_processor, 'vectorstore'):
            logger.info(f"Vectorstore доступен: {doc_processor.vectorstore is not None}")
            if hasattr(doc_processor, 'documents'):
                logger.info(f"Количество документов в коллекции: {len(doc_processor.documents) if doc_processor.documents else 0}")
        
        response = doc_processor.process_query(request.query, ask_agent)
        logger.info(f"Получен ответ от document processor, длина: {len(response)} символов")
        
        return {
            "response": response,
            "query": request.query,
            "success": True,
            "timestamp": datetime.now().isoformat()
        }
    except Exception as e:
        logger.error(f"Ошибка при запросе к документам: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/documents")
async def get_documents():
    """Получить список загруженных документов"""
    logger.info("=== Получение списка документов ===")
    
    if not doc_processor:
        logger.error("Document processor не доступен")
        raise HTTPException(status_code=503, detail="Document processor не доступен")
        
    try:
        doc_list = doc_processor.get_document_list()
        logger.info(f"Список документов: {doc_list}")
        
        return {
            "documents": doc_list,
            "count": len(doc_list) if doc_list else 0,
            "success": True
        }
    except Exception as e:
        logger.error(f"Ошибка при получении списка документов: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/api/documents/{filename}")
async def delete_document(filename: str):
    """Удалить документ по имени файла"""
    logger.info(f"=== Удаление документа: {filename} ===")
    
    if not doc_processor:
        logger.error("Document processor не доступен")
        raise HTTPException(status_code=503, detail="Document processor не доступен")
        
    try:
        # Получаем список документов
        doc_list = doc_processor.get_document_list()
        logger.info(f"Доступные документы до удаления: {doc_list}")
        
        if not doc_list or filename not in doc_list:
            logger.warning(f"Документ {filename} не найден")
            raise HTTPException(status_code=404, detail=f"Документ {filename} не найден")
        
        # Удаляем файл из MinIO, если он там хранится
        documents_bucket = os.getenv('MINIO_DOCUMENTS_BUCKET_NAME', 'memoai-documents')
        if minio_client:
            minio_info = doc_processor.get_image_minio_info(filename)
            if minio_info:
                try:
                    minio_client.delete_file(
                        minio_info["minio_object"],
                        bucket_name=minio_info["minio_bucket"]
                    )
                    logger.info(f"Файл удален из MinIO: {minio_info['minio_bucket']}/{minio_info['minio_object']}")
                except Exception as e:
                    logger.warning(f"Ошибка удаления файла из MinIO: {e}")
        
        # Удаляем документ из процессора
        success = doc_processor.remove_document(filename)
        logger.info(f"Результат удаления: {success}")
        
        if success:
            # Получаем обновленный список документов
            new_doc_list = doc_processor.get_document_list()
            logger.info(f"Документы после удаления: {new_doc_list}")
            
            return {
                "message": f"Документ {filename} успешно удален",
                "success": True,
                "remaining_documents": new_doc_list
            }
        else:
            raise HTTPException(status_code=500, detail="Не удалось удалить документ")
            
    except Exception as e:
        logger.error(f"Ошибка при удалении документа: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/documents/report/generate")
async def generate_confidence_report():
    """Сгенерировать отчет об уверенности модели в распознанном тексте"""
    logger.info("=== Генерация отчета об уверенности ===")
    
    if not doc_processor:
        logger.error("Document processor не доступен")
        raise HTTPException(status_code=503, detail="Document processor не доступен")
    
    try:
        # Получаем данные для отчета
        report_data = doc_processor.get_confidence_report_data()
        logger.info(f"Получены данные отчета: {report_data['total_documents']} документов")
        
        # Формируем текстовый отчет
        report_text = f"""
ОТЧЕТ О СТЕПЕНИ УВЕРЕННОСТИ МОДЕЛИ В РАСПОЗНАННОМ ТЕКСТЕ
{'=' * 80}
Дата генерации: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
{'=' * 80}

ОБЩАЯ ИНФОРМАЦИЯ:
- Всего обработано документов: {report_data['total_documents']}
- Средняя уверенность модели: {report_data['average_confidence']:.2f}%
- Всего слов: {report_data.get('total_words', 0)}
{'=' * 80}

ДЕТАЛЬНАЯ ИНФОРМАЦИЯ ПО ДОКУМЕНТАМ:
"""
        
        # Добавляем распознанный текст с процентами над словами
        for i, doc in enumerate(report_data['documents'], 1):
            report_text += f"""

{i}. {doc['filename']}
   Тип файла: {doc['file_type']}
   Уверенность модели: {doc['confidence']:.2f}%
   Длина распознанного текста: {doc['text_length']} символов
   Количество слов: {doc.get('words_count', 0)}
   {'-' * 80}
   
   РАСПОЗНАННЫЙ ТЕКСТ С УВЕРЕННОСТЬЮ:
"""
            
            # Находим соответствующий отформатированный текст
            formatted_text_info = next((ft for ft in report_data.get('formatted_texts', []) if ft['filename'] == doc['filename']), None)
            
            if formatted_text_info and formatted_text_info.get('words'):
                # Форматируем текст с процентами над словами
                words = formatted_text_info.get('words', [])
                if words:
                    # Группируем слова по строкам для лучшей читаемости
                    line_words = []
                    current_line = []
                    
                    for word_info in words:
                        word = word_info.get('word', '')
                        conf = word_info.get('confidence', 0.0)
                        current_line.append((word, conf))
                        
                        # Каждые 8-10 слов или при достижении символов новой строки
                        if len(current_line) >= 8:
                            line_words.append(current_line)
                            current_line = []
                    
                    if current_line:
                        line_words.append(current_line)
                    
                    # Формируем текст с процентами над словами в красивом формате
                    if line_words:
                        for line in line_words:
                            # Используем табличный формат с фиксированной шириной колонок
                            import re
                            tokens_data = []
                            prev_is_punctuation = False
                            
                            for word, conf in line:
                                is_punctuation = bool(re.match(r'^[^\w\s]+$', word))
                                
                                # Вычисляем ширину колонки на основе длины слова
                                word_width = len(word)
                                # Минимальная ширина колонки - 10 символов (для процента и пробелов)
                                col_width = max(word_width + 2, 10)
                                
                                tokens_data.append({
                                    'word': word,
                                    'conf': conf,
                                    'is_punctuation': is_punctuation,
                                    'col_width': col_width,
                                    'needs_space_before': not prev_is_punctuation and not is_punctuation and tokens_data
                                })
                                prev_is_punctuation = is_punctuation
                            
                            # Формируем строки с выравниванием в табличном формате
                            percent_line = "│"
                            word_line = "│"
                            separator_line = "├"
                            
                            for i, token in enumerate(tokens_data):
                                if token['needs_space_before']:
                                    # Добавляем разделитель между словами
                                    word_line += "│"
                                    percent_line += "│"  # Вертикальный разделитель
                                    separator_line += "┼"
                                
                                # Форматируем процент и слово в колонке
                                percent_str = f"{token['conf']:.0f}%"
                                word_str = token['word']
                                
                                # Выравниваем процент по центру колонки
                                percent_padded = percent_str.center(token['col_width'])
                                # Выравниваем слово по левому краю колонки
                                word_padded = word_str.ljust(token['col_width'])
                                
                                percent_line += percent_padded + "│"
                                word_line += word_padded + "│"
                                separator_line += "─" * token['col_width'] + ("┤" if i == len(tokens_data) - 1 else "┼")
                            
                            # Добавляем в отчет с красивым форматированием
                            report_text += f"   {percent_line}\n"
                            report_text += f"   {separator_line}\n"
                            report_text += f"   {word_line}\n\n"
                    else:
                        report_text += "   [Нет валидных слов для отображения]\n"
                else:
                    report_text += "   [Нет данных о словах]\n"
            else:
                report_text += "   [Нет отформатированного текста]\n"
            
            report_text += f"   {'-' * 80}\n"
        
        # Итоговый процент уверенности
        overall_conf = report_data.get('overall_confidence', report_data.get('average_confidence', 0.0))
        
        report_text += f"""

{'=' * 80}
ИТОГО:
- Итоговая уверенность по всему распознанному тексту: {overall_conf:.2f}%
- Средняя уверенность по документам: {report_data['average_confidence']:.2f}%
- Всего документов: {report_data['total_documents']}
- Всего слов: {report_data.get('total_words', 0)}
{'=' * 80}
"""
        
        # Создаем JSON отчет
        report_json = {
            "generated_at": datetime.now().isoformat(),
            "summary": {
                "total_documents": report_data['total_documents'],
                "average_confidence": round(report_data['average_confidence'], 2),
                "overall_confidence": round(overall_conf, 2),
                "total_words": report_data.get('total_words', 0)
            },
            "documents": report_data['documents']
        }
        
        return {
            "success": True,
            "report_text": report_text,
            "report_json": report_json,
            "summary": {
                "total_documents": report_data['total_documents'],
                "average_confidence": round(report_data['average_confidence'], 2),
                "overall_confidence": round(overall_conf, 2),
                "total_words": report_data.get('total_words', 0)
            }
        }
        
    except Exception as e:
        logger.error(f"Ошибка при генерации отчета: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/documents/report/download")
async def download_confidence_report():
    """Скачать отчет об уверенности в виде Excel файла"""
    logger.info("=== Скачивание отчета об уверенности (Excel) ===")
    
    if not doc_processor:
        logger.error("Document processor не доступен")
        raise HTTPException(status_code=503, detail="Document processor не доступен")
    
    try:
        # Получаем данные для отчета
        report_data = doc_processor.get_confidence_report_data()
        logger.info(f"Получены данные отчета: {report_data['total_documents']} документов")
        
        # Создаем Excel файл
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Отчет об уверенности"
        
        # Стили для заголовков
        header_font = Font(bold=True, size=14, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Стили для подзаголовков
        subheader_font = Font(bold=True, size=12)
        subheader_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
        
        # Стили для процентов уверенности
        high_confidence_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Зеленый для высокой уверенности
        medium_confidence_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Желтый для средней
        low_confidence_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Красный для низкой
        
        # Стили для границ
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        current_row = 1
        
        # Заголовок отчета
        ws.merge_cells(f'A{current_row}:D{current_row}')
        header_cell = ws[f'A{current_row}']
        header_cell.value = "ОТЧЕТ О СТЕПЕНИ УВЕРЕННОСТИ МОДЕЛИ В РАСПОЗНАННОМ ТЕКСТЕ"
        header_cell.font = header_font
        header_cell.fill = header_fill
        header_cell.alignment = header_alignment
        header_cell.border = thin_border
        current_row += 1
        
        # Дата генерации
        ws.merge_cells(f'A{current_row}:D{current_row}')
        date_cell = ws[f'A{current_row}']
        date_cell.value = f"Дата генерации: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        date_cell.alignment = Alignment(horizontal="center")
        current_row += 2
        
        # Общая информация
        if report_data['total_documents'] == 0:
            ws.merge_cells(f'A{current_row}:D{current_row}')
            warning_cell = ws[f'A{current_row}']
            warning_cell.value = "ВНИМАНИЕ: Нет обработанных документов для формирования отчета."
            warning_cell.font = Font(bold=True, color="FF0000")
            warning_cell.alignment = Alignment(horizontal="center")
            current_row += 1
        else:
            # Общая информация
            info_row = current_row
            ws[f'A{info_row}'] = "ОБЩАЯ ИНФОРМАЦИЯ:"
            ws[f'A{info_row}'].font = subheader_font
            ws[f'A{info_row}'].fill = subheader_fill
            current_row += 1
            
            ws[f'A{current_row}'] = "Всего обработано документов:"
            ws[f'B{current_row}'] = report_data['total_documents']
            current_row += 1
            
            ws[f'A{current_row}'] = "Средняя уверенность модели:"
            ws[f'B{current_row}'] = f"{report_data['average_confidence']:.2f}%"
            current_row += 1
            
            ws[f'A{current_row}'] = "Всего слов:"
            ws[f'B{current_row}'] = report_data.get('total_words', 0)
            current_row += 2
            
            # Детальная информация по документам
            for doc_idx, doc in enumerate(report_data.get('documents', []), 1):
                # Заголовок документа
                doc_start_row = current_row
                ws.merge_cells(f'A{current_row}:D{current_row}')
                doc_header = ws[f'A{current_row}']
                doc_header.value = f"{doc_idx}. {doc.get('filename', 'Неизвестный файл')}"
                doc_header.font = subheader_font
                doc_header.fill = subheader_fill
                doc_header.border = thin_border
                current_row += 1
                
                # Информация о документе
                ws[f'A{current_row}'] = "Тип файла:"
                ws[f'B{current_row}'] = doc.get('file_type', 'unknown')
                current_row += 1
                
                ws[f'A{current_row}'] = "Уверенность модели:"
                conf_value = doc.get('confidence', 0.0)
                ws[f'B{current_row}'] = f"{conf_value:.2f}%"
                # Цветовая индикация уверенности
                if conf_value >= 80:
                    ws[f'B{current_row}'].fill = high_confidence_fill
                elif conf_value >= 50:
                    ws[f'B{current_row}'].fill = medium_confidence_fill
                else:
                    ws[f'B{current_row}'].fill = low_confidence_fill
                current_row += 1
                
                ws[f'A{current_row}'] = "Длина текста:"
                ws[f'B{current_row}'] = f"{doc.get('text_length', 0)} символов"
                current_row += 1
                
                ws[f'A{current_row}'] = "Количество слов:"
                ws[f'B{current_row}'] = doc.get('words_count', 0)
                current_row += 2
                
                # Распознанный текст с уверенностью
                formatted_text_info = next((ft for ft in report_data.get('formatted_texts', []) if ft.get('filename') == doc.get('filename')), None)
                
                if formatted_text_info and formatted_text_info.get('words'):
                    words = formatted_text_info.get('words', [])
                    if words:
                        # Заголовок для таблицы слов
                        ws[f'A{current_row}'] = "Слово"
                        ws[f'B{current_row}'] = "Уверенность"
                        ws[f'A{current_row}'].font = Font(bold=True)
                        ws[f'B{current_row}'].font = Font(bold=True)
                        ws[f'A{current_row}'].fill = subheader_fill
                        ws[f'B{current_row}'].fill = subheader_fill
                        ws[f'A{current_row}'].border = thin_border
                        ws[f'B{current_row}'].border = thin_border
                        current_row += 1
                        
                        # Добавляем слова с уверенностью
                        for word_info in words:
                            word = word_info.get('word', '')
                            conf = word_info.get('confidence', 0.0)
                            
                            if word:  # Пропускаем пустые слова
                                ws[f'A{current_row}'] = word
                                ws[f'B{current_row}'] = f"{conf:.1f}%"
                                ws[f'A{current_row}'].border = thin_border
                                ws[f'B{current_row}'].border = thin_border
                                
                                # Цветовая индикация уверенности
                                if conf >= 80:
                                    ws[f'B{current_row}'].fill = high_confidence_fill
                                elif conf >= 50:
                                    ws[f'B{current_row}'].fill = medium_confidence_fill
                                else:
                                    ws[f'B{current_row}'].fill = low_confidence_fill
                                
                                current_row += 1
                
                current_row += 1
            
            # Итоговая информация
            overall_conf = report_data.get('overall_confidence', report_data.get('average_confidence', 0.0))
            ws.merge_cells(f'A{current_row}:D{current_row}')
            summary_header = ws[f'A{current_row}']
            summary_header.value = "ИТОГО"
            summary_header.font = subheader_font
            summary_header.fill = subheader_fill
            summary_header.border = thin_border
            current_row += 1
            
            ws[f'A{current_row}'] = "Итоговая уверенность по всему тексту:"
            ws[f'B{current_row}'] = f"{overall_conf:.2f}%"
            if overall_conf >= 80:
                ws[f'B{current_row}'].fill = high_confidence_fill
            elif overall_conf >= 50:
                ws[f'B{current_row}'].fill = medium_confidence_fill
            else:
                ws[f'B{current_row}'].fill = low_confidence_fill
            current_row += 1
            
            ws[f'A{current_row}'] = "Средняя уверенность по документам:"
            ws[f'B{current_row}'] = f"{report_data['average_confidence']:.2f}%"
            current_row += 1
            
            ws[f'A{current_row}'] = "Всего документов:"
            ws[f'B{current_row}'] = report_data['total_documents']
            current_row += 1
            
            ws[f'A{current_row}'] = "Всего слов:"
            ws[f'B{current_row}'] = report_data.get('total_words', 0)
        
        # Настройка ширины колонок
        ws.column_dimensions['A'].width = 50
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        
        # Сохраняем Excel файл во временный файл
        import tempfile
        temp_dir = tempfile.gettempdir()
        report_filename = f"confidence_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        report_path = os.path.join(temp_dir, report_filename)
        
        try:
            # Убеждаемся, что директория существует
            os.makedirs(temp_dir, exist_ok=True)
            
            # Сохраняем Excel файл
            wb.save(report_path)
            
            logger.info(f"Excel отчет сохранен: {report_path}, размер: {os.path.getsize(report_path)} байт")
            
            # Проверяем, что файл существует
            if not os.path.exists(report_path):
                raise FileNotFoundError(f"Файл отчета не был создан: {report_path}")
            
            # Возвращаем файл для скачивания
            return FileResponse(
                report_path,
                media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                filename=report_filename,
                headers={
                    "Content-Disposition": f"attachment; filename*=UTF-8''{report_filename}"
                }
            )
        except Exception as file_err:
            logger.error(f"Ошибка при сохранении Excel файла отчета: {file_err}")
            raise HTTPException(status_code=500, detail=f"Ошибка при сохранении отчета: {str(file_err)}")
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Ошибка при генерации Excel отчета: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Ошибка при генерации отчета: {str(e)}")

# ================================
# ТРАНСКРИБАЦИЯ
# ================================

@app.post("/api/transcribe/upload")
async def transcribe_file(file: UploadFile = File(...)):
    """Транскрибировать аудио/видео файл с диаризацией по ролям"""
    logger.info(f"=== Начало транскрибации файла с диаризацией: {file.filename} ===")
    
    if not transcriber:
        logger.error("Transcriber не доступен")
        raise HTTPException(status_code=503, detail="Transcriber не доступен")
    
    import tempfile
    temp_dir = tempfile.gettempdir()
    file_path = None
    file_object_name = None
        
    try:
        # Сохраняем файл
        content = await file.read()
        logger.info(f"Файл получен, размер: {len(content)} байт")
        
        # Сохраняем в MinIO или локально
        if minio_client:
            try:
                # Определяем расширение файла
                file_ext = os.path.splitext(file.filename)[1] if file.filename else ""
                file_object_name = minio_client.generate_object_name(prefix="media_", extension=file_ext)
                minio_client.upload_file(content, file_object_name, content_type="application/octet-stream")
                # Получаем локальный путь для обработки
                file_path = minio_client.get_file_path(file_object_name)
                logger.info(f"Файл загружен в MinIO: {file_object_name}")
            except Exception as e:
                logger.warning(f"Ошибка загрузки в MinIO, используем локальный файл: {e}")
                file_path = os.path.join(temp_dir, f"media_{datetime.now().timestamp()}_{file.filename}")
                with open(file_path, "wb") as f:
                    f.write(content)
        else:
            file_path = os.path.join(temp_dir, f"media_{datetime.now().timestamp()}_{file.filename}")
            with open(file_path, "wb") as f:
                f.write(content)
        
        logger.info(f"Временный путь файла: {file_path}")
        logger.info(f"Файл сохранен, размер: {len(content)} байт")
        
        # Транскрибируем с принудительной диаризацией
        logger.info(f"Начинаем транскрибацию с диаризацией по ролям...")
        
        # Проверяем, поддерживает ли транскрайбер диаризацию
        if hasattr(transcriber, 'transcribe_with_diarization'):
            logger.info("Используем принудительную диаризацию...")
            success, result = transcriber.transcribe_with_diarization(file_path)
        else:
            logger.info("Используем стандартную транскрибацию...")
            success, result = transcriber.transcribe_audio_file(file_path)
        
        logger.info(f"Результат транскрибации: success={success}, result_length={len(str(result)) if result else 0}")
        
        if success:
            logger.info("Транскрибация с диаризацией завершена успешно")
            return {
                "transcription": result,
                "filename": file.filename,
                "success": True,
                "timestamp": datetime.now().isoformat(),
                "diarization": True
            }
        else:
            logger.error(f"Ошибка транскрибации: {result}")
            raise HTTPException(status_code=400, detail=result)
            
    except Exception as e:
        logger.error(f"Ошибка в эндпоинте транскрибации: {e}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        # Очистка временных файлов
        try:
            if file_path and os.path.exists(file_path):
                os.remove(file_path)
                logger.info(f"Временный файл удален: {file_path}")
            # Удаляем из MinIO, если был загружен
            if minio_client and file_object_name:
                try:
                    minio_client.delete_file(file_object_name)
                except Exception as e:
                    logger.warning(f"Ошибка удаления файла из MinIO: {e}")
        except Exception as e:
            logger.warning(f"Ошибка очистки временных файлов: {e}")

@app.post("/api/transcribe/upload/diarization")
async def transcribe_file_with_diarization(file: UploadFile = File(...)):
    """Принудительно транскрибировать аудио/видео файл с диаризацией по ролям"""
    logger.info(f"=== Начало принудительной диаризации файла: {file.filename} ===")
    
    if not transcriber:
        logger.error("Transcriber не доступен")
        raise HTTPException(status_code=503, detail="Transcriber не доступен")
    
    import tempfile
    temp_dir = tempfile.gettempdir()
    file_path = None
    file_object_name = None
        
    try:
        # Сохраняем файл
        content = await file.read()
        logger.info(f"Файл получен, размер: {len(content)} байт")
        
        # Сохраняем в MinIO или локально
        if minio_client:
            try:
                # Определяем расширение файла
                file_ext = os.path.splitext(file.filename)[1] if file.filename else ""
                file_object_name = minio_client.generate_object_name(prefix="media_diarization_", extension=file_ext)
                minio_client.upload_file(content, file_object_name, content_type="application/octet-stream")
                # Получаем локальный путь для обработки
                file_path = minio_client.get_file_path(file_object_name)
                logger.info(f"Файл загружен в MinIO: {file_object_name}")
            except Exception as e:
                logger.warning(f"Ошибка загрузки в MinIO, используем локальный файл: {e}")
                file_path = os.path.join(temp_dir, f"media_diarization_{datetime.now().timestamp()}_{file.filename}")
                with open(file_path, "wb") as f:
                    f.write(content)
        else:
            file_path = os.path.join(temp_dir, f"media_diarization_{datetime.now().timestamp()}_{file.filename}")
            with open(file_path, "wb") as f:
                f.write(content)
        
        logger.info(f"Временный путь файла для диаризации: {file_path}")
        logger.info(f"Файл сохранен, размер: {len(content)} байт")
        
        # Принудительная диаризация с WhisperX
        logger.info("Начинаем принудительную диаризацию по ролям...")
        
        if hasattr(transcriber, 'transcribe_with_diarization'):
            success, result = transcriber.transcribe_with_diarization(file_path)
        else:
            logger.warning("Транскрайбер не поддерживает диаризацию, используем стандартную транскрибацию")
            success, result = transcriber.transcribe_audio_file(file_path)
        
        logger.info(f"Результат диаризации: success={success}, result_length={len(str(result)) if result else 0}")
        
        if success:
            logger.info("Диаризация завершена успешно")
            return {
                "transcription": result,
                "filename": file.filename,
                "success": True,
                "timestamp": datetime.now().isoformat(),
                "diarization": True,
                "forced_diarization": True
            }
        else:
            logger.error(f"Ошибка диаризации: {result}")
            raise HTTPException(status_code=400, detail=result)
            
    except Exception as e:
        logger.error(f"Ошибка в эндпоинте диаризации: {e}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        # Очистка временных файлов
        try:
            if file_path and os.path.exists(file_path):
                os.remove(file_path)
                logger.info(f"Временный файл удален: {file_path}")
            # Удаляем из MinIO, если был загружен
            if minio_client and file_object_name:
                try:
                    minio_client.delete_file(file_object_name)
                except Exception as e:
                    logger.warning(f"Ошибка удаления файла из MinIO: {e}")
        except Exception as e:
            logger.warning(f"Ошибка очистки временных файлов: {e}")

@app.post("/api/transcribe/youtube")
async def transcribe_youtube(request: YouTubeTranscribeRequest):
    """Транскрибировать видео с YouTube с диаризацией по ролям"""
    logger.info(f"=== Начало YouTube транскрибации с диаризацией: {request.url} ===")
    
    if not transcriber:
        logger.error("Transcriber не доступен")
        raise HTTPException(status_code=503, detail="Transcriber не доступен")
        
    try:
        logger.info("Начинаем YouTube транскрибацию с диаризацией...")
        success, result = transcriber.transcribe_youtube(request.url)
        logger.info(f"Результат YouTube транскрибации: success={success}, result_length={len(str(result)) if result else 0}")
        
        if success:
            logger.info("YouTube транскрибация с диаризацией завершена успешно")
            return {
                "transcription": result,
                "url": request.url,
                "success": True,
                "timestamp": datetime.now().isoformat(),
                "diarization": True
            }
        else:
            logger.error(f"Ошибка YouTube транскрибации: {result}")
            raise HTTPException(status_code=400, detail=result)
            
    except Exception as e:
        logger.error(f"Ошибка в эндпоинте YouTube транскрибации: {e}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/system/status")
async def get_system_status():
    """Получить статус всех модулей системы"""
    return {
        "modules": {
            "ai_agent": {
                "available": ask_agent is not None,
                "functions": {
                    "ask_agent": ask_agent is not None,
                    "model_settings": model_settings is not None,
                    "update_model_settings": update_model_settings is not None,
                    "reload_model_by_path": reload_model_by_path is not None,
                    "get_model_info": get_model_info is not None,
                    "initialize_model": initialize_model is not None
                }
            },
            "memory": {
                "available": save_dialog_entry is not None,
                "functions": {
                    "save_dialog_entry": save_dialog_entry is not None,
                    "load_dialog_history": load_dialog_history is not None,
                    "clear_dialog_history": clear_dialog_history is not None,
                    "get_recent_dialog_history": get_recent_dialog_history is not None
                }
            },
            "voice": {
                "available": speak_text is not None and recognize_speech_from_file is not None,
                "functions": {
                    "speak_text": speak_text is not None,
                    "recognize_speech": recognize_speech is not None,
                    "recognize_speech_from_file": recognize_speech_from_file is not None,
                    "check_vosk_model": check_vosk_model is not None
                }
            },
            "transcription": {
                "available": transcriber is not None,
                "functions": {
                    "universal_transcriber": UniversalTranscriber is not None,
                    "online_transcriber": OnlineTranscriber is not None
                }
            },
            "document_processor": {
                "available": DocumentProcessor is not None
            }
        },
        "timestamp": datetime.now().isoformat()
    }

# ================================
# ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ
# ================================

def get_current_model_path():
    """Получить путь к текущей загруженной модели"""
    try:
        # Пытаемся получить от AI модуля
        if get_model_info:
            result = get_model_info()
            if result and 'path' in result:
                return result['path']
        
        # Fallback к сохраненным настройкам
        settings = load_app_settings()
        return settings.get('current_model_path')
    except Exception as e:
        logger.error(f"Ошибка получения пути модели: {e}")
        return None

# ================================
# КОНТЕКСТНЫЕ ПРОМПТЫ API
# ================================

@app.get("/api/context-prompts/global")
async def get_global_prompt():
    """Получить глобальный контекстный промпт"""
    try:
        prompt = context_prompt_manager.get_global_prompt()
        return {
            "prompt": prompt,
            "success": True,
            "timestamp": datetime.now().isoformat()
        }
    except Exception as e:
        logger.error(f"Ошибка при получении глобального промпта: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.put("/api/context-prompts/global")
async def update_global_prompt(request: Dict[str, str]):
    """Обновить глобальный контекстный промпт"""
    try:
        prompt = request.get("prompt", "")
        
        success = context_prompt_manager.set_global_prompt(prompt)
        if success:
            return {
                "message": "Глобальный промпт обновлен",
                "success": True,
                "timestamp": datetime.now().isoformat()
            }
        else:
            raise HTTPException(status_code=500, detail="Ошибка при сохранении промпта")
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Ошибка при обновлении глобального промпта: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/context-prompts/models")
async def get_models_with_prompts():
    """Получить список всех моделей с их контекстными промптами"""
    try:
        models = context_prompt_manager.get_models_list()
        return {
            "models": models,
            "success": True,
            "timestamp": datetime.now().isoformat()
        }
    except Exception as e:
        logger.error(f"Ошибка при получении списка моделей: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/context-prompts/model/{model_path:path}")
async def get_model_prompt(model_path: str):
    """Получить контекстный промпт для конкретной модели"""
    try:
        prompt = context_prompt_manager.get_model_prompt(model_path)
        return {
            "model_path": model_path,
            "prompt": prompt,
            "success": True,
            "timestamp": datetime.now().isoformat()
        }
    except Exception as e:
        logger.error(f"Ошибка при получении промпта модели: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.put("/api/context-prompts/model/{model_path:path}")
async def update_model_prompt(model_path: str, request: Dict[str, str]):
    """Обновить контекстный промпт для конкретной модели"""
    try:
        prompt = request.get("prompt", "")
        
        success = context_prompt_manager.set_model_prompt(model_path, prompt)
        if success:
            return {
                "message": f"Промпт для модели {model_path} обновлен",
                "success": True,
                "timestamp": datetime.now().isoformat()
            }
        else:
            raise HTTPException(status_code=500, detail="Ошибка при сохранении промпта")
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Ошибка при обновлении промпта модели: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/context-prompts/custom")
async def get_custom_prompts():
    """Получить все пользовательские промпты"""
    try:
        prompts = context_prompt_manager.get_all_custom_prompts()
        return {
            "prompts": prompts,
            "success": True,
            "timestamp": datetime.now().isoformat()
        }
    except Exception as e:
        logger.error(f"Ошибка при получении пользовательских промптов: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/context-prompts/custom")
async def create_custom_prompt(request: Dict[str, str]):
    """Создать новый пользовательский промпт"""
    try:
        prompt_id = request.get("id", "")
        prompt = request.get("prompt", "")
        description = request.get("description", "")
        
        if not prompt_id.strip() or not prompt.strip():
            raise HTTPException(status_code=400, detail="ID и промпт обязательны")
        
        success = context_prompt_manager.set_custom_prompt(prompt_id, prompt, description)
        if success:
            return {
                "message": f"Пользовательский промпт '{prompt_id}' создан",
                "success": True,
                "timestamp": datetime.now().isoformat()
            }
        else:
            raise HTTPException(status_code=500, detail="Ошибка при создании промпта")
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Ошибка при создании пользовательского промпта: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/api/context-prompts/custom/{prompt_id}")
async def delete_custom_prompt(prompt_id: str):
    """Удалить пользовательский промпт"""
    try:
        success = context_prompt_manager.delete_custom_prompt(prompt_id)
        if success:
            return {
                "message": f"Пользовательский промпт '{prompt_id}' удален",
                "success": True,
                "timestamp": datetime.now().isoformat()
            }
        else:
            raise HTTPException(status_code=404, detail="Промпт не найден")
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Ошибка при удалении пользовательского промпта: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/context-prompts/effective/{model_path:path}")
async def get_effective_prompt(model_path: str, custom_prompt_id: Optional[str] = None):
    """Получить эффективный промпт для модели с учетом приоритетов"""
    try:
        prompt = context_prompt_manager.get_effective_prompt(model_path, custom_prompt_id)
        return {
            "model_path": model_path,
            "custom_prompt_id": custom_prompt_id,
            "prompt": prompt,
            "success": True,
            "timestamp": datetime.now().isoformat()
        }
    except Exception as e:
        logger.error(f"Ошибка при получении эффективного промпта: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# ================================
# АГЕНТНАЯ АРХИТЕКТУРА API
# ================================

class AgentModeRequest(BaseModel):
    mode: str  # "agent", "direct" или "multi-llm"

class MultiLLMModelsRequest(BaseModel):
    models: List[str]  # Список имен моделей для режима multi-llm

class AgentStatusResponse(BaseModel):
    is_initialized: bool
    mode: str
    available_agents: int
    orchestrator_active: bool

@app.get("/api/agent/status")
async def get_agent_status():
    """Получить статус агентной архитектуры"""
    try:
        orchestrator = get_agent_orchestrator()
        if orchestrator:
            status = orchestrator.get_status()
            return AgentStatusResponse(**status)
        else:
            return AgentStatusResponse(
                is_initialized=False,
                mode="unknown",
                available_agents=0,
                orchestrator_active=False
            )
    except Exception as e:
        logger.error(f"Ошибка получения статуса агентной архитектуры: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/agent/mode")
async def set_agent_mode(request: AgentModeRequest):
    """Установить режим работы агентной архитектуры"""
    try:
        orchestrator = get_agent_orchestrator()
        if orchestrator:
            orchestrator.set_mode(request.mode)
            return {
                "message": f"Режим работы изменен на: {request.mode}",
                "success": True,
                "timestamp": datetime.now().isoformat()
            }
        else:
            raise HTTPException(status_code=503, detail="Агентная архитектура не инициализирована")
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Ошибка установки режима агентной архитектуры: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/agent/multi-llm/models")
async def set_multi_llm_models(request: MultiLLMModelsRequest):
    """Установить список моделей для режима multi-llm"""
    try:
        orchestrator = get_agent_orchestrator()
        if orchestrator:
            orchestrator.set_multi_llm_models(request.models)
            return {
                "message": f"Установлены модели для режима multi-llm: {', '.join(request.models)}",
                "success": True,
                "models": request.models,
                "timestamp": datetime.now().isoformat()
            }
        else:
            raise HTTPException(status_code=503, detail="Агентная архитектура не инициализирована")
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Ошибка установки моделей для режима multi-llm: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/agent/multi-llm/models")
async def get_multi_llm_models():
    """Получить список моделей для режима multi-llm"""
    try:
        orchestrator = get_agent_orchestrator()
        if orchestrator:
            models = orchestrator.get_multi_llm_models()
            return {
                "models": models,
                "success": True
            }
        else:
            raise HTTPException(status_code=503, detail="Агентная архитектура не инициализирована")
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Ошибка получения моделей для режима multi-llm: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/agent/agents")
async def get_available_agents():
    """Получить список доступных агентов"""
    try:
        orchestrator = get_agent_orchestrator()
        if orchestrator:
            agents = orchestrator.get_available_agents()
            return {
                "agents": agents,
                "count": len(agents),
                "success": True,
                "timestamp": datetime.now().isoformat()
            }
        else:
            raise HTTPException(status_code=503, detail="Агентная архитектура не инициализирована")
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Ошибка получения списка агентов: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/agent/mcp/status")
async def get_mcp_status():
    """Получить статус MCP серверов"""
    try:
        orchestrator = get_agent_orchestrator()
        if orchestrator:
            # MCP пока не интегрирован в новую архитектуру
            # Возвращаем статус "не инициализирован"
            return {
                "mcp_status": {
                    "initialized": False, 
                    "servers": 0, 
                    "tools": 0,
                    "message": "MCP интеграция в разработке"
                },
                "success": True,
                "timestamp": datetime.now().isoformat()
            }
        else:
            raise HTTPException(status_code=503, detail="Агентная архитектура не инициализирована")
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Ошибка получения статуса MCP: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/agent/agents/{agent_id}/status")
async def set_agent_status(agent_id: str, status: Dict[str, bool]):
    """Установить статус активности агента (теперь инструмента)"""
    try:
        orchestrator = get_agent_orchestrator()
        if not orchestrator:
            raise HTTPException(status_code=503, detail="Агентная архитектура не инициализирована")
        
        is_active = status.get("is_active", True)
        # Используем новый метод для установки статуса инструмента
        orchestrator.set_agent_status(agent_id, is_active)
        success = True
        
        if success:
            return {
                "agent_id": agent_id,
                "is_active": is_active,
                "success": True,
                "message": f"Агент '{agent_id}' {'активирован' if is_active else 'деактивирован'}",
                "timestamp": datetime.now().isoformat()
            }
        else:
            raise HTTPException(status_code=404, detail=f"Агент '{agent_id}' не найден")
            
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Ошибка изменения статуса агента: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/agent/agents/statuses")
async def get_all_agent_statuses():
    """Получить статусы всех агентов (теперь инструментов)"""
    try:
        orchestrator = get_agent_orchestrator()
        if not orchestrator:
            raise HTTPException(status_code=503, detail="Агентная архитектура не инициализирована")
        
        # Используем новый метод для получения статусов инструментов
        statuses = orchestrator.get_all_agent_statuses()
        return {
            "statuses": statuses,
            "success": True,
            "timestamp": datetime.now().isoformat()
        }
            
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Ошибка получения статусов агентов: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/agent/langgraph/status")
async def get_langgraph_status():
    """Получить статус LangGraph оркестратора"""
    try:
        orchestrator = get_agent_orchestrator()
        if orchestrator:
            # LangGraph теперь = сам оркестратор
            # Возвращаем статус оркестратора
            tools = orchestrator.get_available_tools()
            return {
                "langgraph_status": {
                    "is_active": orchestrator.is_initialized,
                    "initialized": orchestrator.is_initialized,
                    "tools_available": len(tools),
                    "memory_enabled": True,
                    "orchestrator_type": "LangGraph",
                    "orchestrator_active": orchestrator.is_orchestrator_active()
                },
                "success": True,
                "timestamp": datetime.now().isoformat()
            }
        else:
            raise HTTPException(status_code=503, detail="Агентная архитектура не инициализирована")
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Ошибка получения статуса LangGraph: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/agent/orchestrator/toggle")
async def toggle_orchestrator(status: Dict[str, bool]):
    """Включить/выключить оркестратор"""
    try:
        orchestrator = get_agent_orchestrator()
        if not orchestrator:
            raise HTTPException(status_code=503, detail="Агентная архитектура не инициализирована")
        
        is_active = status.get("is_active", True)
        
        # Устанавливаем статус оркестратора
        orchestrator.set_orchestrator_status(is_active)
        
        return {
            "success": True,
            "orchestrator_active": is_active,
            "message": f"Оркестратор {'включен' if is_active else 'отключен'}",
            "timestamp": datetime.now().isoformat()
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Ошибка переключения оркестратора: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# ================================
# СТАТИЧЕСКИЕ ФАЙЛЫ И ФРОНТЕНД
# ================================

# Подключаем статические файлы React приложения (только для локального запуска)
# В Docker фронтенд работает в отдельном контейнере на порту 3000
is_docker = os.getenv("DOCKER_ENV", "").lower() == "true"
if not is_docker and os.path.exists("../frontend/build"):
    app.mount("/static", StaticFiles(directory="../frontend/build/static"), name="static")
    
    @app.get("/{path:path}")
    async def serve_react_app(path: str):
        """Отдаем React приложение для всех остальных маршрутов"""
        index_file = "../frontend/build/index.html"
        if os.path.exists(index_file):
            return FileResponse(index_file)
        else:
            return {"message": "Frontend not built"}

if __name__ == "__main__":
    print("Запуск MemoAI Web Backend...")
    print(f"Текущая директория: {os.getcwd()}")
    print(f"Backend директория: {os.path.dirname(os.path.abspath(__file__))}")
    print(f"Корневая директория: {os.path.dirname(os.path.dirname(os.path.abspath(__file__)))}")
    print(f"Python path: {sys.path[:3]}...")
    print("API документация: http://localhost:8000/docs")
    print("WebSocket: ws://localhost:8000/ws/chat")
    
    # Восстанавливаем сохраненную модель
    try:
        settings = load_app_settings()
        saved_model_path = settings.get('current_model_path')
        
        if saved_model_path and os.path.exists(saved_model_path) and reload_model_by_path:
            logger.info(f"Восстанавливаю сохраненную модель: {saved_model_path}")
            success = reload_model_by_path(saved_model_path)
            if success:
                logger.info(f"Модель восстановлена: {saved_model_path}")
            else:
                logger.warning(f"Не удалось восстановить модель: {saved_model_path}")
        else:
            logger.info("Нет сохраненной модели для восстановления")
    except Exception as e:
        logger.error(f"Ошибка восстановления модели: {e}")
    
    uvicorn.run(
        app,  # Передаем объект app напрямую
        host="0.0.0.0",
        port=8000,
        reload=False,  # Отключаем reload для избежания проблем
        log_level="info"
    )
