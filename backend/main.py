# Настройка кодировки для Windows
import sys
import os

# Импортируем утилиту для исправления кодировки
try:
    from utils.encoding_fix import fix_windows_encoding, safe_print
    fix_windows_encoding()
except ImportError:
    # Если утилита недоступна, используем базовую настройку
    if sys.platform == "win32":
        os.system("chcp 65001 >nul 2>&1")
        if hasattr(sys.stdout, 'reconfigure'):
            sys.stdout.reconfigure(encoding='utf-8')
        if hasattr(sys.stderr, 'reconfigure'):
            sys.stderr.reconfigure(encoding='utf-8')

from fastapi import FastAPI, WebSocket, WebSocketDisconnect, HTTPException, UploadFile, File, Form
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse
import uvicorn
import asyncio
import json
import os
import sys
import traceback
from typing import List, Dict, Any, Optional
from datetime import datetime
import logging
from socketio import AsyncServer, ASGIApp
from starlette.applications import Starlette

# Добавляем корневую директорию в путь для импорта модулей
current_dir = os.path.dirname(os.path.abspath(__file__))
root_dir = os.path.dirname(current_dir)

# Загружаем переменные окружения из .env файла (до инициализации логгера)
try:
    from dotenv import load_dotenv
    env_path = os.path.join(root_dir, '.env')
    if os.path.exists(env_path):
        load_dotenv(env_path)
        print(f".env файл загружен: {env_path}")
        # Проверяем MongoDB настройки
        mongodb_user = os.getenv("MONGODB_USER", "").strip()
        mongodb_password = os.getenv("MONGODB_PASSWORD", "").strip()
        if mongodb_user.startswith('#') or mongodb_password.startswith('#'):
            print(f"ВНИМАНИЕ: MONGODB_USER или MONGODB_PASSWORD начинаются с '#', будут игнорироваться")
    else:
        print(f".env файл не найден: {env_path}")
except ImportError:
    print("python-dotenv не установлен, переменные окружения не будут загружены из .env")

# В Docker: /app содержит main.py, agent_llm_svc.py и т.д.
# Для импортов "from backend.xxx" нужно чтобы /app был доступен как /backend
# Создаем временную структуру для импортов
if current_dir == '/app' and not os.path.exists('/app/backend'):
    # В Docker контейнере создаем symbolic link
    os.system('ln -sf /app /app/backend')

sys.path.insert(0, current_dir)  # Для доступа к /app/config
sys.path.insert(0, root_dir)     # Для доступа к / для импортов "from backend.xxx"

# Импортируем конфигурацию
from settings import get_settings

# Получаем настройки для использования во всем приложении
settings = get_settings()

# --- ОБНОВЛЕНО: Добавлены новые микросервисы в конфиг URL ---
urls_config = {
    "frontend_port_1": settings.urls.frontend_port_1,
    "frontend_port_1_ipv4": settings.urls.frontend_port_1_ipv4,
    "frontend_port_2": settings.urls.frontend_port_2,
    "frontend_port_2_ipv4": settings.urls.frontend_port_2_ipv4,
    "frontend_port_3": settings.urls.frontend_port_3,
    "frontend_port_3_ipv4": settings.urls.frontend_port_3_ipv4,
    "backend_port_1": settings.urls.backend_port_1,
    "backend_port_1_ipv4": settings.urls.backend_port_1_ipv4,
    "backend_port_2": settings.urls.backend_port_2,
    "backend_port_2_ipv4": settings.urls.backend_port_2_ipv4,
    "llm_service_port": settings.urls.llm_service_port,
    "frontend_docker": settings.urls.frontend_docker,
    "backend_docker": settings.urls.backend_docker,
    # Обновленные Docker адреса для всех сервисов:
    "llm_service_docker": settings.urls.llm_service_docker,
    "stt_service_docker": settings.urls.stt_service_docker,
    "tts_service_docker": settings.urls.tts_service_docker,
    "ocr_service_docker": settings.urls.ocr_service_docker,
    "diarization_service_docker": settings.urls.diarization_service_docker,
}

# Настройка логирования в самом начале
# Настройка логирования с поддержкой UTF-8
import logging
logging.basicConfig(
    level=logging.DEBUG,
    format='[%(asctime)s] %(levelname)s [Backend] %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S',
    handlers=[
        logging.StreamHandler(sys.stdout)
    ]
)

# Настройка кодировки для обработчиков логирования
for handler in logging.root.handlers:
    if hasattr(handler, 'stream') and hasattr(handler.stream, 'reconfigure'):
        handler.stream.reconfigure(encoding='utf-8')

# Отключаем DEBUG логи от pymongo (MongoDB heartbeat)
logging.getLogger('pymongo').setLevel(logging.WARNING)
logging.getLogger('pymongo.topology').setLevel(logging.WARNING)
logging.getLogger('pymongo.connection').setLevel(logging.WARNING)
logging.getLogger('pymongo.serverSelection').setLevel(logging.WARNING)

logger = logging.getLogger(__name__)
logger.info("Логирование настроено")

# Проверяем, что .env файл был загружен и MongoDB настройки доступны
logger.info("Проверка переменных окружения MongoDB...")
mongodb_host = os.getenv("MONGODB_HOST", "localhost")
mongodb_port = os.getenv("MONGODB_PORT", "27017")
mongodb_user = os.getenv("MONGODB_USER", "").strip()
mongodb_password = os.getenv("MONGODB_PASSWORD", "").strip()
logger.info(f"MONGODB_HOST: {mongodb_host}")
logger.info(f"MONGODB_PORT: {mongodb_port}")
logger.info(f"MONGODB_USER: '{mongodb_user}' (len={len(mongodb_user)})")
logger.info(f"MONGODB_PASSWORD: {'*' * len(mongodb_password) if mongodb_password else ''} (len={len(mongodb_password)})")
if mongodb_user.startswith('#') or mongodb_password.startswith('#'):
    logger.warning("MONGODB_USER или MONGODB_PASSWORD начинаются с '#', будут игнорироваться")

# Импорт authentication router
try:
    logger.info("Попытка импорта auth router...")
    from backend.auth.routes import router as auth_router
    logger.info("auth router импортирован успешно")
except ImportError as e:
    logger.warning(f"auth router недоступен: {e}")
    auth_router = None
except Exception as e:
    logger.warning(f"Ошибка при импорте auth router: {e}")
    auth_router = None

# Импорт prompts gallery router
try:
    logger.info("Попытка импорта prompts router...")
    from backend.api_prompts import router as prompts_router
    logger.info("prompts router импортирован успешно")
except ImportError as e:
    logger.warning(f"prompts router недоступен: {e}")
    prompts_router = None
except Exception as e:
    logger.warning(f"Ошибка при импорте prompts router: {e}")
    prompts_router = None

# Импорт agents gallery router
try:
    logger.info("Попытка импорта agents router...")
    from backend.api_agents import router as agents_router
    logger.info("agents router импортирован успешно")
except ImportError as e:
    logger.warning(f"agents router недоступен: {e}")
    agents_router = None
except Exception as e:
    logger.warning(f"Ошибка при импорте agents router: {e}")
    agents_router = None

# Импорт share router
try:
    logger.info("Попытка импорта share router...")
    from backend.routes.share import router as share_router
    logger.info("share router импортирован успешно")
except ImportError as e:
    logger.warning(f"share router недоступен: {e}")
    share_router = None
except Exception as e:
    logger.warning(f"Ошибка при импорте share router: {e}")
    share_router = None

# Импортируем agent_llm_svc
try:
    logger.info("Попытка импорта agent_llm_svc...")
    from backend.agent_llm_svc import ask_agent, model_settings, update_model_settings, reload_model_by_path, get_model_info, initialize_model
    from backend.context_prompts import context_prompt_manager
    logger.info("agent_llm_svc импортирован успешно")
    if ask_agent:
        logger.info("ask_agent функция доступна")
    else:
        logger.warning("ask_agent функция не доступна")
except ImportError as e:
    logger.error(f"Ошибка импорта agent_llm_svc: {e}")
    print(f"Ошибка импорта agent_llm_svc: {e}")
    print(f"Текущий путь: {os.getcwd()}")
    print(f"Python path: {sys.path}")
    ask_agent = None
    model_settings = None
    update_model_settings = None
    reload_model_by_path = None
    get_model_info = None
    initialize_model = None
    context_prompt_manager = None
except Exception as e:
    logger.error(f"Неожиданная ошибка при импорте agent_llm_svc: {e}")
    import traceback
    logger.error(f"Traceback: {traceback.format_exc()}")
    ask_agent = None
    model_settings = None
    update_model_settings = None
    reload_model_by_path = None
    get_model_info = None
    initialize_model = None
    context_prompt_manager = None

# Общая блокировка для загрузки моделей в режиме multi-llm
# Используется для предотвращения конфликтов при параллельной загрузке
import threading
model_load_lock = threading.Lock()
    
try:
    logger.info("Попытка импорта memory_service (MongoDB)...")
    from backend.database.memory_service import (
        save_dialog_entry, 
        load_dialog_history, 
        clear_dialog_history, 
        get_recent_dialog_history,
        reset_conversation,
        get_or_create_conversation_id,
        remove_last_user_message
    )
    logger.info("memory_service (MongoDB) импортирован успешно")
    logger.info(f"save_dialog_entry импортирован: {save_dialog_entry is not None}, type: {type(save_dialog_entry)}")
    if save_dialog_entry:
        logger.info("save_dialog_entry функция доступна (MongoDB)")
    else:
        logger.error("save_dialog_entry функция не доступна (None или False)")

except ImportError as e:
    logger.error(f"Ошибка импорта memory_service: {e}")
    logger.error("MongoDB memory_service недоступен! Приложение не сможет сохранять диалоги.")
    save_dialog_entry = None
    load_dialog_entry = None
    load_dialog_history = None
    clear_dialog_history = None
    get_recent_dialog_history = None
    reset_conversation = None
    get_or_create_conversation_id = None
    remove_last_user_message = None
except Exception as e:
    logger.error(f"Неожиданная ошибка при импорте memory_service: {e}")
    import traceback
    logger.error(f"Traceback: {traceback.format_exc()}")
    save_dialog_entry = None
    load_dialog_entry = None
    load_dialog_history = None
    clear_dialog_history = None
    get_recent_dialog_history = None
    reset_conversation = None
    get_or_create_conversation_id = None
    remove_last_user_message = None
    
try:
    logger.info("Попытка импорта voice...")
    from backend.voice import speak_text, recognize_speech, recognize_speech_from_file, check_vosk_model
    logger.info("voice импортирован успешно")

except ImportError as e:
    logger.error(f"Ошибка импорта voice: {e}")
    print(f"Ошибка импорта voice: {e}")
    speak_text = None
    recognize_speech = None
    recognize_speech_from_file = None
    check_vosk_model = None
except Exception as e:
    logger.error(f"Неожиданная ошибка при импорте voice: {e}")
    import traceback
    logger.error(f"Traceback: {traceback.format_exc()}")
    speak_text = None
    recognize_speech = None
    recognize_speech_from_file = None
    check_vosk_model = None

# Импорт MinIO клиента для хранения временных файлов
try:
    logger.info("Попытка импорта MinIO клиента...")
    from backend.database.minio import get_minio_client
    minio_client = get_minio_client()
    if minio_client:
        logger.info("MinIO клиент успешно инициализирован")
    else:
        logger.warning("MinIO клиент недоступен, будут использоваться локальные временные файлы")
except ImportError as e:
    logger.warning(f"MinIO клиент недоступен: {e}. Будут использоваться локальные временные файлы")
    minio_client = None
except Exception as e:
    logger.warning(f"Ошибка при инициализации MinIO клиента: {e}. Будут использоваться локальные временные файлы")
    import traceback
    logger.error(f"Traceback: {traceback.format_exc()}")
    minio_client = None

# Импорт клиента для SVC-RAG (RAG‑логика вынесена в отдельный сервис)
try:
    logger.info("Попытка импорта RagClient (SVC-RAG)...")
    from backend.settings.rag_client import get_rag_client
    rag_client = get_rag_client()
    logger.info(f"RagClient инициализирован, base_url={rag_client.base_url}")
except ImportError as e:
    logger.warning(f"RagClient (backend.rag_client) недоступен: {e}")
    rag_client = None
except Exception as e:
    logger.warning(f"Неожиданная ошибка при инициализации RagClient: {e}")
    import traceback
    logger.error(f"Traceback: {traceback.format_exc()}")
    rag_client = None
    
try:
    logger.info("Попытка импорта universal_transcriber...")
    from backend.universal_transcriber import UniversalTranscriber
    logger.info("universal_transcriber импортирован успешно")
except ImportError as e:
    logger.error(f"Ошибка импорта universal_transcriber: {e}")
    import traceback
    logger.error(f"Traceback: {traceback.format_exc()}")
    print("Предупреждение: модуль universal_transcriber не найден")
    UniversalTranscriber = None
except Exception as e:
    logger.error(f"Неожиданная ошибка при импорте universal_transcriber: {e}")
    import traceback
    logger.error(f"Traceback: {traceback.format_exc()}")
    UniversalTranscriber = None

# Импорт агентной архитектуры
try:
    logger.info("Попытка импорта агентной архитектуры...")
    from backend.orchestrator import initialize_agent_orchestrator, get_agent_orchestrator
    logger.info("Агентная архитектура импортирована успешно")
except ImportError as e:
    logger.error(f"Ошибка импорта агентной архитектуры: {e}")
    print("Предупреждение: модуль агентной архитектуры не найден")
    initialize_agent_orchestrator = None
    get_agent_orchestrator = None
except Exception as e:
    logger.error(f"Неожиданная ошибка при импорте агентной архитектуры: {e}")
    logger.error(f"Traceback: {traceback.format_exc()}")
    initialize_agent_orchestrator = None
    get_agent_orchestrator = None

# Импорт модуля баз данных
try:
    logger.info("Попытка импорта database модуля...")
    from backend.database.init_db import (
        init_databases, 
        close_databases,
        get_conversation_repository,
        get_document_repository,
        get_vector_repository
    )
    logger.info("Database модуль импортирован успешно")
    database_available = True
except ImportError as e:
    logger.warning(f"Database модуль недоступен: {e}")
    logger.warning("Приложение будет работать без баз данных (файловый режим)")
    init_databases = None
    close_databases = None
    get_conversation_repository = None
    get_document_repository = None
    get_vector_repository = None
    database_available = False
except Exception as e:
    logger.error(f"Неожиданная ошибка при импорте database модуля: {e}")
    logger.error(f"Traceback: {traceback.format_exc()}")
    init_databases = None
    close_databases = None
    get_conversation_repository = None
    get_document_repository = None
    get_vector_repository = None
    database_available = False

# Глобальный словарь для хранения флагов остановки генерации
stop_generation_flags = {}

# Глобальный флаг для остановки голосового чата
voice_chat_stop_flag = False

# Флаги для остановки транскрибации (по session_id или user_id)
stop_transcription_flags = {}

# Создание Socket.IO сервера
# Все URL читаются из секции urls конфига
socketio_origins = [
    urls_config.get("frontend_port_1"),
    urls_config.get("frontend_port_1_ipv4"),
    urls_config.get("backend_port_1"),
    urls_config.get("backend_port_1_ipv4"),
    urls_config.get("frontend_docker"),
    urls_config.get("backend_docker"),
]
# Фильтруем None значения
socketio_origins = [origin for origin in socketio_origins if origin]
sio = AsyncServer(
    async_mode='asgi',
    cors_allowed_origins=socketio_origins,
    ping_timeout=300,  # ping timeout до 5 минут (для долгих генераций)
    ping_interval=15,  # Отправляем ping каждые 15 секунд
    logger=False,  # Отключено логирование Socket.IO (мешает в консоли)
    engineio_logger=False  # Отключено логирование engine.io (мешает в консоли)
)

# Создание FastAPI приложения с конфигурацией
app_config = settings.app
app = FastAPI(
    title=app_config.name,
    description=app_config.description,
    version=app_config.version,
    debug=app_config.debug
)

# Настройка CORS из конфигурации
cors_origins_from_config = settings.cors.allowed_origins
# Если в настройках не указаны origins, они уже были заполнены автоматически из urls
# Фильтруем None значения
cors_origins_from_config = [origin for origin in cors_origins_from_config if origin]
app.add_middleware(
    CORSMiddleware,
    allow_origins=cors_origins_from_config,
    allow_credentials=settings.cors.allow_credentials,
    allow_methods=settings.cors.allow_methods,
    allow_headers=settings.cors.allow_headers,
)

# Подключаем authentication routes
if auth_router:
    app.include_router(auth_router)
    logger.info("Auth routes подключены (/api/auth/*)")
else:
    logger.warning("Auth routes не подключены (auth_router недоступен)")

# Подключаем prompts gallery routes
if prompts_router:
    app.include_router(prompts_router)
    logger.info("Prompts gallery routes подключены (/api/prompts/*)")
else:
    logger.warning("Prompts gallery routes не подключены (prompts_router недоступен)")

# Подключаем agents gallery routes
if agents_router:
    app.include_router(agents_router)
    logger.info("Agents gallery routes подключены (/api/agents/*)")
else:
    logger.warning("Agents gallery routes не подключены (agents_router недоступен)")

# Подключаем share routes
if share_router:
    app.include_router(share_router)
    logger.info("Share routes подключены (/api/share/*)")
else:
    logger.warning("Share routes не подключены (share_router недоступен)")

# Startup событие для инициализации агентной архитектуры и баз данных
@app.on_event("startup")
async def startup_event():
    """Инициализация при запуске приложения"""
    logger.info("Запуск приложения...")
    
    # Инициализируем базы данных
    logger.info(f"Проверка доступности баз данных: init_databases={init_databases is not None}, database_available={database_available}")
    if init_databases and database_available:
        try:
            logger.info("Инициализация баз данных...")
            success = await init_databases()
            if success:
                logger.info("Базы данных успешно инициализированы")
                logger.info("- MongoDB: готов для хранения диалогов")
                logger.info("- PostgreSQL + pgvector: готов для RAG системы")
                
                # Пересоздаем пул PostgreSQL в текущем event loop (FastAPI)
                # Это необходимо, так как пул был создан в другом event loop при инициализации
                try:
                    from backend.database.init_db import postgresql_connection
                    if postgresql_connection:
                        logger.info("Пересоздание пула PostgreSQL в event loop FastAPI...")
                        await postgresql_connection.ensure_pool()
                        logger.info("Пул PostgreSQL пересоздан в event loop FastAPI")
                except Exception as e:
                    logger.warning(f"Не удалось пересоздать пул PostgreSQL: {e}")
                
                # Проверяем статус MinIO
                if minio_client:
                    logger.info(f"- MinIO: готов для хранения файлов (endpoint: {minio_client.endpoint})")
                else:
                    logger.warning("- MinIO: не инициализирован, используется локальное хранение")
            else:
                logger.warning("Не удалось инициализировать некоторые базы данных")
                logger.warning("Приложение продолжит работу в файловом режиме")
        except Exception as e:
            logger.error(f"Ошибка инициализации баз данных: {e}")
            logger.error(f"Traceback: {traceback.format_exc()}")
            logger.warning("Приложение продолжит работу в файловом режиме")
    else:
        if not init_databases:
            logger.warning("init_databases не импортирован или недоступен")
        if not database_available:
            logger.warning("database_available = False")
        logger.warning("Базы данных не настроены, используется файловый режим")
    
    # Инициализируем агентную архитектуру
    if initialize_agent_orchestrator:
        try:
            success = await initialize_agent_orchestrator()
            if success:
                logger.info("Агентная архитектура успешно инициализирована")
            else:
                logger.warning("Не удалось инициализировать агентную архитектуру")
        except Exception as e:
            logger.error(f"Ошибка инициализации агентной архитектуры: {e}")
    
    logger.info("Приложение запущено")

# Shutdown событие для корректного закрытия подключений
@app.on_event("shutdown")
async def shutdown_event():
    """Корректное закрытие при остановке приложения"""
    logger.info("Остановка приложения...")
    
    # Закрываем подключения к базам данных
    if close_databases and database_available:
        try:
            logger.info("Закрытие подключений к базам данных...")
            await close_databases()
            logger.info("Подключения к базам данных закрыты")
        except Exception as e:
            logger.error(f"Ошибка при закрытии баз данных: {e}")
    
    logger.info("Приложение остановлено")

# Создание Starlette приложения для Socket.IO
starlette_app = Starlette()
socket_app = ASGIApp(sio, starlette_app)

# Монтирование Socket.IO
app.mount("/socket.io", socket_app)

logger.info("=== Инициализация сервисов ===")

# Инициализируем базы данных (для чатов/промптов/агентов и др.)
if init_databases and database_available:
    try:
        logger.info("Инициализация баз данных...")
        import asyncio
        try:
            loop = asyncio.get_event_loop()
        except RuntimeError:
            loop = asyncio.new_event_loop()
            asyncio.set_event_loop(loop)
        if loop.is_running():
            logger.info("Event loop уже запущен (uvicorn), init_databases() будет вызываться в другом месте при необходимости")
        else:
            logger.info("Вызываем init_databases()...")
            success = loop.run_until_complete(init_databases())
            if success:
                logger.info("Базы данных успешно инициализированы")
            else:
                logger.error("Не удалось инициализировать базы данных")
    except Exception as e:
        logger.error(f"Ошибка инициализации баз данных: {e}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
else:
    if not init_databases:
        logger.warning("init_databases недоступен")
    if not database_available:
        logger.warning("database_available = False")

try:
    if UniversalTranscriber:
        logger.info("Инициализация UniversalTranscriber с движком whisperx...")
        transcriber = UniversalTranscriber(engine="whisperx")
        if transcriber:
            logger.info("UniversalTranscriber инициализирован успешно")
        else:
            logger.warning("UniversalTranscriber не удалось создать")
    else:
        logger.warning("UniversalTranscriber не доступен")
        transcriber = None
except Exception as e:
    logger.error(f"Ошибка инициализации UniversalTranscriber: {e}")
    logger.error(f"Traceback: {traceback.format_exc()}")
    transcriber = None

logger.info("=== Инициализация сервисов завершена ===")

# Глобальные настройки транскрибации
current_transcription_engine = "whisperx"
current_transcription_language = "ru"

# Глобальная переменная для хранения выбранной стратегии RAG
current_rag_strategy = "auto"  # auto, reranking, hierarchical, hybrid, standard

# Глобальные настройки памяти
memory_max_messages = 20
memory_include_system_prompts = True
memory_clear_on_restart = False

# Путь к файлу настроек
SETTINGS_FILE = os.path.join(os.path.dirname(__file__), "..", "settings.json")

def load_app_settings():
    """Загрузить настройки приложения из файла """
    global current_transcription_engine, current_transcription_language, memory_max_messages, memory_include_system_prompts, memory_clear_on_restart, current_rag_strategy
    
    try:
        if os.path.exists(SETTINGS_FILE):
            with open(SETTINGS_FILE, 'r', encoding='utf-8') as f:
                settings_data = json.load(f)
            
            current_transcription_engine = settings_data.get('transcription_engine', 'whisperx')
            current_transcription_language = settings_data.get('transcription_language', 'ru')
            memory_max_messages = settings_data.get('memory_max_messages', 20)
            memory_include_system_prompts = settings_data.get('memory_include_system_prompts', True)
            memory_clear_on_restart = settings_data.get('memory_clear_on_restart', False)
            current_rag_strategy = settings_data.get('rag_strategy', 'auto')
            
            logger.info(f"Настройки загружены: engine={current_transcription_engine}, language={current_transcription_language}, memory_max_messages={memory_max_messages}, rag_strategy={current_rag_strategy}")
            return settings_data
    except Exception as e:
        logger.error(f"Ошибка загрузки настроек: {e}")
    
    return {
        'transcription_engine': current_transcription_engine,
        'transcription_language': current_transcription_language,
        'memory_max_messages': memory_max_messages,
        'memory_include_system_prompts': memory_include_system_prompts,
        'memory_clear_on_restart': memory_clear_on_restart,
        'rag_strategy': current_rag_strategy,
        'current_model_path': None
    }

def save_app_settings(settings_to_save):
    """Сохранить настройки приложения в файл"""
    try:
        existing_settings = {}
        if os.path.exists(SETTINGS_FILE):
            with open(SETTINGS_FILE, 'r', encoding='utf-8') as f:
                existing_settings = json.load(f)
        
        existing_settings.update(settings_to_save)
        with open(SETTINGS_FILE, 'w', encoding='utf-8') as f:
            json.dump(existing_settings, f, ensure_ascii=False, indent=2)
        
        logger.info(f"Настройки сохранены: {settings_to_save}")
        return True
    except Exception as e:
        logger.error(f"Ошибка сохранения настроек: {e}")
        return False

# Загружаем настройки при старте 
loaded_settings = load_app_settings()

# Очищаем память при перезапуске 
if memory_clear_on_restart and clear_dialog_history:
    try:
        logger.info("Очистка памяти при перезапуске (настройка включена)")
        clear_dialog_history()
        logger.info("Память очищена при перезапуске")
    except Exception as e:
        logger.warning(f"Не удалось очистить память при перезапуске: {e}")

# WebSocket менеджер
class ConnectionManager:
    def __init__(self):
        self.active_connections: List[WebSocket] = []

    async def connect(self, websocket: WebSocket):
        await websocket.accept()
        self.active_connections.append(websocket)
        logger.info(f"WebSocket connection established. Total connections: {len(self.active_connections)}")

    def disconnect(self, websocket: WebSocket):
        self.active_connections.remove(websocket)
        logger.info(f"WebSocket connection closed. Total connections: {len(self.active_connections)}")

    async def send_personal_message(self, message: str, websocket: WebSocket):
        await websocket.send_text(message)

    async def broadcast(self, message: str):
        for connection in self.active_connections:
            try:
                await connection.send_text(message)
            except:
                pass

manager = ConnectionManager()

# Socket.IO события 
@sio.event
async def connect(sid, environ):
    logger.info(f"Socket.IO client connected: {sid}")
    stop_generation_flags[sid] = False
    await sio.emit('connected', {'data': 'Connected to astrachat'}, room=sid)

@sio.event
async def disconnect(sid):
    logger.info(f"Socket.IO client disconnected: {sid}")
    if sid in stop_generation_flags:
        del stop_generation_flags[sid]

@sio.event
async def ping(sid, data):
    """Обработка heartbeat ping от клиента"""
    try:
        await sio.emit('pong', {
            'timestamp': data.get('timestamp', 0),
            'server_time': datetime.now().isoformat()
        }, room=sid)
    except Exception as e:
        logger.error(f"Ошибка обработки ping: {e}")

@sio.event
async def stop_generation(sid, data):
    """Обработка команды остановки генерации """
    logger.info(f"Socket.IO: получена команда остановки генерации от {sid}")
    stop_generation_flags[sid] = True
    await sio.emit('generation_stopped', {
        'content': 'Генерация остановлена',
        'timestamp': datetime.now().isoformat()
    }, room=sid)

@sio.event
async def stop_transcription(sid, data):
    """Обработка команды остановки транскрибации """
    logger.info(f"Socket.IO: получена команда остановки транскрибации от {sid}")
    stop_transcription_flags[sid] = True
    await sio.emit('transcription_stopped', {
        'message': 'Транскрибация остановлена',
        'timestamp': datetime.now().isoformat()
    }, room=sid)


def _is_structure_query(text: str) -> bool:
    """Запрос про оглавление/структуру/главы — тогда добавляем начало документа в RAG."""
    if not text or len(text.strip()) < 3:
        return False
    t = text.lower().strip()
    keywords = ("оглавление", "содержание", "главы", "глава", "пункт", "подпункт", "структура работы", "структуру работы", "названия глав", "какие главы")
    return any(k in t for k in keywords)


def _terminal_chat_inference_banner(
    *,
    sid: str,
    conversation_id,
    user_preview: str,
    mode_label: str,
    model_path_for_call: str = None,
    extra_line: str = None,
):
    """
    Явный вывод в терминал (stdout) и в лог: какая модель, настройки и системный промпт
    участвуют в ответе. Агент с UI не передаётся по WebSocket — на сервере действуют
    загруженная модель, model_settings и глобальный промпт (после PUT из галереи и т.д.).
    """
    import json

    lines = [
        "",
        "=" * 76,
        "  [ЧАТ] Генерация ответа — что использует сервер СЕЙЧАС",
        "=" * 76,
        f"  Режим: {mode_label}",
        f"  Socket: {sid[:20]}…  |  conversation_id: {conversation_id}",
        f"  Текст запроса (начало): {(user_preview or '')[:160]!r}",
    ]
    path = model_path_for_call if model_path_for_call is not None else get_current_model_path()
    lines.append(f"  Модель (путь для вызова LLM): {path!r}")
    if get_model_info:
        try:
            info = get_model_info()
            if info:
                lines.append(f"  get_model_info: {json.dumps(info, ensure_ascii=False, default=str)[:500]}")
        except Exception as ex:
            lines.append(f"  get_model_info: ошибка {ex}")
    if model_settings:
        try:
            st = model_settings.get_all()
            lines.append(f"  Настройки модели: {json.dumps(st, ensure_ascii=False, default=str)}")
        except Exception as ex:
            lines.append(f"  Настройки модели: недоступны ({ex})")
    else:
        lines.append("  Настройки модели: модуль недоступен")
    if context_prompt_manager:
        try:
            gp = context_prompt_manager.get_global_prompt() or ""
            prev = gp[:500] + ("…" if len(gp) > 500 else "")
            lines.append(f"  Глобальный системный промпт ({len(gp)} симв., начало): {prev!r}")
        except Exception as ex:
            lines.append(f"  Глобальный промпт: ошибка {ex}")
    else:
        lines.append("  Глобальный промпт: менеджер недоступен")
    if extra_line:
        lines.append(f"  {extra_line}")
    lines.append(
        "  Примечание: блок get_model_info — глобальное состояние llm-svc (что загружено в память)."
    )
    lines.append(
        "  Реальный ответ строится по полю model в POST /v1/chat/completions (см. лог generate_response выше);"
    )
    lines.append(
        "  при выбранном агенте туда подставляется модель из конструктора (llm-svc://id)."
    )
    lines.append("=" * 76)
    block = "\n".join(lines)
    print(block, flush=True)
    logger.info(block)


async def _resolve_agent_chat_params(agent_id_raw):
    """
    Модель и параметры из карточки агента (конструктор), а не глобальная загрузка на сервере.
    В запрос к llm-svc уходит явный model id (llm-svc://...) и max_tokens/temperature из model_settings.
    """
    empty = {"model_path": None, "max_tokens": None, "temperature": None, "system_prompt": None}
    if agent_id_raw is None:
        return empty
    try:
        aid = int(agent_id_raw)
    except (TypeError, ValueError):
        return empty
    try:
        from backend.database.init_db import get_agent_repository

        repo = get_agent_repository()
        if repo is None:
            return empty
        ag = await repo.get_agent(aid, None)
        if not ag:
            return empty
        cfg = ag.config if isinstance(ag.config, dict) else {}
        mp = str(cfg.get("model") or cfg.get("model_path") or "").strip()
        out = {**empty}
        if mp:
            low = mp.lower()
            # 1lm-svc:// вместо llm-svc:// — иначе клиент не подставляет id модели в запрос
            if low.startswith("1lm-svc://"):
                mp = "llm-svc://" + mp[10:]
                low = mp.lower()
            if low.startswith("llm-svc://"):
                out["model_path"] = mp
            elif "/" in mp or mp.lower().endswith(".gguf") or (len(mp) > 2 and mp[1] == ":"):
                out["model_path"] = mp
            else:
                out["model_path"] = f"llm-svc://{mp}"
        ms = cfg.get("model_settings")
        if isinstance(ms, dict):
            if ms.get("output_tokens") is not None:
                try:
                    out["max_tokens"] = int(ms["output_tokens"])
                except (TypeError, ValueError):
                    pass
            if ms.get("temperature") is not None:
                try:
                    out["temperature"] = float(ms["temperature"])
                except (TypeError, ValueError):
                    pass
        sp = (ag.system_prompt or "").strip()
        if sp:
            out["system_prompt"] = sp
        logger.info(
            f"[chat] agent_id={aid} → model_path={out['model_path']}, "
            f"max_tokens={out['max_tokens']}, temperature={out['temperature']}"
        )
        return out
    except Exception as ex:
        logger.warning(f"_resolve_agent_chat_params: {ex}")
        return empty


@sio.event
async def chat_message(sid, data):
    """Обработка сообщений чата через Socket.IO"""
    if not ask_agent or not save_dialog_entry:
        logger.error("AI services недоступны")
        await sio.emit('chat_error', {'error': 'AI services not available'}, room=sid)
        return
        
    try:
        user_message = data.get("message", "")
        streaming = data.get("streaming", True)
        logger.info(f"Socket.IO chat: {user_message[:50]}...")
        
        stop_generation_flags[sid] = False
        user_message_id = data.get("message_id", None)
        conversation_id = data.get("conversation_id", None)
        use_kb_rag = bool(data.get("use_kb_rag", False))
        use_memory_library_rag = bool(data.get("use_memory_library_rag", False))
        agent_profile = await _resolve_agent_chat_params(data.get("agent_id"))
        
        if conversation_id:
            import backend.database.memory_service as memory_service_module
            memory_service_module.current_conversation_id = conversation_id
        
        # Получаем историю 
        history = await get_recent_dialog_history(max_entries=memory_max_messages, conversation_id=conversation_id)
        
        # Сохраняем сообщение пользователя
        try:
            await save_dialog_entry("user", user_message, None, user_message_id, conversation_id)
        except RuntimeError as e:
            if "MongoDB" in str(e):
                logger.error(f"MongoDB недоступен: {e}")
                await sio.emit('chat_error', {'error': 'MongoDB недоступен.'}, room=sid)
                return
            raise
        
        orchestrator = get_agent_orchestrator()
        use_agent_mode = orchestrator and orchestrator.get_mode() == "agent"
        use_multi_llm_mode = orchestrator and orchestrator.get_mode() == "multi-llm"
        
        logger.info(f"[DEBUG] orchestrator: {orchestrator is not None}")
        logger.info(f"[DEBUG] use_agent_mode: {use_agent_mode}")
        logger.info(f"[DEBUG] use_multi_llm_mode: {use_multi_llm_mode}")
        
        # Функция для отправки чанков 
        async def async_stream_callback(chunk: str, accumulated_text: str):
            try:
                await sio.emit('chat_chunk', {'chunk': chunk, 'accumulated': accumulated_text}, room=sid)
            except: pass
        
        loop = asyncio.get_event_loop()
        def sync_stream_callback(chunk: str, accumulated_text: str):
            if stop_generation_flags.get(sid, False): return False
            asyncio.run_coroutine_threadsafe(async_stream_callback(chunk, accumulated_text), loop)
            return True
        
        # --- РЕЖИМ MULTI-LLM ---
        if use_multi_llm_mode:
            logger.info("РЕЖИМ MULTI-LLM: Параллельная генерация")
            multi_llm_models = orchestrator.get_multi_llm_models()
            if not multi_llm_models:
                await sio.emit('chat_error', {'error': 'Модели не выбраны'}, room=sid)
                return

            _terminal_chat_inference_banner(
                sid=sid,
                conversation_id=conversation_id,
                user_preview=user_message,
                mode_label=f"MULTI-LLM — модели: {', '.join(multi_llm_models)}",
                extra_line="Ниже для каждой модели — отдельный блок перед вызовом LLM.",
            )

            doc_context = None
            if rag_client:
                try:
                    strategy = current_rag_strategy if 'current_rag_strategy' in globals() else "auto"
                    hits = await rag_client.search(user_message, k=8, strategy=strategy)
                    if hits:
                        parts = []
                        total_len = 0
                        MAX_RAG_CONTEXT_CHARS = 12000
                        for i, (content, score, doc_id, chunk_idx) in enumerate(hits, 1):
                            frag = f"Фрагмент {i} (document_id={doc_id}, чанк {chunk_idx}, релевантность: {score:.2f}):\n{content}\n"
                            if total_len + len(frag) > MAX_RAG_CONTEXT_CHARS:
                                frag = frag[: max(0, MAX_RAG_CONTEXT_CHARS - total_len - 80)] + "\n... [обрезано]\n"
                                parts.append(frag)
                                break
                            parts.append(frag)
                            total_len += len(frag)
                        doc_context = "\n".join(parts)
                except Exception as e:
                    logger.error(f"Socket.IO: ошибка при получении контекста документов через SVC-RAG: {e}")
            
            final_user_message = user_message
            if doc_context:
                final_user_message = f"Контекст: {doc_context}\nВопрос: {user_message}"

            # KB search для multi-llm режима
            if use_kb_rag and rag_client:
                try:
                    kb_hits_multi = await rag_client.kb_search(user_message, k=8)
                    if kb_hits_multi:
                        kb_parts_multi = []
                        total_len = 0
                        MAX_KB_CONTEXT_CHARS = 10000
                        for i, (content, score, doc_id, chunk_idx) in enumerate(kb_hits_multi, 1):
                            frag = f"Фрагмент БЗ {i} (doc_id={doc_id}, чанк {chunk_idx}, релевантность: {score:.2f}):\n{content}\n"
                            if total_len + len(frag) > MAX_KB_CONTEXT_CHARS:
                                frag = frag[: max(0, MAX_KB_CONTEXT_CHARS - total_len - 60)] + "\n... [обрезано]\n"
                                kb_parts_multi.append(frag)
                                break
                            kb_parts_multi.append(frag)
                            total_len += len(frag)
                        kb_context_multi = "\n".join(kb_parts_multi)
                        final_user_message = f"База Знаний (постоянные документы):\n{kb_context_multi}\n\n" + final_user_message
                except Exception as e:
                    logger.error(f"Socket.IO multi-llm: ошибка поиска по Базе Знаний: {e}")

            if use_memory_library_rag and rag_client:
                try:
                    mem_hits = await rag_client.memory_rag_search(user_message, k=8)
                    if mem_hits:
                        mem_parts = []
                        total_len = 0
                        MAX_MEM = 10000
                        for i, (content, score, doc_id, chunk_idx) in enumerate(mem_hits, 1):
                            frag = (
                                f"Фрагмент библиотеки памяти {i} (doc_id={doc_id}, чанк {chunk_idx}, релевантность: {score:.2f}):\n{content}\n"
                            )
                            if total_len + len(frag) > MAX_MEM:
                                frag = frag[: max(0, MAX_MEM - total_len - 60)] + "\n... [обрезано]\n"
                                mem_parts.append(frag)
                                break
                            mem_parts.append(frag)
                            total_len += len(frag)
                        _mem_block = "\n".join(mem_parts)
                        final_user_message = (
                            f"Документы из настроек (библиотека памяти):\n{_mem_block}\n\n"
                            + final_user_message
                        )
                except Exception as e:
                    logger.error(f"Socket.IO multi-llm: ошибка memory_rag: {e}")

            loop = asyncio.get_running_loop()
            
            async def generate_single_model_response(model_name: str):
                try:
                    await sio.emit('multi_llm_start', {'model': model_name, 'models': multi_llm_models}, room=sid)
                    
                    # Логика путей для микросервисов 
                    # Определяем путь к модели
                    if model_name.startswith("llm-svc://"):
                        # Модель из llm-svc - не требует загрузки
                        model_path = model_name
                    else:
                        # Локальная модель - нужно загрузить её перед генерацией
                        model_path = os.path.join("models", model_name) if not os.path.isabs(model_name) else model_name
                        
                        # Загружаем модель для этого запроса
                        # Используем общую блокировку для предотвращения конфликтов при параллельной загрузке
                        with model_load_lock:
                            logger.info(f"Socket.IO: Загрузка модели {model_name} для генерации...")
                            if reload_model_by_path:
                                # Перезагружаем модель с блокировкой
                                success = reload_model_by_path(model_path)
                                if not success:
                                    logger.error(f"Socket.IO: Не удалось загрузить модель {model_name}")
                                    return {"model": model_name, "response": f"Ошибка: Не удалось загрузить модель {model_name}", "error": True}
                                
                                logger.info(f"Socket.IO: Модель {model_name} успешно загружена")
                                # Небольшая задержка после загрузки для стабилизации
                                import time
                                time.sleep(0.5)
                            else:
                                logger.warning(f"Socket.IO: Функция reload_model_by_path недоступна, используем текущую модель")

                    _terminal_chat_inference_banner(
                        sid=sid,
                        conversation_id=data.get("conversation_id"),
                        user_preview=final_user_message,
                        mode_label=f"MULTI-LLM — сейчас модель «{model_name}»",
                        model_path_for_call=model_path,
                    )

                    # Функция для отправки чанков от конкретной модели
                    def model_stream_callback(chunk: str, acc_text: str):
                        try:
                            # Планируем отправку чанка в event loop
                            asyncio.run_coroutine_threadsafe(
                                sio.emit('multi_llm_chunk', {
                                    'model': model_name,
                                    'chunk': chunk,
                                    'accumulated': acc_text
                                }, room=sid),
                                loop
                            )
                        except Exception as e:
                            logger.error(f"Socket.IO: Ошибка отправки чанка от модели {model_name}: {e}")
                        return True
                    
                    # Для режима multi-llm используем пустую историю
                    multi_llm_history = []
                    
                    # Генерируем ответ
                    response = None
                    if streaming:
                        # Потоковая генерация для каждой модели
                        import concurrent.futures
                        with concurrent.futures.ThreadPoolExecutor() as executor:
                            response = await asyncio.get_event_loop().run_in_executor(
                                executor,
                                ask_agent,
                                final_user_message,
                                multi_llm_history,
                                None,  # max_tokens
                                True,  # streaming
                                model_stream_callback,
                                model_path,
                                None   # custom_prompt_id
                            )
                    else:
                        # Обычная генерация
                        import concurrent.futures
                        with concurrent.futures.ThreadPoolExecutor() as executor:
                            response = await asyncio.get_event_loop().run_in_executor(
                                executor,
                                ask_agent,
                                final_user_message,
                                multi_llm_history,
                                None,  # max_tokens
                                False, # streaming
                                None,
                                model_path,
                                None
                            )
                    
                    # Проверяем, является ли ответ ошибкой
                    has_error = False
                    if response:
                        error_indicators = [
                            "Извините, произошла ошибка",
                            "llama_decode returned -1",
                            "Ошибка",
                            "error",
                            "Error"
                        ]
                        has_error = any(indicator.lower() in response.lower() for indicator in error_indicators)
                    
                    if has_error:
                        logger.warning(f"Socket.IO: Модель {model_name} вернула ошибку: {response[:100]}")
                        return {"model": model_name, "response": response, "error": True}
                    else:
                        return {"model": model_name, "response": response}

                except Exception as e:
                    logger.error(f"Socket.IO: Исключение при генерации от модели {model_name}: {e}")
                    import traceback
                    logger.error(traceback.format_exc())
                    return {"model": model_name, "response": f"Ошибка: {str(e)}", "error": True}
            
            # Запускаем параллельную генерацию от всех моделей
            tasks = [generate_single_model_response(model) for model in multi_llm_models]
            results = await asyncio.gather(*tasks, return_exceptions=True)
            
            # Отправляем результаты для каждой модели
            for i, result in enumerate(results):
                if isinstance(result, Exception):
                    logger.error(f"Socket.IO: Исключение при генерации: {result}")
                    await sio.emit('multi_llm_complete', {
                        'model': 'unknown',
                        'response': f'Ошибка: {str(result)}',
                        'error': True,
                        'index': i,
                        'total': len(multi_llm_models)
                    }, room=sid)
                else:
                    await sio.emit('multi_llm_complete', {
                        'model': result.get('model', 'unknown'),
                        'response': result.get('response', ''),
                        'error': result.get('error', False),
                        'index': i,
                        'total': len(multi_llm_models)
                    }, room=sid)
            
            logger.info("Socket.IO: Все ответы от моделей сгенерированы")
            return
        
        # Если включен агентный режим, используем агентную архитектуру
        if use_agent_mode:
            logger.info("Socket.IO: АГЕНТНАЯ АРХИТЕКТУРА: Переключение на агентный режим обработки")
            logger.info(f"Socket.IO: Запрос пользователя: '{user_message[:100]}{'...' if len(user_message) > 100 else ''}'")
            
            await sio.emit('chat_thinking', {
                'status': 'processing',
                'message': 'Обрабатываю запрос через агентную архитектуру...'
            }, room=sid)
            
            async def agent_stream_callback(chunk: str, accumulated_text: str):
                try:
                    if stop_generation_flags.get(sid, False):
                        return False
                    await sio.emit('chat_chunk', {'chunk': chunk, 'accumulated': accumulated_text}, room=sid)
                    return True
                except Exception as e:
                    logger.error(f"[agent_stream_callback] Ошибка: {e}")
                    return True
            
            context = {
                "history": history, "user_message": user_message,
                "selected_model": None, "socket_id": sid, "streaming": streaming,
            }
            
            try:
                from backend.tools.prompt_tools import set_tool_context
            except ModuleNotFoundError:
                from tools.prompt_tools import set_tool_context
            
            extended_context = context.copy()
            extended_context["sio"] = sio
            extended_context["socket_id"] = sid
            extended_context["stream_callback"] = agent_stream_callback if streaming else None
            extended_context["_main_event_loop"] = asyncio.get_running_loop()
            set_tool_context(extended_context)

            agent_effective_message = user_message
            if use_kb_rag and rag_client:
                try:
                    _kb = await rag_client.kb_search(user_message, k=6)
                    if _kb:
                        _parts, _tl = [], 0
                        for i, (c, s, did, ch) in enumerate(_kb, 1):
                            frag = f"БЗ {i} (doc={did}): {c}\n"
                            if _tl + len(frag) > 8000:
                                break
                            _parts.append(frag)
                            _tl += len(frag)
                        if _parts:
                            agent_effective_message = (
                                "База Знаний (документы):\n" + "\n".join(_parts) + "\n\n" + agent_effective_message
                            )
                except Exception as _e:
                    logger.error(f"Agent mode KB RAG: {_e}")
            if use_memory_library_rag and rag_client:
                try:
                    _mh = await rag_client.memory_rag_search(user_message, k=6)
                    if _mh:
                        _parts, _tl = [], 0
                        for i, (c, s, did, ch) in enumerate(_mh, 1):
                            frag = f"Библиотека памяти {i} (doc={did}): {c}\n"
                            if _tl + len(frag) > 8000:
                                break
                            _parts.append(frag)
                            _tl += len(frag)
                        if _parts:
                            agent_effective_message = (
                                "Документы из настроек (библиотека памяти):\n"
                                + "\n".join(_parts)
                                + "\n\n"
                                + agent_effective_message
                            )
                except Exception as _e:
                    logger.error(f"Agent mode memory_rag: {_e}")

            _terminal_chat_inference_banner(
                sid=sid,
                conversation_id=data.get("conversation_id"),
                user_preview=user_message,
                mode_label="Оркестратор агентов (agent architecture)",
                extra_line="Базовая модель на сервере — та, что ниже; оркестратор может дергать LLM несколько раз.",
            )

            try:
                response = await orchestrator.process_message(
                    agent_effective_message, history=history, context=context
                )
                logger.info(f"Socket.IO: АГЕНТНАЯ АРХИТЕКТУРА: Получен ответ, длина: {len(response) if response else 0}")
                
                if stop_generation_flags.get(sid, False):
                    stop_generation_flags[sid] = False
                    await sio.emit('generation_stopped', {'message': 'Генерация остановлена'}, room=sid)
                    return
                
                if response is None:
                    await sio.emit('chat_error', {'error': 'Не удалось получить ответ от агента'}, room=sid)
                    return
                
                await sio.emit('chat_complete', {
                    'response': response if response else "",
                    'timestamp': datetime.now().isoformat(),
                    'was_streaming': streaming
                }, room=sid)
            except Exception as orchestrator_error:
                logger.error(f"Socket.IO: Ошибка в оркестраторе: {orchestrator_error}")
                import traceback
                logger.error(traceback.format_exc())
                await sio.emit('chat_error', {'error': f"Ошибка выполнения: {str(orchestrator_error)}"}, room=sid)
                if sid in stop_generation_flags:
                    stop_generation_flags[sid] = False
                return
            
            try:
                conversation_id = data.get("conversation_id", None)
                await save_dialog_entry("assistant", response, None, None, conversation_id)
            except RuntimeError as e:
                if "MongoDB" in str(e):
                    logger.warning(f"MongoDB недоступен. Ответ не будет сохранен: {e}")
                else:
                    logger.error(f"Ошибка при сохранении ответа: {e}")
            return
        
        # Иначе используем прямой режим
        logger.info("Socket.IO: ПРЯМОЙ РЕЖИМ: Переключение на прямое общение с LLM")
        logger.info(f"Socket.IO: Запрос пользователя: '{user_message[:100]}{'...' if len(user_message) > 100 else ''}'")

        # ЛОГИКА ОБРАБОТКИ С ДОКУМЕНТАМИ (через SVC-RAG)
        # Если есть контекст RAG — отправляем только текст
        final_message = user_message
        images = None
        if rag_client:
            try:
                strategy = current_rag_strategy if 'current_rag_strategy' in globals() else "auto"
                hits = await rag_client.search(user_message, k=8, strategy=strategy)
                if hits:
                    # Для запросов про оглавление/структуру добавляем начало каждого документа (часто там оглавление)
                    if _is_structure_query(user_message):
                        seen = set()
                        start_chunks = []
                        for _c, _s, doc_id, chunk_idx in hits:
                            if doc_id is not None and (doc_id, chunk_idx) not in seen:
                                seen.add((doc_id, chunk_idx))
                        for doc_id in {d for _c, _s, d, _i in hits if d is not None}:
                            try:
                                first = await rag_client.get_document_start_chunks(doc_id, max_chunks=2)
                                for c, sc, did, idx in first:
                                    if (did, idx) not in seen:
                                        start_chunks.append((c, sc, did, idx))
                                        seen.add((did, idx))
                            except Exception:
                                pass
                        if start_chunks:
                            hits = start_chunks + hits
                    parts = []
                    total_len = 0
                    # Ограничиваем суммарный размер RAG-контекста, чтобы влезали история и лимит контекста модели
                    MAX_RAG_CONTEXT_CHARS = 12000
                    for i, (content, score, doc_id, chunk_idx) in enumerate(hits, 1):
                        frag = f"Фрагмент {i} (document_id={doc_id}, чанк {chunk_idx}, релевантность: {score:.2f}):\n{content}\n"
                        if total_len + len(frag) > MAX_RAG_CONTEXT_CHARS:
                            frag = frag[: max(0, MAX_RAG_CONTEXT_CHARS - total_len - 80)] + "\n... [обрезано по лимиту контекста]\n"
                            parts.append(frag)
                            total_len += len(frag)
                            break
                        parts.append(frag)
                        total_len += len(frag)
                    doc_context = "\n".join(parts)
                    final_message = (
                        f"Документы (RAG):\n{doc_context}\n"
                        f"Вопрос: {user_message}\n"
                        f"Ответь на основе этих документов. Перечисляй только то, что явно есть в фрагментах; не дублируй одни и те же пункты. Если полного оглавления в контексте нет — так и скажи."
                    )
            except Exception as e:
                logger.error(f"Socket.IO: ошибка при получении контекста документов через SVC-RAG: {e}")

        # База знаний + библиотека памяти: один проход поиска, контекст для LLM и трейс для UI
        document_search_trace = None
        kb_hits_ctx = []
        mem_hits_ctx = []
        if rag_client and (use_kb_rag or use_memory_library_rag):
            if use_kb_rag:
                try:
                    kb_hits_ctx = list(await rag_client.kb_search(user_message, k=8) or [])
                except Exception as e:
                    logger.error(f"Socket.IO: ошибка поиска по Базе Знаний: {e}")
            if use_memory_library_rag:
                try:
                    mem_hits_ctx = list(await rag_client.memory_rag_search(user_message, k=8) or [])
                except Exception as e:
                    logger.error(f"Socket.IO: ошибка memory_rag: {e}")
            kb_id_name = {}
            mem_id_name = {}
            try:
                if use_kb_rag and kb_hits_ctx:
                    for d in await rag_client.kb_list_documents():
                        kb_id_name[d["id"]] = d.get("filename") or str(d["id"])
            except Exception:
                pass
            try:
                if use_memory_library_rag and mem_hits_ctx:
                    for d in await rag_client.memory_rag_list_documents():
                        mem_id_name[d["id"]] = d.get("filename") or str(d["id"])
            except Exception:
                pass
            hits_out = []
            files_used = set()
            for content, score, doc_id, chunk_idx in kb_hits_ctx:
                if doc_id is None:
                    continue
                fn = kb_id_name.get(doc_id, f"doc_{doc_id}")
                files_used.add(fn)
                hits_out.append({
                    "file": fn,
                    "anchor": f"chunk@{chunk_idx}({fn})",
                    "relevance": round(float(score), 4),
                    "content": (content or "")[:12000],
                    "chunkIndex": chunk_idx,
                    "documentId": doc_id,
                    "store": "kb",
                })
            for content, score, doc_id, chunk_idx in mem_hits_ctx:
                if doc_id is None:
                    continue
                fn = mem_id_name.get(doc_id, f"doc_{doc_id}")
                files_used.add(fn)
                hits_out.append({
                    "file": fn,
                    "anchor": f"chunk@{chunk_idx}({fn})",
                    "relevance": round(float(score), 4),
                    "content": (content or "")[:12000],
                    "chunkIndex": chunk_idx,
                    "documentId": doc_id,
                    "store": "memory",
                })
            document_search_trace = {
                "query": user_message,
                "sourceFiles": sorted(files_used),
                "hits": hits_out,
            }

        if kb_hits_ctx:
            kb_parts = []
            total_len = 0
            MAX_KB_CONTEXT_CHARS = 10000
            for i, (content, score, doc_id, chunk_idx) in enumerate(kb_hits_ctx, 1):
                frag = f"Фрагмент БЗ {i} (doc_id={doc_id}, чанк {chunk_idx}, релевантность: {score:.2f}):\n{content}\n"
                if total_len + len(frag) > MAX_KB_CONTEXT_CHARS:
                    frag = frag[: max(0, MAX_KB_CONTEXT_CHARS - total_len - 60)] + "\n... [обрезано]\n"
                    kb_parts.append(frag)
                    break
                kb_parts.append(frag)
                total_len += len(frag)
            kb_context = "\n".join(kb_parts)
            final_message = f"База Знаний (постоянные документы):\n{kb_context}\n\n" + final_message
            logger.info(f"KB RAG: добавлен контекст ({len(kb_hits_ctx)} чанков)")

        if mem_hits_ctx:
            mem_parts = []
            total_len = 0
            MAX_MEM = 10000
            for i, (content, score, doc_id, chunk_idx) in enumerate(mem_hits_ctx, 1):
                frag = (
                    f"Фрагмент библиотеки памяти {i} (doc_id={doc_id}, чанк {chunk_idx}, релевантность: {score:.2f}):\n{content}\n"
                )
                if total_len + len(frag) > MAX_MEM:
                    frag = frag[: max(0, MAX_MEM - total_len - 60)] + "\n... [обрезано]\n"
                    mem_parts.append(frag)
                    break
                mem_parts.append(frag)
                total_len += len(frag)
            mem_ctx = "\n".join(mem_parts)
            final_message = f"Документы из настроек (библиотека памяти):\n{mem_ctx}\n\n" + final_message
            logger.info(f"memory_library RAG: добавлен контекст ({len(mem_hits_ctx)} чанков)")

        # Генерация ответа (при выбранном агенте — модель и параметры из БД, иначе глобальная)
        base_model_path = get_current_model_path()
        eff_model_path = agent_profile["model_path"] or base_model_path
        logger.info(
            f"Socket.IO: model_path для вызова LLM = {eff_model_path!r} "
            f"(из агента={'да' if agent_profile['model_path'] else 'нет'})"
        )
        _terminal_chat_inference_banner(
            sid=sid,
            conversation_id=conversation_id,
            user_preview=final_message,
            mode_label="Прямой чат с LLM (одна модель)"
            + (" — параметры из выбранного агента" if data.get("agent_id") else ""),
            model_path_for_call=eff_model_path,
            extra_line="RAG/KB уже учтены в final_message при необходимости.",
        )

        def _run_ask_agent(stream: bool, cb):
            return ask_agent(
                final_message,
                history=history,
                max_tokens=agent_profile["max_tokens"],
                streaming=stream,
                stream_callback=cb,
                model_path=eff_model_path,
                custom_prompt_id=None,
                images=images,
                system_prompt=agent_profile["system_prompt"],
                temperature=agent_profile["temperature"],
            )

        if streaming:
            import concurrent.futures
            with concurrent.futures.ThreadPoolExecutor() as executor:
                response = await asyncio.get_event_loop().run_in_executor(
                    executor, lambda: _run_ask_agent(True, sync_stream_callback)
                )
            logger.info(f"Socket.IO: получен потоковый ответ, длина: {len(response) if response else 0} символов")
            
            if response is None or stop_generation_flags.get(sid, False):
                stop_generation_flags[sid] = False
                await sio.emit('generation_stopped', {'message': 'Генерация остановлена'}, room=sid)
                return
        else:
            import concurrent.futures
            with concurrent.futures.ThreadPoolExecutor() as executor:
                response = await asyncio.get_event_loop().run_in_executor(
                    executor, lambda: _run_ask_agent(False, None)
                )
            logger.info(f"Socket.IO: получен ответ, длина: {len(response)} символов")
        
        if stop_generation_flags.get(sid, False):
            stop_generation_flags[sid] = False
            await sio.emit('generation_stopped', {'message': 'Генерация остановлена'}, room=sid)
            return
        
        try:
            conversation_id = data.get("conversation_id", None)
            _assist_meta = {"document_search": document_search_trace} if document_search_trace else None
            await save_dialog_entry("assistant", response, _assist_meta, None, conversation_id)
        except RuntimeError as e:
            if "MongoDB" in str(e):
                logger.warning(f"MongoDB недоступен. Ответ не будет сохранен: {e}")
            else:
                logger.error(f"Ошибка при сохранении ответа: {e}")
        
        if sid in stop_generation_flags:
            stop_generation_flags[sid] = False
        
        _complete_payload = {
            "response": response,
            "timestamp": datetime.now().isoformat(),
            "was_streaming": streaming,
        }
        if document_search_trace is not None:
            _complete_payload["document_search"] = document_search_trace
        await sio.emit("chat_complete", _complete_payload, room=sid)
        logger.info(f"Socket.IO: финальное сообщение отправлено (streaming={streaming}, response_len={len(response) if response else 0})")

    except Exception as e:
        logger.error(f"Socket.IO chat error: {e}")
        import traceback
        logger.error(traceback.format_exc())
        try:
            await sio.emit('chat_error', {
                'error': str(e)
            }, room=sid)
        except:
            logger.error("Не удалось отправить сообщение об ошибке клиенту")
    finally:
        # Финальная гарантия очистки флага
        if sid in stop_generation_flags:
            stop_generation_flags[sid] = False

# ================================
# МОДЕЛИ ДАННЫХ 
# ================================
from pydantic import BaseModel

class ChatMessage(BaseModel):
    message: str
    streaming: bool = True

class ModelSettings(BaseModel):
    context_size: int = 2048
    output_tokens: int = 512
    temperature: float = 0.7
    top_p: float = 0.95
    repeat_penalty: float = 1.05
    top_k: int = 40
    min_p: float = 0.05
    frequency_penalty: float = 0.0
    presence_penalty: float = 0.0
    use_gpu: bool = False
    streaming: bool = True

class VoiceSettings(BaseModel):
    voice_id: str = "ru"
    speech_rate: float = 1.0
    voice_speaker: str = "baya"

class MemorySettings(BaseModel):
    max_messages: int = 20
    include_system_prompts: bool = True
    clear_on_restart: bool = False

class ModelLoadRequest(BaseModel):
    model_path: str

class ModelLoadResponse(BaseModel):
    message: str
    success: bool

# ================================
# ОСНОВНЫЕ API ENDPOINTS
# ================================

@app.get("/")
async def root():
    """Главная страница API"""
    return {"message": "astrachat Web API", "status": "active", "version": "1.0.0"}

@app.get("/socket-test")
async def socket_test():
    """Тестовый endpoint для проверки Socket.IO (УЛУЧШЕНО: использование всех настроенных URL)"""
    return {
        "socketio_status": "active",
        "endpoint": "/socket.io/",
        # Берем все фронтенд порты из нашего нового конфига
        "cors_origins": [
            urls_config.get("frontend_port_1"), 
            urls_config.get("frontend_port_1_ipv4"),
            urls_config.get("frontend_port_2"),
            urls_config.get("frontend_port_3")
        ],
        "ping_timeout": 120,
        "ping_interval": 25
    }

@app.get("/health")
async def health_check():
    """Проверка состояния системы"""
    try:
        model_info = get_model_info() if get_model_info else {"loaded": False}
        vosk_status = check_vosk_model() if check_vosk_model else False
        
        # Проверяем доступность RAG-сервиса (SVC-RAG) по возможности
        rag_available = False
        if rag_client:
            try:
                health = await rag_client.health()
                rag_available = bool(health)
            except Exception:
                rag_available = False

        return {
            "status": "healthy",
            "timestamp": datetime.now().isoformat(),
            "services": {
                "llm_model": model_info.get("loaded", False),
                "vosk_model": vosk_status,
                "rag_service": rag_available,
                "transcriber": UniversalTranscriber is not None
            }
        }
    except Exception as e:
        return {
            "status": "unhealthy",
            "error": str(e),
            "timestamp": datetime.now().isoformat()
        }

# ================================
# ЧАТ API
# ================================

@app.post("/api/chat")
async def chat_with_ai(message: ChatMessage):
    """Отправить сообщение AI и получить ответ"""
    if not ask_agent:
        raise HTTPException(status_code=503, detail="AI agent не доступен")
    if not save_dialog_entry:
        raise HTTPException(status_code=503, detail="Memory service не доступен")
        
    try:
        logger.info(f"Chat request: {message.message[:50]}...")
        
        # Получаем историю диалога
        if get_recent_dialog_history:
            history = await get_recent_dialog_history(max_entries=memory_max_messages)
        else:
            history = []
        
        # Проверяем, доступна ли агентная архитектура
        orchestrator = get_agent_orchestrator()
        use_agent_mode = orchestrator and orchestrator.get_mode() == "agent"
        
        logger.info(f"DEBUG: orchestrator = {orchestrator is not None}")
        if orchestrator:
            logger.info(f"DEBUG: orchestrator.get_mode() = '{orchestrator.get_mode()}'")
        logger.info(f"DEBUG: use_agent_mode = {use_agent_mode}")
        
        if use_agent_mode:
            logger.info("АГЕНТНАЯ АРХИТЕКТУРА: Переключение на агентный режим обработки")
            logger.info(f"Запрос пользователя: '{message.message[:100]}{'...' if len(message.message) > 100 else ''}'")
            _terminal_chat_inference_banner(
                sid="HTTP-POST-/api/chat",
                conversation_id=None,
                user_preview=message.message,
                mode_label="REST /api/chat — оркестратор агентов",
            )
            # Используем агентную архитектуру
            context = {
                "history": history,
                "user_message": message.message,
            }
            response = await orchestrator.process_message(message.message, context)
            logger.info(f"АГЕНТНАЯ АРХИТЕКТУРА: Получен ответ, длина: {len(response)} символов")
        else:
            logger.info("ПРЯМОЙ РЕЖИМ: Переключение на прямое общение с LLM")
            logger.info(f"Запрос пользователя: '{message.message[:100]}{'...' if len(message.message) > 100 else ''}'")
            # Если RAG доступен, сначала пробуем получить контекст документов
            response = None
            if rag_client:
                try:
                    strategy = current_rag_strategy if 'current_rag_strategy' in globals() else "auto"
                    hits = await rag_client.search(message.message, k=12, strategy=strategy)
                    if hits:
                        if _is_structure_query(message.message):
                            seen = set()
                            start_chunks = []
                            for _c, _s, doc_id, chunk_idx in hits:
                                if doc_id is not None:
                                    seen.add((doc_id, chunk_idx))
                            for doc_id in {d for _c, _s, d, _i in hits if d is not None}:
                                try:
                                    first = await rag_client.get_document_start_chunks(doc_id, max_chunks=2)
                                    for c, sc, did, idx in first:
                                        if (did, idx) not in seen:
                                            start_chunks.append((c, sc, did, idx))
                                            seen.add((did, idx))
                                except Exception:
                                    pass
                            if start_chunks:
                                hits = start_chunks + hits
                        parts = []
                        total_len = 0
                        MAX_RAG_CONTEXT_CHARS = 12000
                        for i, (content, score, doc_id, chunk_idx) in enumerate(hits, 1):
                            frag = f"Фрагмент {i} (document_id={doc_id}, чанк {chunk_idx}, релевантность: {score:.2f}):\n{content}\n"
                            if total_len + len(frag) > MAX_RAG_CONTEXT_CHARS:
                                frag = frag[: max(0, MAX_RAG_CONTEXT_CHARS - total_len - 80)] + "\n... [обрезано]\n"
                                parts.append(frag)
                                break
                            parts.append(frag)
                            total_len += len(frag)
                        doc_context = "\n".join(parts)
                        prompt = f"""На основе предоставленного контекста из документов ответь на вопрос пользователя.
Если информации в контексте недостаточно, укажи это.
Отвечай только на основе информации из контекста. Не придумывай информацию.
Перечисляй только то, что явно есть во фрагментах; не дублируй одни и те же пункты.

Контекст из документов:

{doc_context}

Вопрос пользователя: {message.message}

Ответ:"""
                        current_model_path = get_current_model_path()
                        _terminal_chat_inference_banner(
                            sid="HTTP-POST-/api/chat",
                            conversation_id=None,
                            user_preview=prompt,
                            mode_label="REST /api/chat — ответ с RAG",
                            model_path_for_call=current_model_path,
                        )
                        response = ask_agent(
                            prompt,
                            history=[],
                            streaming=False,
                            model_path=current_model_path,
                        )
                except Exception as e:
                    logger.error(f"ПРЯМОЙ РЕЖИМ: ошибка при получении контекста документов через SVC-RAG: {e}")

            # Если RAG недоступен или не дал результатов - обычный запрос к модели
            if not response:
                logger.info("ПРЯМОЙ РЕЖИМ: Используем обычный AI agent без контекста документов")
                current_model_path = get_current_model_path()
                _terminal_chat_inference_banner(
                    sid="HTTP-POST-/api/chat",
                    conversation_id=None,
                    user_preview=message.message,
                    mode_label="REST /api/chat — прямой LLM (без RAG)",
                    model_path_for_call=current_model_path,
                )
                response = ask_agent(
                    message.message,
                    history=history,
                    streaming=False,  # Для REST API используем обычный режим
                    model_path=current_model_path
                )
            else:
                logger.info("ПРЯМОЙ РЕЖИМ: контекст документов недоступен, используем обычный AI agent")
                # Отправляем запрос к модели без контекста документов
                current_model_path = get_current_model_path()
                _terminal_chat_inference_banner(
                    sid="HTTP-POST-/api/chat",
                    conversation_id=None,
                    user_preview=message.message,
                    mode_label="REST /api/chat — прямой LLM (fallback)",
                    model_path_for_call=current_model_path,
                )
                response = ask_agent(
                    message.message,
                    history=history,
                    streaming=False,  # Для REST API используем обычный режим
                    model_path=current_model_path
                )
                logger.info(f"ПРЯМОЙ РЕЖИМ: Получен ответ от AI agent, длина: {len(response)} символов")
        
        # Сохраняем в память
        await save_dialog_entry("user", message.message)
        await save_dialog_entry("assistant", response)
        
        return {
            "response": response,
            "timestamp": datetime.now().isoformat(),
            "success": True
        }
        
    except Exception as e:
        logger.error(f"Chat error: {e}")
        raise HTTPException(status_code=500, detail=str(e))
@app.put("/api/messages/{conversation_id}/{message_id}")
async def update_message(conversation_id: str, message_id: str, request: Dict[str, str]):
    """Обновить содержимое сообщения в MongoDB"""
    try:
        if not get_conversation_repository:
            raise HTTPException(status_code=503, detail="MongoDB repository не доступен")
        
        conversation_repo = get_conversation_repository()
        content = request.get("content", "")
        old_content = request.get("old_content", None)  # Старое содержимое для поиска
        
        if not content:
            raise HTTPException(status_code=400, detail="Поле 'content' обязательно")
        
        success = await conversation_repo.update_message(conversation_id, message_id, content, old_content)
        
        if success:
            return {
                "message": "Сообщение обновлено",
                "success": True,
                "timestamp": datetime.now().isoformat()
            }
        else:
            raise HTTPException(status_code=404, detail="Сообщение не найдено")
            
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Ошибка при обновлении сообщения: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.websocket("/ws/chat")
async def websocket_chat(websocket: WebSocket):
    """WebSocket для потокового чата с AI"""
    if not ask_agent or not save_dialog_entry:
        await websocket.close(code=1008, reason="AI services not available")
        return
        
    await manager.connect(websocket)
    try:
        while True:
            # Получаем сообщение от клиента
            data = await websocket.receive_text()
            message_data = json.loads(data)
            
            # Получаем сообщение чата
            user_message = message_data.get("message", "")
            streaming = message_data.get("streaming", True)
            
            logger.info(f"WebSocket chat: {user_message[:50]}...")
            
            # Получаем историю
            if get_recent_dialog_history:
                history = await get_recent_dialog_history(max_entries=memory_max_messages)
            else:
                history = []
            
            # Сохраняем сообщение пользователя
            await save_dialog_entry("user", user_message)
            
            # Проверяем, доступна ли агентная архитектура
            orchestrator = get_agent_orchestrator()
            use_agent_mode = orchestrator and orchestrator.get_mode() == "agent"
            use_multi_llm_mode = orchestrator and orchestrator.get_mode() == "multi-llm"
            
            logger.info(f"WebSocket DEBUG: orchestrator = {orchestrator is not None}")
            if orchestrator:
                logger.info(f"WebSocket DEBUG: orchestrator.get_mode() = '{orchestrator.get_mode()}'")
            logger.info(f"WebSocket DEBUG: use_agent_mode = {use_agent_mode}, use_multi_llm_mode = {use_multi_llm_mode}")
            
            # Функция для отправки частей ответа 
            def stream_callback(chunk: str, accumulated_text: str):
                try:
                    logger.info(f"WebSocket: отправляем чанк, длина: {len(chunk)} символов, накоплено: {len(accumulated_text)} символов")
                    asyncio.create_task(websocket.send_text(json.dumps({
                        "type": "chunk",
                        "chunk": chunk,
                        "accumulated": accumulated_text
                    })))
                    logger.info("WebSocket: чанк успешно отправлен")
                    return True  
                except Exception as e:
                    logger.error(f"WebSocket: ошибка отправки чанка: {e}")
                    return False
            
            try:
                # Если включен режим multi-llm, генерируем ответы от нескольких моделей параллельно
                if use_multi_llm_mode:
                    logger.info("WebSocket: РЕЖИМ MULTI-LLM: Параллельная генерация")
                    logger.info(f"WebSocket: Запрос пользователя: '{user_message[:100]}{'...' if len(user_message) > 100 else ''}'")
                    
                    multi_llm_models = orchestrator.get_multi_llm_models()
                    if not multi_llm_models:
                        logger.warning("WebSocket: Режим multi-llm активирован, но модели не выбраны")
                        await websocket.send_text(json.dumps({
                            "type": "error",
                            "error": "Режим multi-llm активирован, но модели не выбраны"
                        }))
                        continue
                    
                    logger.info(f"WebSocket: Генерируем ответы от моделей: {multi_llm_models}")
                    
                    # Проверяем документы через SVC-RAG
                    doc_context = None
                    if rag_client:
                        try:
                            strategy = current_rag_strategy if 'current_rag_strategy' in globals() else "auto"
                            hits = await rag_client.search(user_message, k=8, strategy=strategy)
                            if hits:
                                parts = []
                                total_len = 0
                                MAX_RAG_CONTEXT_CHARS = 12000
                                for i, (content, score, doc_id, chunk_idx) in enumerate(hits, 1):
                                    frag = f"Фрагмент {i} (document_id={doc_id}, чанк {chunk_idx}, релевантность: {score:.2f}):\n{content}\n"
                                    if total_len + len(frag) > MAX_RAG_CONTEXT_CHARS:
                                        frag = frag[: max(0, MAX_RAG_CONTEXT_CHARS - total_len - 80)] + "\n... [обрезано]\n"
                                        parts.append(frag)
                                        break
                                    parts.append(frag)
                                    total_len += len(frag)
                                doc_context = "\n".join(parts)
                                logger.info(f"WebSocket: Получен контекст документов для multi-llm через SVC-RAG, длина: {len(doc_context)} символов")
                        except Exception as e:
                            logger.error(f"WebSocket: Ошибка при получении контекста документов через SVC-RAG: {e}")
            
                    # Формируем финальное сообщение с контекстом
                    final_user_message = user_message
                    if doc_context:
                        final_user_message = f"""Контекст из загруженных документов:
                        {doc_context}
                        Вопрос пользователя: {user_message}
                        Пожалуйста, ответьте на вопрос пользователя, используя информацию из предоставленных документов. Если в документах нет информации для ответа, честно скажите об этом."""
                    
                    # Функция для генерации ответа от одной модели
                    async def generate_single_model_response(model_name: str):
                        try:
                            logger.info(f"WebSocket: Начинаем генерацию от модели: {model_name}")
                            
                            await websocket.send_text(json.dumps({
                                "type": "multi_llm_start",
                                "model": model_name,
                                "total_models": len(multi_llm_models),
                                "models": multi_llm_models
                            }))
                            
                            # Определяем путь к модели 
                            if model_name.startswith("llm-svc://"):
                                model_path = model_name
                            else:
                                model_path = os.path.join("models", model_name) if not os.path.isabs(model_name) else model_name
                            
                            # Для режима multi-llm используем пустую историю 
                            multi_llm_history = []
                            
                            # Генерируем ответ 
                            if streaming:
                                accumulated_text = ""
                                def model_stream_callback(chunk: str, acc_text: str):
                                    nonlocal accumulated_text
                                    accumulated_text = acc_text
                                    try:
                                        asyncio.create_task(websocket.send_text(json.dumps({
                                            "type": "multi_llm_chunk",
                                            "model": model_name,
                                            "chunk": chunk,
                                            "accumulated": acc_text
                                        })))
                                    except Exception as e:
                                        logger.error(f"WebSocket: Ошибка отправки чанка от модели {model_name}: {e}")
                                    return True
                                
                                response = ask_agent(
                                    final_user_message,
                                    history=multi_llm_history,
                                    streaming=True,
                                    stream_callback=model_stream_callback,
                                    model_path=model_path
                                )
                                return {"model": model_name, "response": accumulated_text}
                            else:
                                response = ask_agent(
                                    final_user_message,
                                    history=multi_llm_history,
                                    streaming=False,
                                    model_path=model_path
                                )
                                return {"model": model_name, "response": response}
                        except Exception as e:
                            logger.error(f"WebSocket: Исключение при генерации от модели {model_name}: {e}")
                            return {"model": model_name, "response": f"Ошибка: {str(e)}", "error": True}
                    
                    # Запуск параллельной генерации
                    tasks = [generate_single_model_response(model) for model in multi_llm_models]
                    results = await asyncio.gather(*tasks, return_exceptions=True)
                    
                    # Отправляем результаты
                    for result in results:
                        if isinstance(result, Exception):
                            await websocket.send_text(json.dumps({
                                "type": "multi_llm_complete", "model": "unknown", "response": f"Ошибка: {str(result)}", "error": True
                            }))
                        else:
                            await websocket.send_text(json.dumps({
                                "type": "multi_llm_complete", "model": result.get("model", "unknown"), 
                                "response": result.get("response", ""), "error": result.get("error", False)
                            }))
                    
                    logger.info("WebSocket: Все ответы от моделей сгенерированы")
                    continue
                
                # --- ЛОГИКА АГЕНТНОЙ АРХИТЕКТУРЫ (Начало) ---
                if use_agent_mode:
                    # Обычная генерация
                        response = ask_agent(
                            user_message,
                            history=history,
                            streaming=False,
                            model_path=get_current_model_path()
                        )
                        logger.info(f"WebSocket: получен ответ от AI agent, длина: {len(response)} символов")
                
                # Сохраняем ответ
                await save_dialog_entry("assistant", response)
                
                # Отправляем финальное сообщение
                await websocket.send_text(json.dumps({
                    "type": "complete",
                    "response": response,
                    "timestamp": datetime.now().isoformat()
                }))
                
            except Exception as e:
                await websocket.send_text(json.dumps({
                    "type": "error",
                    "error": str(e)
                }))
                
    except WebSocketDisconnect:
        logger.info("WebSocket отключен клиентом - нормальное отключение")
        try:
            manager.disconnect(websocket)
        except Exception as e:
            logger.warning(f"Ошибка при отключении WebSocket в менеджере: {e}")
    except Exception as e:
        logger.error(f"WebSocket error: {e}")
        manager.disconnect(websocket)

async def process_audio_data(websocket: WebSocket, data: bytes):
    """Обработка аудио данных от WebSocket клиента"""
    import tempfile
    temp_dir = tempfile.gettempdir()
    audio_object_name = None
    audio_file = None
    
    # Проверяем флаг остановки голосового чата
    if globals().get('voice_chat_stop_flag', False):
        logger.info("Обработка аудио данных остановлена - установлен флаг остановки")
        return
    
    logger.info(f"Начинаю обработку аудио данных размером {len(data)} байт")
    
    try:
        # Проверяем, что получили действительно аудио данные
        if len(data) < 100:  
            logger.warning(f"Получены данные слишком маленького размера: {len(data)} байт")
            await websocket.send_text(json.dumps({
                "type": "error",
                "error": "Получены некорректные аудио данные"
            }))
            return
        
        # Определяем реальный формат аудио по сигнатуре файла
        if data[:4] == b'RIFF' and b'WAVE' in data[:12]:
            audio_ext = ".wav"
            audio_content_type = "audio/wav"
        elif data[:4] == b'\x1a\x45\xdf\xa3':  # EBML/WebM (MediaRecorder default)
            audio_ext = ".webm"
            audio_content_type = "audio/webm"
        elif data[:4] == b'OggS':
            audio_ext = ".ogg"
            audio_content_type = "audio/ogg"
        elif data[:3] == b'ID3' or data[:2] == b'\xff\xfb':
            audio_ext = ".mp3"
            audio_content_type = "audio/mpeg"
        else:
            audio_ext = ".webm"  # Default: браузеры обычно шлют webm
            audio_content_type = "audio/webm"
        
        logger.info(f"Определён формат аудио: {audio_ext} ({audio_content_type})")
        
        # Сохраняем файл в MinIO или локально
        if minio_client:
            try:
                audio_object_name = minio_client.generate_object_name(prefix="voice_", extension=audio_ext)
                minio_client.upload_file(data, audio_object_name, content_type=audio_content_type)
                audio_file = minio_client.get_file_path(audio_object_name)
                logger.info(f"Аудио файл загружен в MinIO: {audio_object_name}")
            except Exception as e:
                logger.warning(f"Ошибка загрузки в MinIO, используем локальный файл: {e}")
                audio_file = os.path.join(temp_dir, f"voice_{datetime.now().timestamp()}{audio_ext}")
                with open(audio_file, "wb") as f:
                    f.write(data)
        else:
            audio_file = os.path.join(temp_dir, f"voice_{datetime.now().timestamp()}{audio_ext}")
            with open(audio_file, "wb") as f:
                f.write(data)
        
        # Распознаем речь 
        logger.info(f"Обрабатываю аудио файл: {audio_file}")
        
        if not recognize_speech_from_file:
            logger.warning("recognize_speech_from_file функция не доступна")
            await websocket.send_text(json.dumps({
                "type": "error",
                "error": "Модуль распознавания речи недоступен."
            }))
            return
            
        # Запускаем STT в executor чтобы не блокировать event loop
        loop = asyncio.get_event_loop()
        recognized_text = await loop.run_in_executor(
            None, lambda: recognize_speech_from_file(audio_file)
        )
        logger.info(f"РАСПОЗНАННЫЙ ТЕКСТ: '{recognized_text}'")
        
        if recognized_text and recognized_text.strip():
            await websocket.send_text(json.dumps({
                "type": "speech_recognized",
                "text": recognized_text,
                "timestamp": datetime.now().isoformat()
            }))
            
            # Получаем ответ от AI
            if get_recent_dialog_history:
                history = await get_recent_dialog_history(max_entries=memory_max_messages)
            else:
                history = []
            
            try:
                current_model_path = get_current_model_path()
                voice_system_prompt = (
                    "Ты - голосовой AI-ассистент AstraChat. "
                    "ВСЕГДА отвечай на русском языке, даже если пользователь написал на другом языке. "
                    "Давай краткие, понятные ответы, подходящие для голосового общения. "
                    "Не используй emoji, markdown-разметку, специальные символы или код в ответах. "
                    "Отвечай как в устном разговоре — просто и по существу."
                )
                # Запускаем LLM в executor чтобы не блокировать event loop
                ai_response = await loop.run_in_executor(
                    None, lambda: ask_agent(recognized_text, history=history, streaming=False, model_path=current_model_path, system_prompt=voice_system_prompt)
                )
                logger.info(f"ОТВЕТ ОТ LLM: '{ai_response[:100]}...'")
            except Exception as ai_error:
                logger.error(f"Ошибка обращения к AI: {ai_error}")
                await websocket.send_text(json.dumps({
                    "type": "speech_error",
                    "error": f"Ошибка AI модуля: {str(ai_error)}"
                }))
                return
            
            await save_dialog_entry("user", recognized_text)
            await save_dialog_entry("assistant", ai_response)
            
            await websocket.send_text(json.dumps({
                "type": "ai_response",
                "text": ai_response,
                "timestamp": datetime.now().isoformat()
            }))
            
            # Синтезируем речь
            # Запускаем в executor чтобы не блокировать event loop
            speech_file = os.path.join(temp_dir, f"speech_{datetime.now().timestamp()}.wav")
            try:
                loop = asyncio.get_event_loop()
                speech_ok = await loop.run_in_executor(
                    None, lambda: speak_text(ai_response, speaker='baya', voice_id='ru', save_to_file=speech_file)
                )
                if speech_ok:
                    if os.path.exists(speech_file) and os.path.getsize(speech_file) > 44:
                        with open(speech_file, "rb") as f:
                            audio_data = f.read()
                        await websocket.send_bytes(audio_data)
                        try: os.remove(speech_file)
                        except: pass
                else:
                    await websocket.send_text(json.dumps({"type": "tts_error", "error": "Ошибка синтеза речи"}))
            except Exception as tts_err:
                logger.error(f"Ошибка TTS в voice WebSocket: {tts_err}")
                await websocket.send_text(json.dumps({"type": "tts_error", "error": str(tts_err)}))
        else:
            await websocket.send_text(json.dumps({"type": "speech_error", "error": "Речь не распознана"}))
            
    except Exception as e:
        logger.error(f"Ошибка обработки аудио: {e}")
    finally:
        try:
            if audio_file and os.path.exists(audio_file): os.remove(audio_file)
        except: pass
@app.websocket("/ws/voice")
async def websocket_voice(websocket: WebSocket):
    """WebSocket для голосового чата в реальном времени"""
    
    # Принимаем соединение в любом случае
    await manager.connect(websocket)
    
    # Проверяем доступность сервисов после подключения
    if not ask_agent or not save_dialog_entry:
        logger.warning("AI services недоступны для WebSocket /ws/voice")
        try:
            await websocket.send_text(json.dumps({
                "type": "error",
                "error": "AI сервисы недоступны. Проверьте настройки модели."
            }))
        except Exception as e:
            logger.warning(f"Не удалось отправить сообщение об ошибке: {e}")
        
    try:
        while True:
            # Получаем сообщение через низкоуровневый receive() - он не теряет данные
            raw_message = await websocket.receive()
            
            if raw_message.get("type") == "websocket.disconnect":
                logger.info("WebSocket отключен клиентом (receive)")
                break
            
            if "text" in raw_message:
                # Текстовое сообщение (JSON команда)
                message = raw_message["text"]
                logger.info(f"Получено текстовое сообщение: {message[:100]}...")
                
                try:
                    data = json.loads(message)
                    
                    if data.get("type") == "start_listening":
                        logger.info("Получена команда start_listening")
                        await websocket.send_text(json.dumps({
                            "type": "listening_started",
                            "message": "Готов к приему голоса"
                        }))
                    elif data.get("type") == "stop_processing":
                        logger.info("Получена команда stop_processing")
                        globals()['voice_chat_stop_flag'] = True
                        await websocket.send_text(json.dumps({
                            "type": "processing_stopped",
                            "message": "Обработка остановлена"
                        }))
                    elif data.get("type") == "reset_processing":
                        logger.info("Получена команда reset_processing")
                        globals()['voice_chat_stop_flag'] = False
                        await websocket.send_text(json.dumps({
                            "type": "processing_reset",
                            "message": "Обработка возобновлена"
                        }))
                    else:
                        logger.warning(f"Неизвестный тип сообщения: {data.get('type', 'unknown')}")
                        
                except json.JSONDecodeError as e:
                    logger.error(f"Ошибка парсинга JSON: {e}")
                    
            elif "bytes" in raw_message:
                # Бинарные данные (аудио)
                data = raw_message["bytes"]
                logger.info(f"Получены аудио данные размером: {len(data)} байт")
                
                try:
                    await process_audio_data(websocket, data)
                except Exception as process_error:
                    logger.error(f"Ошибка обработки аудио данных: {process_error}")
                    try:
                        await websocket.send_text(json.dumps({
                            "type": "error",
                            "error": f"Ошибка обработки аудио: {str(process_error)}"
                        }))
                    except: pass
                    
    except WebSocketDisconnect:
        logger.info("WebSocket отключен клиентом")
        try: manager.disconnect(websocket)
        except: pass
    except Exception as e:
        logger.error(f"Voice WebSocket error: {e}")
        import traceback
        logger.error(f"Full traceback: {traceback.format_exc()}")
        try:
            await websocket.send_text(json.dumps({"type": "error", "error": f"Временная ошибка: {str(e)}"}))
        except: pass

# ================================
# ИСТОРИЯ ДИАЛОГОВ
# ================================

@app.get("/api/history")
async def get_chat_history(limit: int = None):
    """Получить историю диалогов"""
    if limit is None:
        limit = memory_max_messages if 'memory_max_messages' in globals() else 20
    
    if not get_recent_dialog_history:
        logger.error("memory_service недоступен")
        raise HTTPException(status_code=503, detail="Memory service недоступен")
    
    try:
        history = await get_recent_dialog_history(max_entries=limit)
        return {
            "history": history,
            "count": len(history),
            "max_messages": memory_max_messages,
            "timestamp": datetime.now().isoformat(),
            "source": "memory_service"
        }
    except Exception as e:
        logger.error(f"Ошибка получения истории: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/api/history")
async def clear_chat_history():
    """Очистить историю диалогов"""
    if not clear_dialog_history:
        raise HTTPException(status_code=503, detail="Memory service недоступен")
    try:
        result = await clear_dialog_history()
        return {"message": "История очищена", "success": True, "source": "memory_service"}
    except Exception as e:
        logger.error(f"Ошибка очистки истории: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# ================================
# УПРАВЛЕНИЕ МОДЕЛЯМИ
# ================================

@app.get("/api/models/current")
async def get_current_model():
    """Получить информацию о текущей модели"""
    if get_model_info:
        try:
            result = get_model_info()
            if result and 'path' in result:
                save_app_settings({
                    'current_model_path': result['path'],
                    'current_model_name': result.get('name', 'Unknown'),
                    'current_model_status': result.get('status', 'loaded')
                })
            return result
        except Exception as e:
            logger.error(f"Ошибка получения информации о модели: {e}")
    
    # Fallback к сохраненным настройкам
    try:
        app_settings = load_app_settings()
        current_model_path = app_settings.get('current_model_path')
        if current_model_path and os.path.exists(current_model_path):
            file_size = os.path.getsize(current_model_path)
            return {
                "name": app_settings.get('current_model_name', os.path.basename(current_model_path)),
                "path": current_model_path,
                "status": "loaded_from_settings",
                "size": file_size,
                "size_mb": round(file_size / (1024 * 1024), 2),
                "type": "gguf"
            }
    except: pass
    
    return {"name": "Модель не загружена", "path": "", "status": "not_loaded"}

@app.get("/api/models")
async def get_models():
    """Получить список доступных моделей"""
    return await get_available_models()

@app.get("/api/models/available")
async def get_available_models():
    """Получить список доступных моделей"""
    try:
        # Проверяем, используется ли llm-svc 
        use_llm_svc = os.getenv('USE_LLM_SVC', 'false').lower() == 'true'
        
        if use_llm_svc:
            logger.info("[Backend] Запрос списка моделей через llm-svc")
            try:
                from backend.llm_client import get_llm_service
                service = await get_llm_service()
                models_data = await service.client.get_models()
                
                models = []
                for model_data in models_data:
                    models.append({
                        "name": model_data.get("id", "Unknown"),
                        "path": f"llm-svc://{model_data.get('id', 'unknown')}",
                        "size": model_data.get("size", 0),
                        "size_mb": model_data.get("size_mb", 0),
                        "object": model_data.get("object", "model"),
                        "owned_by": model_data.get("owned_by", "llm-svc")
                    })
                return {"models": models}
            except Exception as e:
                logger.error("")
                logger.error("=" * 100)
                logger.error("[Backend] ОШИБКА ПОЛУЧЕНИЯ СПИСКА МОДЕЛЕЙ ИЗ LLM-SVC")
                logger.error("=" * 100)
                logger.error(f"Ошибка: {str(e)}")
                logger.error(f"Тип ошибки: {type(e).__name__}")
                logger.error("")
                logger.error("Возможные причины:")
                logger.error("1. Контейнер llm-svc не запущен или еще загружает модель")
                logger.error("2. Проблемы с сетью Docker между контейнерами")
                logger.error("3. llm-svc не отвечает на запросы (модель еще загружается)")
                logger.error("")
                logger.error("=" * 100)
                return {"models": [], "error": str(e), "warning": "llm-svc недоступен"}
        else:
            # Локальный режим
            models_dir = "models"
            if not os.path.exists(models_dir): return {"models": []}
            models = []
            for file in os.listdir(models_dir):
                if file.endswith('.gguf'):
                    file_path = os.path.join(models_dir, file)
                    size = os.path.getsize(file_path)
                    models.append({
                        "name": file, "path": file_path, "size": size, "size_mb": round(size / (1024 * 1024), 2)
                    })
            return {"models": models}
    except Exception as e:
        logger.error(f"Ошибка получения списка моделей: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# ================================
# УПРАВЛЕНИЕ МОДЕЛЯМИ
# ================================

@app.post("/api/models/load")
async def load_model(request: ModelLoadRequest):
    """Загрузить модель по указанному пути (теперь поддерживает llm-svc://)"""
    if not reload_model_by_path:
        logger.warning("reload_model_by_path функция не доступна")
        return ModelLoadResponse(
            message="Функция загрузки модели недоступна. Проверьте инициализацию AI agent.", 
            success=False
        )
    
    try:
        logger.info(f"Загружаю модель: {request.model_path}")
        
        # Проверяем, что путь не является директорией
        if os.path.isdir(request.model_path):
            logger.error(f"Передан путь к директории вместо файла модели: {request.model_path}")
            return ModelLoadResponse(
                message=f"Ошибка: передан путь к директории вместо файла модели: {request.model_path}",
                success=False
            )
        
        # Функция reload_model_by_path (в agent_llm_svc.py) теперь умеет работать по сети
        success = reload_model_by_path(request.model_path)
        if success:
            logger.info(f"Модель успешно загружена: {request.model_path}")
            
            # Сохраняем информацию о загруженной модели
            if request.model_path.startswith("llm-svc://"):
                model_name = request.model_path.replace("llm-svc://", "")
            else:
                model_name = os.path.basename(request.model_path)
            
            save_app_settings({
                'current_model_path': request.model_path,
                'current_model_name': model_name,
                'current_model_status': 'loaded'
            })
            
            return ModelLoadResponse(message="Модель успешно загружена", success=True)
        else:
            logger.error(f"Не удалось загрузить модель: {request.model_path}")
            return ModelLoadResponse(message="Не удалось загрузить модель", success=False)
    except Exception as e:
        logger.error(f"Ошибка загрузки модели: {e}")
        return ModelLoadResponse(message=f"Ошибка загрузки модели: {str(e)}", success=False)

@app.get("/api/models/settings")
async def get_model_settings():
    """Получить настройки модели"""
    if not model_settings:
        logger.warning("model_settings не доступен, возвращаю дефолтные настройки")
        return {
            "context_size": 2048, "output_tokens": 512, "temperature": 0.7,
            "top_p": 0.95, "repeat_penalty": 1.05, "top_k": 40, "min_p": 0.05,
            "frequency_penalty": 0.0, "presence_penalty": 0.0, "use_gpu": False,
            "streaming": True, "streaming_speed": 50
        }
    try:
        result = model_settings.get_all()
        logger.info(f"Настройки модели: {result}")
        return result
    except Exception as e:
        logger.error(f"Ошибка получения настроек модели: {e}")
        return {
            "context_size": 2048, "output_tokens": 512, "temperature": 0.7,
            "top_p": 0.95, "repeat_penalty": 1.05, "top_k": 40, "min_p": 0.05,
            "frequency_penalty": 0.0, "presence_penalty": 0.0, "use_gpu": False,
            "streaming": True, "streaming_speed": 50
        }

@app.put("/api/models/settings")
async def update_model_settings_api(settings_data: ModelSettings):
    """Обновить настройки модели """
    if not update_model_settings:
        raise HTTPException(status_code=503, detail="AI agent не доступен")
    try:
        success = update_model_settings(settings_data.dict())
        if success:
            return {"message": "Настройки обновлены", "success": True}
        else:
            raise HTTPException(status_code=400, detail="Не удалось обновить настройки")
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/models/settings/reset")
async def reset_model_settings():
    """Сбросить настройки модели к рекомендуемым значениям"""
    if not model_settings:
        raise HTTPException(status_code=503, detail="AI agent не доступен")
    try:
        model_settings.reset_to_defaults()
        return {
            "message": "Настройки сброшены к рекомендуемым значениям", 
            "success": True,
            "settings": model_settings.get_all()
        }
    except Exception as e:
        logger.error(f"Ошибка сброса настроек модели: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/models/settings/recommended")
async def get_recommended_settings():
    """Получить рекомендуемые настройки модели """
    if not model_settings:
        raise HTTPException(status_code=503, detail="AI agent не доступен")
    try:
        return {
            "recommended": model_settings.get_recommended_settings(),
            "max_values": model_settings.get_max_values()
        }
    except Exception as e:
        logger.error(f"Ошибка получения рекомендуемых настроек: {e}")
        raise HTTPException(status_code=500, detail=str(e))
    
# ================================
# ГОЛОСОВЫЕ ФУНКЦИИ
# ================================

class VoiceSynthesizeRequest(BaseModel):
    text: str
    voice_id: str = "ru"
    voice_speaker: str = "baya"
    speech_rate: float = 1.0

class TranscriptionSettings(BaseModel):
    engine: str = "whisperx"  # whisperx или vosk
    language: str = "ru"
    auto_detect: bool = True

class YouTubeTranscribeRequest(BaseModel):
    url: str

class DocumentQueryRequest(BaseModel):
    query: str

class RAGSettings(BaseModel):
    strategy: str = "auto"  # auto, reranking, hierarchical, hybrid, standard

@app.post("/api/voice/synthesize")
async def synthesize_speech(request: VoiceSynthesizeRequest):
    """Синтезировать речь из текста"""
    if not speak_text:
        logger.warning("speak_text функция не доступна")
        raise HTTPException(status_code=503, detail="Модуль синтеза речи недоступен.")
    
    import tempfile
    temp_dir = tempfile.gettempdir()
    audio_file = os.path.join(temp_dir, f"speech_{datetime.now().timestamp()}.wav")
    
    try:
        logger.info(f"Синтезирую речь: '{request.text[:100]}...'")
        
        # speak_text (из voice.py) теперь умеет ходить в микросервис
        # Запускаем в executor чтобы не блокировать event loop
        loop = asyncio.get_event_loop()
        success = await loop.run_in_executor(
            None, lambda: speak_text(
                text=request.text, 
                speaker=request.voice_speaker, 
                voice_id=request.voice_id, 
                speech_rate=request.speech_rate,
                save_to_file=audio_file
            )
        )
        
        if success and os.path.exists(audio_file):
            logger.info(f"Аудиофайл создан: {audio_file}, размер: {os.path.getsize(audio_file)} байт")
            
            # Сохраняем в MinIO, если доступен
            if minio_client:
                try:
                    with open(audio_file, "rb") as f:
                        audio_data = f.read()
                    audio_object_name = minio_client.generate_object_name(prefix="speech_", extension=".wav")
                    minio_client.upload_file(audio_data, audio_object_name, content_type="audio/wav")
                except Exception as e:
                    logger.warning(f"Ошибка сохранения в MinIO: {e}")
            
            # Создаем временную копию
            temp_copy = os.path.join(temp_dir, f"speech_copy_{datetime.now().timestamp()}.wav")
            import shutil
            shutil.copy2(audio_file, temp_copy)
            
            async def cleanup_temp_file():
                try:
                    if os.path.exists(temp_copy): os.remove(temp_copy)
                except: pass
            
            return FileResponse(
                temp_copy, 
                media_type="audio/wav", 
                filename="speech.wav", 
                background=cleanup_temp_file
            )
        else:
            raise HTTPException(status_code=500, detail="Не удалось создать аудиофайл")
            
    except Exception as e:
        logger.error(f"Ошибка синтеза речи: {e}")
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        try:
            if os.path.exists(audio_file): os.remove(audio_file)
        except: pass

@app.post("/api/voice/recognize")
async def recognize_speech_api(audio_file: UploadFile = File(...)):
    """Распознать речь из аудиофайла"""
    if not recognize_speech_from_file:
        return {"text": "", "success": False, "error": "Модуль распознавания речи недоступен."}
    
    import tempfile
    temp_dir = tempfile.gettempdir()
    file_path = None
    file_object_name = None
    
    try:
        content = await audio_file.read()
        logger.info(f"Получен аудиофайл: {audio_file.filename}, размер: {len(content)} байт")
        
        if minio_client:
            try:
                file_object_name = minio_client.generate_object_name(prefix="audio_", extension=".wav")
                minio_client.upload_file(content, file_object_name, content_type="audio/wav")
                file_path = minio_client.get_file_path(file_object_name)
            except Exception as e:
                logger.warning(f"Ошибка MinIO: {e}")
                file_path = os.path.join(temp_dir, f"audio_{datetime.now().timestamp()}.wav")
                with open(file_path, "wb") as f: f.write(content)
        else:
            file_path = os.path.join(temp_dir, f"audio_{datetime.now().timestamp()}.wav")
            with open(file_path, "wb") as f: f.write(content)
        
        # recognize_speech_from_file (из voice.py) теперь умеет ходить в микросервис
        text = recognize_speech_from_file(file_path)
        logger.info(f"Распознанный текст: '{text}'")
        
        return {
            "text": text,
            "success": True,
            "timestamp": datetime.now().isoformat()
        }
        
    except Exception as e:
        logger.error(f"Ошибка распознавания речи: {e}")
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        try:
            if file_path and os.path.exists(file_path): os.remove(file_path)
            if minio_client and file_object_name:
                try: minio_client.delete_file(file_object_name)
                except: pass
        except: pass

@app.get("/api/voice/settings")
async def get_voice_settings():
    return {"voice_id": "ru", "speech_rate": 1.0, "voice_speaker": "baya"}

@app.put("/api/voice/settings")
async def update_voice_settings(settings_data: VoiceSettings):
    return {"message": "Настройки обновлены", "success": True, "settings": settings_data.dict()}

@app.get("/api/transcription/settings")
async def get_transcription_settings():
    global current_transcription_engine, current_transcription_language
    return {
        "engine": current_transcription_engine, 
        "language": current_transcription_language, 
        "auto_detect": True
    }

@app.put("/api/transcription/settings")
async def update_transcription_settings(settings_data: TranscriptionSettings):
    global current_transcription_engine, current_transcription_language, transcriber
    try:
        if settings_data.engine:
            current_transcription_engine = settings_data.engine.lower()
            if transcriber and hasattr(transcriber, 'switch_engine'):
                transcriber.switch_engine(current_transcription_engine)
        
        if settings_data.language:
            current_transcription_language = settings_data.language
            if transcriber and hasattr(transcriber, 'set_language'):
                transcriber.set_language(current_transcription_language)
        
        save_app_settings({
            'transcription_engine': current_transcription_engine, 
            'transcription_language': current_transcription_language
        })
        return {"message": "Настройки обновлены", "success": True}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# ================================
# НАСТРОЙКИ ПАМЯТИ 
# ================================

@app.get("/api/memory/settings")
async def get_memory_settings():
    global memory_max_messages, memory_include_system_prompts, memory_clear_on_restart
    return {
        "max_messages": memory_max_messages, 
        "include_system_prompts": memory_include_system_prompts, 
        "clear_on_restart": memory_clear_on_restart
    }

@app.put("/api/memory/settings")
async def update_memory_settings(settings_data: MemorySettings):
    global memory_max_messages, memory_include_system_prompts, memory_clear_on_restart
    try:
        memory_max_messages = settings_data.max_messages
        memory_include_system_prompts = settings_data.include_system_prompts
        memory_clear_on_restart = settings_data.clear_on_restart
        save_app_settings({
            'memory_max_messages': memory_max_messages, 
            'memory_include_system_prompts': memory_include_system_prompts, 
            'memory_clear_on_restart': memory_clear_on_restart
        })
        return {"message": "Настройки памяти обновлены", "success": True}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/memory/status")
async def get_memory_status():
    try:
        if not get_recent_dialog_history:
            raise HTTPException(status_code=503, detail="Memory service не доступен")
        history = await get_recent_dialog_history(max_entries=memory_max_messages)
        return {
            "message_count": len(history), 
            "max_messages": memory_max_messages, 
            "success": True
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/memory/clear")
async def clear_memory():
    """Очистить память"""
    try:
        if not clear_dialog_history:
            raise HTTPException(status_code=503, detail="Memory service не доступен")
        
        result = await clear_dialog_history()
        logger.info(f"Память очищена: {result}")
        
        return {
            "message": "Память успешно очищена",
            "success": True,
            "result": result
        }
        
    except Exception as e:
        logger.error(f"Ошибка очистки памяти: {e}")
        raise HTTPException(status_code=500, detail=f"Ошибка очистки памяти: {str(e)}")

# ================================
# РАБОТА С ДОКУМЕНТАМИ 
# ================================

@app.post("/api/documents/upload")
async def upload_document(file: UploadFile = File(...)):
    """Загрузить и обработать документ (через SVC-RAG)."""
    logger.info(f"=== Загрузка документа: {file.filename} ===")

    if not rag_client:
        logger.error("RAG service client не доступен")
        raise HTTPException(status_code=503, detail="RAG service недоступен")

    file_object_name = None
    documents_bucket = os.getenv('MINIO_DOCUMENTS_BUCKET_NAME', 'astrachat-documents')
        
    try:
        # Читаем содержимое файла в память
        content = await file.read()
        logger.info(f"Файл получен, размер: {len(content)} байт")
        
        # Определяем тип файла
        file_extension = os.path.splitext(file.filename)[1].lower() if file.filename else ""
        is_image = file_extension in ['.jpg', '.jpeg', '.png', '.webp']
        
        # Определяем content_type
        content_type_map = {
            '.pdf': 'application/pdf',
            '.docx': 'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
            '.doc': 'application/msword',
            '.txt': 'text/plain',
            '.xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            '.xls': 'application/vnd.ms-excel',
            '.jpg': 'image/jpeg',
            '.jpeg': 'image/jpeg',
            '.png': 'image/png',
            '.webp': 'image/webp'
        }
        content_type = content_type_map.get(file_extension, 'application/octet-stream')
        
        # Сохраняем в MinIO или локально
        if minio_client:
            try:
                # Генерируем имя объекта
                file_object_name = minio_client.generate_object_name(
                    prefix="doc_" if not is_image else "img_",
                    extension=file_extension
                )
                
                # Загружаем в MinIO в bucket для документов
                minio_client.upload_file(
                    content, 
                    file_object_name, 
                    content_type=content_type,
                    bucket_name=documents_bucket
                )
                logger.info(f"Документ загружен в MinIO: {documents_bucket}/{file_object_name}")
            except Exception as e:
                logger.warning(f"Ошибка загрузки в MinIO: {e}")
                # Продолжаем обработку даже если не удалось загрузить в MinIO
                file_object_name = None
        
        # Отправляем документ в SVC-RAG для индексации
        logger.info("Отправляем документ в SVC-RAG для индексации...")
        try:
            rag_result = await rag_client.upload_document(
                file_bytes=content,
                filename=file.filename or file_object_name or "unknown",
                minio_object=file_object_name,
                minio_bucket=documents_bucket if minio_client and file_object_name else None,
                original_path=None,
            )
        except Exception as e:
            logger.error(f"Ошибка при обращении к SVC-RAG: {e}")
            # В случае ошибки удаляем файл из MinIO, если он был загружен
            if minio_client and file_object_name:
                try:
                    minio_client.delete_file(file_object_name, bucket_name=documents_bucket)
                    logger.info(f"Файл удален из MinIO после ошибки: {documents_bucket}/{file_object_name}")
                except Exception as e2:
                    logger.warning(f"Ошибка удаления файла из MinIO: {e2}")
            raise HTTPException(status_code=502, detail=f"Ошибка RAG-сервиса: {e}")

        if not rag_result.get("ok"):
            detail = rag_result.get("error", "Ошибка индексации документа в RAG")
            # В случае ошибки удаляем файл из MinIO, если он был загружен
            if minio_client and file_object_name:
                try:
                    minio_client.delete_file(file_object_name, bucket_name=documents_bucket)
                    logger.info(f"Файл удален из MinIO после ошибки RAG: {documents_bucket}/{file_object_name}")
                except Exception as e:
                    logger.warning(f"Ошибка удаления файла из MinIO: {e}")
            raise HTTPException(status_code=400, detail=detail)

        logger.info("Обработка завершена, файл хранится в MinIO и проиндексирован в SVC-RAG")

        result = {
            "message": "Документ успешно загружен и обработан",
            "filename": file.filename,
            "success": True,
            "rag_document_id": rag_result.get("document_id"),
        }

        if is_image and minio_client and file_object_name:
            result["minio_object"] = file_object_name
            result["minio_bucket"] = documents_bucket

        return result
            
    except Exception as e:
        logger.error(f"Ошибка при загрузке документа: {str(e)}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/documents/query")
async def query_document(request: DocumentQueryRequest):
    """Задать вопрос по загруженному документу"""
    logger.info(f"=== Запрос к документам: {request.query[:50]}... ===")

    if not rag_client:
        logger.error("RAG service client не доступен")
        raise HTTPException(status_code=503, detail="RAG service недоступен")

    try:
        if not ask_agent:
            logger.error("AI agent не доступен")
            raise HTTPException(status_code=503, detail="AI agent не доступен")

        # Выполняем поиск через SVC-RAG
        # Используем текущую стратегию RAG 
        strategy = current_rag_strategy if 'current_rag_strategy' in globals() else "auto"
        logger.info(f"Выполняем поиск в SVC-RAG: strategy={strategy}")
        hits = await rag_client.search(
            query=request.query,
            k=12,
            strategy=strategy,
        )
        logger.info(f"SVC-RAG вернул {len(hits)} фрагментов")

        if not hits:
            response_text = "В загруженных документах не найдено информации по вашему запросу."
        else:
            if _is_structure_query(request.query):
                seen = set()
                start_chunks = []
                for _c, _s, doc_id, chunk_idx in hits:
                    if doc_id is not None:
                        seen.add((doc_id, chunk_idx))
                for doc_id in {d for _c, _s, d, _i in hits if d is not None}:
                    try:
                        first = await rag_client.get_document_start_chunks(doc_id, max_chunks=2)
                        for c, sc, did, idx in first:
                            if (did, idx) not in seen:
                                start_chunks.append((c, sc, did, idx))
                                seen.add((did, idx))
                    except Exception:
                        pass
                if start_chunks:
                    hits = start_chunks + hits
            # Формируем контекст из фрагментов
            context_parts = []
            total_len = 0
            MAX_RAG_CONTEXT_CHARS = 12000
            for i, (content, score, doc_id, chunk_idx) in enumerate(hits, 1):
                frag = f"Фрагмент {i} (document_id={doc_id}, чанк {chunk_idx}, релевантность: {score:.2f}):\n{content}\n"
                if total_len + len(frag) > MAX_RAG_CONTEXT_CHARS:
                    frag = frag[: max(0, MAX_RAG_CONTEXT_CHARS - total_len - 80)] + "\n... [обрезано]\n"
                    context_parts.append(frag)
                    break
                context_parts.append(frag)
                total_len += len(frag)
            doc_context = "\n".join(context_parts)

            prompt = f"""На основе предоставленного контекста из документов ответь на вопрос пользователя.
Если информации в контексте недостаточно, укажи это.
Отвечай только на основе информации из контекста. Не придумывай информацию.
Перечисляй только то, что явно есть во фрагментах; не дублируй одни и те же пункты.

Контекст из документов:

{doc_context}

Вопрос пользователя: {request.query}

Ответ:"""

            logger.info("Отправляем запрос к LLM с контекстом документов (через SVC-RAG)...")
            response_text = ask_agent(prompt)
            logger.info(f"Получен ответ от LLM, длина: {len(response_text) if response_text else 0} символов")

        return {
            "response": response_text,
            "query": request.query,
            "success": True,
            "timestamp": datetime.now().isoformat()
        }
    except Exception as e:
        logger.error(f"Ошибка при запросе к документам: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/documents")
async def get_documents():
    """Получить список загруженных документов"""
    logger.info("=== Получение списка документов ===")

    if not rag_client:
        logger.error("RAG service client не доступен")
        raise HTTPException(status_code=503, detail="RAG service недоступен")

    try:
        docs = await rag_client.list_documents()
        filenames = [d.get("filename") for d in docs]
        logger.info(f"Список документов (SVC-RAG): {filenames}")

        return {
            "documents": filenames,
            "count": len(filenames),
            "success": True
        }
    except Exception as e:
        logger.error(f"Ошибка при получении списка документов: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/api/documents/{filename}")
async def delete_document(filename: str):
    """Удалить документ по имени файла"""
    logger.info(f"Удаление документа: {filename} ===")

    if not rag_client:
        logger.error("RAG service client не доступен")
        raise HTTPException(status_code=503, detail="RAG service недоступен")

    try:
        # Получаем список документов из SVC-RAG и проверяем наличие
        docs = await rag_client.list_documents()
        filenames = [d.get("filename") for d in docs]
        logger.info(f"Доступные документы до удаления: {filenames}")

        if filename not in filenames:
            logger.warning(f"Документ {filename} не найден")
            raise HTTPException(status_code=404, detail=f"Документ {filename} не найден")

        # Удаляем файл из MinIO, если он там хранится
        documents_bucket = os.getenv('MINIO_DOCUMENTS_BUCKET_NAME', 'astrachat-documents')
        if minio_client:
            try:
                minio_info = await rag_client.get_image_minio_info(filename)
            except Exception as e:
                logger.warning(f"Не удалось получить информацию о MinIO из SVC-RAG: {e}")
                minio_info = None
            if minio_info:
                try:
                    minio_client.delete_file(
                        minio_info["minio_object"],
                        bucket_name=minio_info["minio_bucket"]
                    )
                    logger.info(f"Файл удален из MinIO: {minio_info['minio_bucket']}/{minio_info['minio_object']}")
                except Exception as e:
                    logger.warning(f"Ошибка удаления файла из MinIO: {e}")
        
        # Удаляем документ из RAG
        try:
            await rag_client.delete_document_by_filename(filename)
        except Exception as e:
            logger.error(f"Ошибка удаления документа из SVC-RAG: {e}")
            raise HTTPException(status_code=502, detail=f"Ошибка RAG-сервиса при удалении: {e}")

        # Получаем обновленный список документов
        new_docs = await rag_client.list_documents()
        new_filenames = [d.get("filename") for d in new_docs]
        logger.info(f"Документы после удаления: {new_filenames}")

        return {
            "message": f"Документ {filename} успешно удален",
            "success": True,
            "remaining_documents": new_filenames
        }
            
    except Exception as e:
        logger.error(f"Ошибка при удалении документа: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/documents/report/generate")
async def generate_confidence_report():
    """Сгенерировать отчет об уверенности модели в распознанном тексте"""
    logger.info("=== Генерация отчета об уверенности ===")

    if not rag_client:
        logger.error("RAG service client не доступен")
        raise HTTPException(status_code=503, detail="RAG service недоступен")

    try:
        # Получаем данные для отчета из SVC-RAG
        report_data = await rag_client.get_confidence_report()
        logger.info(f"Получены данные отчета: {report_data['total_documents']} документов")
        
        # Формируем текстовый отчет
        report_text = f"""
        ОТЧЕТ О СТЕПЕНИ УВЕРЕННОСТИ МОДЕЛИ В РАСПОЗНАННОМ ТЕКСТЕ
        {'=' * 80}
        Дата генерации: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
        {'=' * 80}
        
        ОБЩАЯ ИНФОРМАЦИЯ:
        - Всего обработано документов: {report_data['total_documents']}
        - Средняя уверенность модели: {report_data['average_confidence']:.2f}%
        - Всего слов: {report_data.get('total_words', 0)}
        {'=' * 80}
        ДЕТАЛЬНАЯ ИНФОРМАЦИЯ ПО ДОКУМЕНТАМ:
        """
        
        # Добавляем распознанный текст с процентами над словами
        for i, doc in enumerate(report_data['documents'], 1):
            report_text += f"""

{i}. {doc['filename']}
   Тип файла: {doc['file_type']}
   Уверенность модели: {doc['confidence']:.2f}%
   Длина распознанного текста: {doc['text_length']} символов
   Количество слов: {doc.get('words_count', 0)}
   {'-' * 80}
   
   РАСПОЗНАННЫЙ ТЕКСТ С УВЕРЕННОСТЬЮ:
"""
            
            # Находим соответствующий отформатированный текст
            formatted_text_info = next((ft for ft in report_data.get('formatted_texts', []) if ft['filename'] == doc['filename']), None)
            
            if formatted_text_info and formatted_text_info.get('words'):
                # Форматируем текст с процентами над словами
                words = formatted_text_info.get('words', [])
                if words:
                    # Группируем слова по строкам для лучшей читаемости
                    line_words = []
                    current_line = []
                    
                    for word_info in words:
                        word = word_info.get('word', '')
                        conf = word_info.get('confidence', 0.0)
                        current_line.append((word, conf))
                        
                        # Каждые 8-10 слов или при достижении символов новой строки
                        if len(current_line) >= 8:
                            line_words.append(current_line)
                            current_line = []
                    
                    if current_line:
                        line_words.append(current_line)
                    
                    # Формируем текст с процентами над словами в красивом формате
                    if line_words:
                        for line in line_words:
                            # Используем табличный формат с фиксированной шириной колонок
                            import re
                            tokens_data = []
                            prev_is_punctuation = False
                            
                            for word, conf in line:
                                is_punctuation = bool(re.match(r'^[^\w\s]+$', word))
                                
                                # Вычисляем ширину колонки на основе длины слова
                                word_width = len(word)
                                # Минимальная ширина колонки - 10 символов (для процента и пробелов)
                                col_width = max(word_width + 2, 10)
                                
                                tokens_data.append({
                                    'word': word,
                                    'conf': conf,
                                    'is_punctuation': is_punctuation,
                                    'col_width': col_width,
                                    'needs_space_before': not prev_is_punctuation and not is_punctuation and tokens_data
                                })
                                prev_is_punctuation = is_punctuation
                            
                            # Формируем строки с выравниванием в табличном формате
                            percent_line = "│"
                            word_line = "│"
                            separator_line = "├"
                            
                            for i, token in enumerate(tokens_data):
                                if token['needs_space_before']:
                                    # Добавляем разделитель между словами
                                    word_line += "│"
                                    percent_line += "│"  # Вертикальный разделитель
                                    separator_line += "┼"
                                
                                # Форматируем процент и слово в колонке
                                percent_str = f"{token['conf']:.0f}%"
                                word_str = token['word']
                                
                                # Выравниваем процент по центру колонки
                                percent_padded = percent_str.center(token['col_width'])
                                # Выравниваем слово по левому краю колонки
                                word_padded = word_str.ljust(token['col_width'])
                                
                                percent_line += percent_padded + "│"
                                word_line += word_padded + "│"
                                separator_line += "─" * token['col_width'] + ("┤" if i == len(tokens_data) - 1 else "┼")
                            
                            # Добавляем в отчет с красивым форматированием
                            report_text += f"{percent_line}\n"
                            report_text += f"{separator_line}\n"
                            report_text += f"{word_line}\n\n"
                    else:
                        report_text += "[Нет валидных слов для отображения]\n"
                else:
                    report_text += "[Нет данных о словах]\n"
            else:
                report_text += "[Нет отформатированного текста]\n"
            
            report_text += f"{'-' * 80}\n"
        
        # Итоговый процент уверенности
        overall_conf = report_data.get('overall_confidence', report_data.get('average_confidence', 0.0))
        
        report_text += f"""

{'=' * 80}
ИТОГО:
- Итоговая уверенность по всему распознанному тексту: {overall_conf:.2f}%
- Средняя уверенность по документам: {report_data['average_confidence']:.2f}%
- Всего документов: {report_data['total_documents']}
- Всего слов: {report_data.get('total_words', 0)}
{'=' * 80}
"""
        
        # Создаем JSON отчет
        report_json = {
            "generated_at": datetime.now().isoformat(),
            "summary": {
                "total_documents": report_data['total_documents'],
                "average_confidence": round(report_data['average_confidence'], 2),
                "overall_confidence": round(overall_conf, 2),
                "total_words": report_data.get('total_words', 0)
            },
            "documents": report_data['documents']
        }
        
        return {
            "success": True,
            "report_text": report_text,
            "report_json": report_json,
            "summary": {
                "total_documents": report_data['total_documents'],
                "average_confidence": round(report_data['average_confidence'], 2),
                "overall_confidence": round(overall_conf, 2),
                "total_words": report_data.get('total_words', 0)
            }
        }
        
    except Exception as e:
        logger.error(f"Ошибка при генерации отчета: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/documents/report/download")
async def download_confidence_report():
    """Скачать отчет об уверенности в виде Excel файла"""
    logger.info("=== Скачивание отчета об уверенности (Excel) ===")
    
    try:
        # Получаем данные для отчета из SVC-RAG
        if not rag_client:
            logger.error("RAG service client не доступен")
            raise HTTPException(status_code=503, detail="RAG service недоступен")
        report_data = await rag_client.get_confidence_report()
        logger.info(f"Получены данные отчета: {report_data['total_documents']} документов")
        
        # Создаем Excel файл
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Отчет об уверенности"
        
        # Стили для заголовков
        header_font = Font(bold=True, size=14, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Стили для подзаголовков
        subheader_font = Font(bold=True, size=12)
        subheader_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
        
        # Стили для процентов уверенности
        high_confidence_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Зеленый для высокой уверенности
        medium_confidence_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Желтый для средней
        low_confidence_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Красный для низкой
        
        # Стили для границ
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        current_row = 1
        
        # Заголовок отчета
        ws.merge_cells(f'A{current_row}:D{current_row}')
        header_cell = ws[f'A{current_row}']
        header_cell.value = "ОТЧЕТ О СТЕПЕНИ УВЕРЕННОСТИ МОДЕЛИ В РАСПОЗНАННОМ ТЕКСТЕ"
        header_cell.font = header_font
        header_cell.fill = header_fill
        header_cell.alignment = header_alignment
        header_cell.border = thin_border
        current_row += 1
        
        # Дата генерации
        ws.merge_cells(f'A{current_row}:D{current_row}')
        date_cell = ws[f'A{current_row}']
        date_cell.value = f"Дата генерации: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        date_cell.alignment = Alignment(horizontal="center")
        current_row += 2
        
        # Общая информация
        if report_data['total_documents'] == 0:
            ws.merge_cells(f'A{current_row}:D{current_row}')
            warning_cell = ws[f'A{current_row}']
            warning_cell.value = "ВНИМАНИЕ: Нет обработанных документов для формирования отчета."
            warning_cell.font = Font(bold=True, color="FF0000")
            warning_cell.alignment = Alignment(horizontal="center")
            current_row += 1
        else:
            # Общая информация
            info_row = current_row
            ws[f'A{info_row}'] = "ОБЩАЯ ИНФОРМАЦИЯ:"
            ws[f'A{info_row}'].font = subheader_font
            ws[f'A{info_row}'].fill = subheader_fill
            current_row += 1
            
            ws[f'A{current_row}'] = "Всего обработано документов:"
            ws[f'B{current_row}'] = report_data['total_documents']
            current_row += 1
            
            ws[f'A{current_row}'] = "Средняя уверенность модели:"
            ws[f'B{current_row}'] = f"{report_data['average_confidence']:.2f}%"
            current_row += 1
            ws[f'A{current_row}'] = "Всего слов:"
            ws[f'B{current_row}'] = report_data.get('total_words', 0)
            current_row += 2
            
            # Детальная информация по документам 
            for doc_idx, doc in enumerate(report_data.get('documents', []), 1):
                # Заголовок документа
                doc_start_row = current_row
                ws.merge_cells(f'A{current_row}:D{current_row}')
                doc_header = ws[f'A{current_row}']
                doc_header.value = f"{doc_idx}. {doc.get('filename', 'Неизвестный файл')}"
                doc_header.font = subheader_font
                doc_header.fill = subheader_fill
                doc_header.border = thin_border
                current_row += 1
                
                # Информация о документе 
                ws[f'A{current_row}'] = "Тип файла:"
                ws[f'B{current_row}'] = doc.get('file_type', 'unknown')
                current_row += 1
                
                ws[f'A{current_row}'] = "Уверенность модели:"
                conf_value = doc.get('confidence', 0.0)
                ws[f'B{current_row}'] = f"{conf_value:.2f}%"
                if conf_value >= 80:
                    ws[f'B{current_row}'].fill = high_confidence_fill
                elif conf_value >= 50:
                    ws[f'B{current_row}'].fill = medium_confidence_fill
                else:
                    ws[f'B{current_row}'].fill = low_confidence_fill
                current_row += 1
                
                ws[f'A{current_row}'] = "Длина текста:"
                ws[f'B{current_row}'] = f"{doc.get('text_length', 0)} символов"
                current_row += 1
                
                ws[f'A{current_row}'] = "Количество слов:"
                ws[f'B{current_row}'] = doc.get('words_count', 0)
                current_row += 2
                
                # Распознанный текст с уверенностью 
                formatted_text_info = next((ft for ft in report_data.get('formatted_texts', []) if ft.get('filename') == doc.get('filename')), None)
                
                if formatted_text_info and formatted_text_info.get('words'):
                    words = formatted_text_info.get('words', [])
                    if words:
                        ws[f'A{current_row}'] = "Слово"
                        ws[f'B{current_row}'] = "Уверенность"
                        ws[f'A{current_row}'].font = Font(bold=True)
                        ws[f'B{current_row}'].font = Font(bold=True)
                        ws[f'A{current_row}'].fill = subheader_fill
                        ws[f'B{current_row}'].fill = subheader_fill
                        ws[f'A{current_row}'].border = thin_border
                        ws[f'B{current_row}'].border = thin_border
                        current_row += 1
                        
                        for word_info in words:
                            word = word_info.get('word', '')
                            conf = word_info.get('confidence', 0.0)
                            
                            if word:  
                                ws[f'A{current_row}'] = word
                                ws[f'B{current_row}'] = f"{conf:.1f}%"
                                ws[f'A{current_row}'].border = thin_border
                                ws[f'B{current_row}'].border = thin_border
                                
                                if conf >= 80:
                                    ws[f'B{current_row}'].fill = high_confidence_fill
                                elif conf >= 50:
                                    ws[f'B{current_row}'].fill = medium_confidence_fill
                                else:
                                    ws[f'B{current_row}'].fill = low_confidence_fill
                                current_row += 1
                current_row += 1
            
            # Итоговая информация 
            overall_conf = report_data.get('overall_confidence', report_data.get('average_confidence', 0.0))
            ws.merge_cells(f'A{current_row}:D{current_row}')
            summary_header = ws[f'A{current_row}']
            summary_header.value = "ИТОГО"
            summary_header.font = subheader_font
            summary_header.fill = subheader_fill
            summary_header.border = thin_border
            current_row += 1
            
            ws[f'A{current_row}'] = "Итоговая уверенность по всему тексту:"
            ws[f'B{current_row}'] = f"{overall_conf:.2f}%"
            if overall_conf >= 80:
                ws[f'B{current_row}'].fill = high_confidence_fill
            elif overall_conf >= 50:
                ws[f'B{current_row}'].fill = medium_confidence_fill
            else:
                ws[f'B{current_row}'].fill = low_confidence_fill
            current_row += 1
            
            ws[f'A{current_row}'] = "Средняя уверенность по документам:"
            ws[f'B{current_row}'] = f"{report_data['average_confidence']:.2f}%"
            current_row += 1
            
            ws[f'A{current_row}'] = "Всего документов:"
            ws[f'B{current_row}'] = report_data['total_documents']
            current_row += 1
            
            ws[f'A{current_row}'] = "Всего слов:"
            ws[f'B{current_row}'] = report_data.get('total_words', 0)
        
        ws.column_dimensions['A'].width = 50
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        
        import tempfile
        temp_dir = tempfile.gettempdir()
        report_filename = f"confidence_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        report_path = os.path.join(temp_dir, report_filename)
        
        try:
            os.makedirs(temp_dir, exist_ok=True)
            wb.save(report_path)
            logger.info(f"Excel отчет сохранен: {report_path}")
            
            return FileResponse(
                report_path,
                media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                filename=report_filename,
                headers={"Content-Disposition": f"attachment; filename*=UTF-8''{report_filename}"}
            )
        except Exception as file_err:
            logger.error(f"Ошибка при сохранении Excel: {file_err}")
            raise HTTPException(status_code=500, detail=str(file_err))
        
    except HTTPException: raise
    except Exception as e:
        logger.error(f"Ошибка при генерации Excel: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))

# ================================
# RAG НАСТРОЙКИ
# ================================

@app.get("/api/rag/settings")
async def get_rag_settings():
    """Получить настройки RAG и информацию о методе (через стратегию и конфиг SVC-RAG)."""
    global current_rag_strategy
    applied_method = "unknown"
    method_description = ""

    user_strategy = current_rag_strategy
    if user_strategy == "auto":
        applied_method = "auto"
        method_description = (
            "Автоматический выбор между reranking, иерархическим, гибридным и стандартным поиском "
            "в зависимости от настроек RAG-сервиса."
        )
    elif user_strategy == "reranking":
        applied_method = "reranking"
        method_description = "Reranking (переранжирование результатов поиска) включён."
    elif user_strategy == "hierarchical":
        applied_method = "hierarchical"
        method_description = "Иерархический поиск по многоуровневым суммаризациям документа."
    elif user_strategy == "hybrid":
        applied_method = "hybrid"
        method_description = "Гибридный поиск: комбинирует векторный поиск и BM25."
    elif user_strategy == "standard":
        applied_method = "standard"
        method_description = "Стандартный векторный поиск по чанкам документов."

    return {"strategy": current_rag_strategy, "applied_method": applied_method, "method_description": method_description}

@app.put("/api/rag/settings")
async def update_rag_settings(settings_data: RAGSettings):
    """Обновить настройки RAG """
    global current_rag_strategy
    valid_strategies = ["auto", "reranking", "hierarchical", "hybrid", "standard"]
    if settings_data.strategy not in valid_strategies:
        raise HTTPException(status_code=400, detail=f"Недопустимая стратегия.")
    
    try:
        current_rag_strategy = settings_data.strategy
        logger.info(f"[RAG SETTINGS] Стратегия изменена на: '{current_rag_strategy}'")
        save_app_settings({'rag_strategy': current_rag_strategy})
        return {"message": "Настройки RAG обновлены", "success": True, "strategy": current_rag_strategy}
    except Exception as e:
        logger.error(f"Ошибка обновления RAG: {e}")
        raise HTTPException(status_code=500, detail=str(e))
# ================================
# ТРАНСКРИБАЦИЯ
# ================================

@app.post("/api/transcribe/upload")
async def transcribe_file(
    file: UploadFile = File(...),
    request_id: Optional[str] = Form(None)
):
    """Транскрибировать аудио/видео файл с диаризацией по ролям"""
    import uuid
    
    # Получаем request_id из формы или генерируем новый
    transcription_id = request_id if request_id else str(uuid.uuid4())
    logger.info(f"=== Начало транскрибации файла с диаризацией: {file.filename}, ID: {transcription_id} ===")
    
    if not transcriber:
        logger.error("Transcriber не доступен")
        raise HTTPException(status_code=503, detail="Transcriber не доступен")
    
    # Сбрасываем флаг остановки для этого ID транскрибации
    stop_transcription_flags[transcription_id] = False
    
    import tempfile
    temp_dir = tempfile.gettempdir()
    file_path = None
    file_object_name = None
        
    try:
        # Сохраняем файл 
        content = await file.read()
        logger.info(f"Файл получен, размер: {len(content)} байт")
        
        if minio_client:
            try:
                file_ext = os.path.splitext(file.filename)[1] if file.filename else ""
                file_object_name = minio_client.generate_object_name(prefix="media_", extension=file_ext)
                minio_client.upload_file(content, file_object_name, content_type="application/octet-stream")
                file_path = minio_client.get_file_path(file_object_name)
                logger.info(f"Файл загружен в MinIO: {file_object_name}")
            except Exception as e:
                logger.warning(f"Ошибка загрузки в MinIO, используем локальный файл: {e}")
                file_path = os.path.join(temp_dir, f"media_{datetime.now().timestamp()}_{file.filename}")
                with open(file_path, "wb") as f:
                    f.write(content)
        else:
            file_path = os.path.join(temp_dir, f"media_{datetime.now().timestamp()}_{file.filename}")
            with open(file_path, "wb") as f:
                f.write(content)
        
        logger.info(f"Временный путь файла: {file_path}")
        
        # Проверяем флаг остановки перед началом транскрибации
        if stop_transcription_flags.get(transcription_id, False):
            logger.info(f"Транскрибация {transcription_id} была остановлена до начала")
            raise HTTPException(status_code=499, detail="Транскрибация была остановлена")
        
        # Транскрибируем
        logger.info(f"Начинаем транскрибацию с диаризацией по ролям...")
        
        import concurrent.futures
        loop = asyncio.get_event_loop()
        
        def _transcribe():
            # Проверка флага внутри потока
            if stop_transcription_flags.get(transcription_id, False):
                return False, "Транскрибация была остановлена"
            
            try:
                if hasattr(transcriber, 'transcribe_with_diarization'):
                    logger.info("Используем принудительную диаризацию...")
                    # Этот вызов в UniversalTranscriber теперь пойдет в микросервис
                    result = transcriber.transcribe_with_diarization(file_path)
                else:
                    logger.info("Используем стандартную транскрибацию...")
                    result = transcriber.transcribe_audio_file(file_path)
                
                if stop_transcription_flags.get(transcription_id, False):
                    return False, "Транскрибация была остановлена"
                
                return result
            except Exception as e:
                logger.error(f"Ошибка транскрибации: {e}")
                return False, str(e)
        
        with concurrent.futures.ThreadPoolExecutor() as executor:
            success, result = await loop.run_in_executor(executor, _transcribe)
        
        if transcription_id in stop_transcription_flags:
            del stop_transcription_flags[transcription_id]
        
        if success:
            return {
                "transcription": result,
                "filename": file.filename,
                "success": True,
                "timestamp": datetime.now().isoformat(),
                "diarization": True,
                "transcription_id": transcription_id
            }
        else:
            if "остановлена" in str(result).lower():
                raise HTTPException(status_code=499, detail=result)
            raise HTTPException(status_code=400, detail=result)
            
    except Exception as e:
        logger.error(f"Ошибка в эндпоинте транскрибации: {e}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        # Очистка 
        try:
            if file_path and os.path.exists(file_path):
                os.remove(file_path)
            if minio_client and file_object_name:
                try: minio_client.delete_file(file_object_name)
                except: pass
        except: pass

@app.post("/api/transcribe/stop")
async def stop_transcription(request: Dict[str, Any]):
    """Остановить транскрибацию по ID"""
    transcription_id = request.get('transcription_id')
    if not transcription_id:
        raise HTTPException(status_code=400, detail="transcription_id обязателен")
    
    stop_transcription_flags[transcription_id] = True
    logger.info(f"Установлен флаг остановки для {transcription_id}")
    return {"success": True, "message": "Команда остановки отправлена", "transcription_id": transcription_id}

@app.post("/api/transcribe/upload/diarization")
async def transcribe_file_with_diarization(file: UploadFile = File(...)):
    """Принудительно транскрибировать аудио/видео с диаризацией"""
    logger.info(f"=== Начало принудительной диаризации: {file.filename} ===")
    
    if not transcriber:
        raise HTTPException(status_code=503, detail="Transcriber не доступен")
    
    import tempfile
    temp_dir = tempfile.gettempdir()
    file_path = None
    file_object_name = None
        
    try:
        content = await file.read()
        if minio_client:
            try:
                file_ext = os.path.splitext(file.filename)[1] if file.filename else ""
                file_object_name = minio_client.generate_object_name(prefix="media_diarization_", extension=file_ext)
                minio_client.upload_file(content, file_object_name, content_type="application/octet-stream")
                file_path = minio_client.get_file_path(file_object_name)
            except Exception as e:
                logger.warning(f"MinIO error: {e}")
                file_path = os.path.join(temp_dir, f"media_diarization_{datetime.now().timestamp()}_{file.filename}")
                with open(file_path, "wb") as f: f.write(content)
        else:
            file_path = os.path.join(temp_dir, f"media_diarization_{datetime.now().timestamp()}_{file.filename}")
            with open(file_path, "wb") as f: f.write(content)
        
        import concurrent.futures
        loop = asyncio.get_event_loop()
        
        def _transcribe():
            if hasattr(transcriber, 'transcribe_with_diarization'):
                return transcriber.transcribe_with_diarization(file_path)
            return transcriber.transcribe_audio_file(file_path)
        
        with concurrent.futures.ThreadPoolExecutor() as executor:
            success, result = await loop.run_in_executor(executor, _transcribe)
        
        if success:
            return {"transcription": result, "filename": file.filename, "success": True, "diarization": True}
        else:
            raise HTTPException(status_code=400, detail=result)
            
    except Exception as e:
        logger.error(f"Ошибка эндпоинта диаризации: {e}")
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        # Очистка 
        try:
            if file_path and os.path.exists(file_path):
                os.remove(file_path)
                logger.info(f"Временный файл удален: {file_path}")
            
            if minio_client and file_object_name:
                try:
                    minio_client.delete_file(file_object_name)
                except Exception as e:
                    logger.warning(f"Ошибка удаления файла из MinIO: {e}")
        except Exception as e:
            logger.warning(f"Ошибка очистки временных файлов: {e}")

@app.post("/api/transcribe/youtube")
async def transcribe_youtube(request: YouTubeTranscribeRequest):
    """Транскрибировать видео с YouTube """
    logger.info(f"=== Начало YouTube транскрибации: {request.url} ===")
    
    if not transcriber:
        logger.error("Transcriber не доступен")
        raise HTTPException(status_code=503, detail="Transcriber не доступен")
        
    try:
        logger.info("Начинаем YouTube транскрибацию...")
        import concurrent.futures
        loop = asyncio.get_event_loop()
        
        with concurrent.futures.ThreadPoolExecutor() as executor:
            success, result = await loop.run_in_executor(
                executor, 
                transcriber.transcribe_youtube, 
                request.url
            )
        
        if success:
            logger.info("YouTube транскрибация завершена успешно")
            return {
                "transcription": result,
                "url": request.url,
                "success": True,
                "timestamp": datetime.now().isoformat(),
                "diarization": True
            }
        else:
            logger.error(f"Ошибка YouTube транскрибации: {result}")
            raise HTTPException(status_code=400, detail=result)
            
    except Exception as e:
        logger.error(f"Ошибка YouTube транскрибации: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/system/status")
async def get_system_status():
    """Получить статус всех модулей и МИКРОСЕРВИСОВ системы (УЛУЧШЕНО)"""
    
    # 1. Проверка локальных модулей 
    module_status = {
        "ai_agent": {
            "available": ask_agent is not None,
            "functions": {
                "ask_agent": ask_agent is not None,
                "model_settings": model_settings is not None,
                "update_model_settings": update_model_settings is not None,
                "reload_model_by_path": reload_model_by_path is not None,
                "get_model_info": get_model_info is not None,
                "initialize_model": initialize_model is not None
            }
        },
        "memory": {
            "available": save_dialog_entry is not None,
            "functions": {
                "save_dialog_entry": save_dialog_entry is not None,
                "load_dialog_history": load_dialog_history is not None,
                "clear_dialog_history": clear_dialog_history is not None,
                "get_recent_dialog_history": get_recent_dialog_history is not None
            }
        },
        "voice": {
            "available": speak_text is not None and recognize_speech_from_file is not None,
            "functions": {
                "speak_text": speak_text is not None,
                "recognize_speech": recognize_speech is not None,
                "recognize_speech_from_file": recognize_speech_from_file is not None,
                "check_vosk_model": check_vosk_model is not None
            }
        },
        "transcription": {
            "available": transcriber is not None,
            "functions": {
                "universal_transcriber": UniversalTranscriber is not None
            }
        },
        "rag_service": {
            "available": rag_client is not None
        }
    }

    # 2. Проверка связи с 5 микросервисами по сети 
    services_health = {}
    try:
        from backend.llm_client import get_llm_service
        service = await get_llm_service()
        
        # Запускаем проверки параллельно, чтобы не тормозить API
        tasks = [
            service.client.health_check(),              # LLM
            service.client.get_transcription_health(), # STT
            service.client.get_tts_health(),           # TTS
            service.client.get_ocr_health()            # OCR
        ]
        
        # Для диаризации добавим простую проверку через httpx
        import httpx
        async def check_diar():
            try:
                async with httpx.AsyncClient(timeout=2.0) as c:
                    r = await c.get(f"{service.client.diarization_url}/health")
                    return r.json() if r.status_code == 200 else {"status": "error"}
            except: return {"status": "unreachable"}

        tasks.append(check_diar())
        
        results = await asyncio.gather(*tasks, return_exceptions=True)
        
        services_health = {
            "llm": results[0] if not isinstance(results[0], Exception) else {"status": "error"},
            "stt": results[1] if not isinstance(results[1], Exception) else {"status": "error"},
            "tts": results[2] if not isinstance(results[2], Exception) else {"status": "error"},
            "ocr": results[3] if not isinstance(results[3], Exception) else {"status": "error"},
            "diarization": results[4] if not isinstance(results[4], Exception) else {"status": "error"}
        }
    except Exception as e:
        logger.error(f"Ошибка при сетевой проверке сервисов: {e}")
        services_health = {"error": "Could not connect to health-check client"}

    return {
        "modules": module_status,
        "microservices": services_health,
        "timestamp": datetime.now().isoformat()
    }
# ================================
# ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ
# ================================

def get_current_model_path():
    """Получить путь к текущей загруженной модели"""
    try:
        # Пытаемся получить от AI модуля
        if get_model_info:
            result = get_model_info()
            if result and 'path' in result:
                return result['path']
        
        # Fallback к сохраненным настройкам
        app_settings = load_app_settings()
        return app_settings.get('current_model_path')
    except Exception as e:
        logger.error(f"Ошибка получения пути модели: {e}")
        return None

# ================================
# КОНТЕКСТНЫЕ ПРОМПТЫ API 
# ================================

@app.get("/api/context-prompts/global")
async def get_global_prompt():
    """Получить глобальный контекстный промпт"""
    try:
        prompt = context_prompt_manager.get_global_prompt()
        return {
            "prompt": prompt,
            "success": True,
            "timestamp": datetime.now().isoformat()
        }
    except Exception as e:
        logger.error(f"Ошибка при получении глобального промпта: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.put("/api/context-prompts/global")
async def update_global_prompt(request: Dict[str, str]):
    """Обновить глобальный контекстный промпт"""
    try:
        prompt = request.get("prompt", "")
        
        success = context_prompt_manager.set_global_prompt(prompt)
        if success:
            return {
                "message": "Глобальный промпт обновлен",
                "success": True,
                "timestamp": datetime.now().isoformat()
            }
        else:
            raise HTTPException(status_code=500, detail="Ошибка при сохранении промпта")
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Ошибка при обновлении глобального промпта: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/context-prompts/models")
async def get_models_with_prompts():
    """Получить список всех моделей с их контекстными промптами"""
    try:
        models = context_prompt_manager.get_models_list()
        return {
            "models": models,
            "success": True,
            "timestamp": datetime.now().isoformat()
        }
    except Exception as e:
        logger.error(f"Ошибка при получении списка моделей: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/context-prompts/model/{model_path:path}")
async def get_model_prompt(model_path: str):
    """Получить контекстный промпт для конкретной модели"""
    try:
        prompt = context_prompt_manager.get_model_prompt(model_path)
        return {
            "model_path": model_path,
            "prompt": prompt,
            "success": True,
            "timestamp": datetime.now().isoformat()
        }
    except Exception as e:
        logger.error(f"Ошибка при получении промпта модели: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.put("/api/context-prompts/model/{model_path:path}")
async def update_model_prompt(model_path: str, request: Dict[str, str]):
    """Обновить контекстный промпт для конкретной модели"""
    try:
        prompt = request.get("prompt", "")
        
        success = context_prompt_manager.set_model_prompt(model_path, prompt)
        if success:
            return {
                "message": f"Промпт для модели {model_path} обновлен",
                "success": True,
                "timestamp": datetime.now().isoformat()
            }
        else:
            raise HTTPException(status_code=500, detail="Ошибка при сохранении промпта")
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Ошибка при обновлении промпта модели: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/context-prompts/custom")
async def get_custom_prompts():
    """Получить все пользовательские промпты"""
    try:
        prompts = context_prompt_manager.get_all_custom_prompts()
        return {
            "prompts": prompts,
            "success": True,
            "timestamp": datetime.now().isoformat()
        }
    except Exception as e:
        logger.error(f"Ошибка при получении пользовательских промптов: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/context-prompts/custom")
async def create_custom_prompt(request: Dict[str, str]):
    """Создать новый пользовательский промпт"""
    try:
        prompt_id = request.get("id", "")
        prompt = request.get("prompt", "")
        description = request.get("description", "")
        
        if not prompt_id.strip() or not prompt.strip():
            raise HTTPException(status_code=400, detail="ID и промпт обязательны")
        
        success = context_prompt_manager.set_custom_prompt(prompt_id, prompt, description)
        if success:
            return {
                "message": f"Пользовательский промпт '{prompt_id}' создан",
                "success": True,
                "timestamp": datetime.now().isoformat()
            }
        else:
            raise HTTPException(status_code=500, detail="Ошибка при создании промпта")
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Ошибка при создании пользовательского промпта: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.delete("/api/context-prompts/custom/{prompt_id}")
async def delete_custom_prompt(prompt_id: str):
    """Удалить пользовательский промпт"""
    try:
        success = context_prompt_manager.delete_custom_prompt(prompt_id)
        if success:
            return {
                "message": f"Пользовательский промпт '{prompt_id}' удален",
                "success": True,
                "timestamp": datetime.now().isoformat()
            }
        else:
            raise HTTPException(status_code=404, detail="Промпт не найден")
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Ошибка при удалении пользовательского промпта: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/context-prompts/effective/{model_path:path}")
async def get_effective_prompt(model_path: str, custom_prompt_id: Optional[str] = None):
    """Получить эффективный промпт для модели с учетом приоритетов"""
    try:
        prompt = context_prompt_manager.get_effective_prompt(model_path, custom_prompt_id)
        return {
            "model_path": model_path,
            "custom_prompt_id": custom_prompt_id,
            "prompt": prompt,
            "success": True,
            "timestamp": datetime.now().isoformat()
        }
    except Exception as e:
        logger.error(f"Ошибка при получении эффективного промпта: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# ================================
# АГЕНТНАЯ АРХИТЕКТУРА API 
# ================================

class AgentModeRequest(BaseModel):
    mode: str  # "agent", "direct" или "multi-llm"

class MultiLLMModelsRequest(BaseModel):
    models: List[str]  # Список имен моделей для режима multi-llm

class AgentStatusResponse(BaseModel):
    is_initialized: bool
    mode: str
    available_agents: int
    orchestrator_active: bool

@app.get("/api/agent/status")
async def get_agent_status():
    """Получить статус агентной архитектуры"""
    try:
        orchestrator = get_agent_orchestrator()
        if orchestrator:
            status = orchestrator.get_status()
            return AgentStatusResponse(**status)
        else:
            return AgentStatusResponse(
                is_initialized=False,
                mode="unknown",
                available_agents=0,
                orchestrator_active=False
            )
    except Exception as e:
        logger.error(f"Ошибка получения статуса агентной архитектуры: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/agent/mode")
async def set_agent_mode(request: AgentModeRequest):
    """Установить режим работы агентной архитектуры"""
    try:
        orchestrator = get_agent_orchestrator()
        if orchestrator:
            orchestrator.set_mode(request.mode)
            return {
                "message": f"Режим работы изменен на: {request.mode}",
                "success": True,
                "timestamp": datetime.now().isoformat()
            }
        else:
            raise HTTPException(status_code=503, detail="Агентная архитектура не инициализирована")
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Ошибка установки режима агентной архитектуры: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/agent/multi-llm/models")
async def set_multi_llm_models(request: MultiLLMModelsRequest):
    """Установить список моделей для режима multi-llm"""
    try:
        orchestrator = get_agent_orchestrator()
        if orchestrator:
            orchestrator.set_multi_llm_models(request.models)
            return {
                "message": f"Установлены модели для режима multi-llm: {', '.join(request.models)}",
                "success": True,
                "models": request.models,
                "timestamp": datetime.now().isoformat()
            }
        else:
            raise HTTPException(status_code=503, detail="Агентная архитектура не инициализирована")
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Ошибка установки моделей для режима multi-llm: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/agent/multi-llm/models")
async def get_multi_llm_models():
    """Получить список моделей для режима multi-llm"""
    try:
        orchestrator = get_agent_orchestrator()
        if orchestrator:
            models = orchestrator.get_multi_llm_models()
            return {
                "models": models,
                "success": True
            }
        else:
            raise HTTPException(status_code=503, detail="Агентная архитектура не инициализирована")
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Ошибка получения моделей для режима multi-llm: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/agent/agents")
async def get_available_agents():
    """Получить список доступных агентов"""
    try:
        orchestrator = get_agent_orchestrator()
        if orchestrator:
            agents = orchestrator.get_available_agents()
            return {
                "agents": agents,
                "count": len(agents),
                "success": True,
                "timestamp": datetime.now().isoformat()
            }
        else:
            raise HTTPException(status_code=503, detail="Агентная архитектура не инициализирована")
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Ошибка получения списка агентов: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/agent/mcp/status")
async def get_mcp_status():
    """Получить статус MCP серверов"""
    try:
        orchestrator = get_agent_orchestrator()
        if orchestrator:
            return {
                "mcp_status": {
                    "initialized": False, 
                    "servers": 0, 
                    "tools": 0,
                    "message": "MCP интеграция в разработке"
                },
                "success": True,
                "timestamp": datetime.now().isoformat()
            }
        else:
            raise HTTPException(status_code=503, detail="Агентная архитектура не инициализирована")
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Ошибка получения статуса MCP: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/agent/agents/{agent_id}/status")
async def set_agent_status(agent_id: str, status: Dict[str, bool]):
    """Установить статус активности агента"""
    try:
        orchestrator = get_agent_orchestrator()
        if not orchestrator:
            raise HTTPException(status_code=503, detail="Агентная архитектура не инициализирована")
        
        is_active = status.get("is_active", True)
        orchestrator.set_agent_status(agent_id, is_active)
        
        return {
            "agent_id": agent_id,
            "is_active": is_active,
            "success": True,
            "message": f"Агент '{agent_id}' {'активирован' if is_active else 'деактивирован'}",
            "timestamp": datetime.now().isoformat()
        }
    except Exception as e:
        logger.error(f"Ошибка изменения статуса агента: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/agent/agents/statuses")
async def get_all_agent_statuses():
    """Получить статусы всех агентов"""
    try:
        orchestrator = get_agent_orchestrator()
        if not orchestrator:
            raise HTTPException(status_code=503, detail="Агентная архитектура не инициализирована")
        
        statuses = orchestrator.get_all_agent_statuses()
        return {
            "statuses": statuses,
            "success": True,
            "timestamp": datetime.now().isoformat()
        }
    except Exception as e:
        logger.error(f"Ошибка получения статусов агентов: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/agent/langgraph/status")
async def get_langgraph_status():
    """Получить статус LangGraph оркестратора"""
    try:
        orchestrator = get_agent_orchestrator()
        if orchestrator:
            tools = orchestrator.get_available_tools()
            return {
                "langgraph_status": {
                    "is_active": orchestrator.is_initialized,
                    "initialized": orchestrator.is_initialized,
                    "tools_available": len(tools),
                    "memory_enabled": True,
                    "orchestrator_type": "LangGraph",
                    "orchestrator_active": orchestrator.is_orchestrator_active()
                },
                "success": True, "timestamp": datetime.now().isoformat()
            }
        else:
            raise HTTPException(status_code=503, detail="Агентная архитектура не инициализирована")
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Ошибка получения статуса LangGraph: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/agent/orchestrator/toggle")
async def toggle_orchestrator(status: Dict[str, bool]):
    """Включить/выключить оркестратор"""
    try:
        orchestrator = get_agent_orchestrator()
        if not orchestrator:
            raise HTTPException(status_code=503, detail="Агентная архитектура не инициализирована")
        
        is_active = status.get("is_active", True)
        
        # Устанавливаем статус оркестратора
        orchestrator.set_orchestrator_status(is_active)
        
        return {
            "success": True,
            "orchestrator_active": is_active,
            "message": f"Оркестратор {'включен' if is_active else 'отключен'}",
            "timestamp": datetime.now().isoformat()
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Ошибка переключения оркестратора: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# ================================
# БАЗА ЗНАНИЙ (Knowledge Base RAG)
# ================================

@app.post("/api/kb/documents")
async def kb_upload_document(file: UploadFile = File(...)):
    """Загрузить документ в постоянную Базу Знаний (KB RAG)."""
    if not rag_client:
        raise HTTPException(status_code=503, detail="RAG service недоступен")
    try:
        content = await file.read()
        if not content:
            raise HTTPException(status_code=400, detail="Файл пустой")
        result = await rag_client.kb_upload_document(
            file_bytes=content,
            filename=file.filename or "unknown",
        )
        return result
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Ошибка загрузки в Базу Знаний: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/kb/documents")
async def kb_list_documents():
    """Список документов в постоянной Базе Знаний."""
    if not rag_client:
        raise HTTPException(status_code=503, detail="RAG service недоступен")
    try:
        docs = await rag_client.kb_list_documents()
        return {"documents": docs}
    except Exception as e:
        logger.error(f"Ошибка получения списка Базы Знаний: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.delete("/api/kb/documents/{document_id}")
async def kb_delete_document(document_id: int):
    """Удалить документ из постоянной Базы Знаний."""
    if not rag_client:
        raise HTTPException(status_code=503, detail="RAG service недоступен")
    try:
        result = await rag_client.kb_delete_document(document_id)
        return result
    except Exception as e:
        logger.error(f"Ошибка удаления из Базы Знаний: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ================================
# Библиотека памяти RAG (настройки): MinIO + memory_rag_* в Postgres
# ================================

@app.post("/api/memory-rag/documents")
async def memory_rag_upload_document(file: UploadFile = File(...)):
    """Загрузить файл: MinIO (отдельный bucket) + индексация в memory_rag_documents/vectors."""
    if not rag_client:
        raise HTTPException(status_code=503, detail="RAG service недоступен")
    try:
        content = await file.read()
        if not content:
            raise HTTPException(status_code=400, detail="Файл пустой")
        fn = file.filename or "unknown"
        ext = os.path.splitext(fn)[1] or ".bin"
        memory_bucket = settings.minio.memory_rag_bucket_name
        file_object_name = None
        if minio_client:
            try:
                minio_client.ensure_bucket(memory_bucket)
                file_object_name = minio_client.generate_object_name(prefix="memrag_", extension=ext)
                minio_client.upload_file(
                    content,
                    file_object_name,
                    content_type=file.content_type or "application/octet-stream",
                    bucket_name=memory_bucket,
                )
            except Exception as e:
                logger.error(f"MinIO загрузка memory-rag: {e}")
                raise HTTPException(status_code=500, detail=f"MinIO: {e}")
        try:
            result = await rag_client.memory_rag_index_document(
                file_bytes=content,
                filename=fn,
                minio_object=file_object_name,
                minio_bucket=memory_bucket if file_object_name else None,
            )
        except Exception as e:
            if minio_client and file_object_name:
                try:
                    minio_client.delete_file(file_object_name, bucket_name=memory_bucket)
                except Exception:
                    pass
            logger.error(f"SVC-RAG memory-rag индексация: {e}")
            raise HTTPException(status_code=422, detail=str(e))
        return result
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Ошибка загрузки memory-rag: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/memory-rag/documents")
async def memory_rag_list_documents():
    if not rag_client:
        raise HTTPException(status_code=503, detail="RAG service недоступен")
    try:
        docs = await rag_client.memory_rag_list_documents()
        return {"documents": docs}
    except Exception as e:
        logger.error(f"Ошибка списка memory-rag: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.delete("/api/memory-rag/documents/{document_id}")
async def memory_rag_delete_document(document_id: int):
    if not rag_client:
        raise HTTPException(status_code=503, detail="RAG service недоступен")
    try:
        out = await rag_client.memory_rag_delete_document(document_id)
        if not out.get("ok"):
            raise HTTPException(status_code=404, detail="Документ не найден")
        mo, mb = out.get("minio_object"), out.get("minio_bucket")
        if minio_client and mo and mb:
            try:
                minio_client.delete_file(mo, bucket_name=mb)
            except Exception as ex:
                logger.warning(f"Не удалось удалить объект MinIO memory-rag: {ex}")
        return {"ok": True, "document_id": document_id}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Ошибка удаления memory-rag: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ================================
# СТАТИЧЕСКИЕ ФАЙЛЫ И ФРОНТЕНД
# ================================

# Подключаем статические файлы React приложения (только для локального запуска)
# В Docker фронтенд работает в отдельном контейнере на порту 3000
is_docker = os.getenv("DOCKER_ENV", "").lower() == "true"
if not is_docker and os.path.exists("../frontend/build"):
    app.mount("/static", StaticFiles(directory="../frontend/build/static"), name="static")
    
    @app.get("/{path:path}")
    async def serve_react_app(path: str):
        """Отдаем React приложение для всех остальных маршрутов"""
        index_file = "../frontend/build/index.html"
        if os.path.exists(index_file):
            return FileResponse(index_file)
        else:
            return {"message": "Frontend not built"}

if __name__ == "__main__":
    print("Запуск astrachat Web Backend...")
    print(f"Текущая директория: {os.getcwd()}")
    print(f"Backend директория: {os.path.dirname(os.path.abspath(__file__))}")
    print(f"Корневая директория: {os.path.dirname(os.path.dirname(os.path.abspath(__file__)))}")
    print(f"Python path: {sys.path[:3]}...")
    print(f"API документация: {urls_config.get('backend_port_1')}/docs")
    backend_url = urls_config.get('backend_port_1', '').replace('http://', 'ws://')
    print(f"WebSocket: {backend_url}/ws/chat")
    
    # Восстанавливаем сохраненную модель
    try:
        app_settings = load_app_settings()
        saved_model_path = app_settings.get('current_model_path')
        
        # Проверяем, что путь валидный
        if saved_model_path and reload_model_by_path:
            # Если путь начинается с llm-svc://, модель уже доступна через llm-svc
            if saved_model_path.startswith("llm-svc://"):
                logger.info(f"Модель из llm-svc уже доступна: {saved_model_path}")
            # Проверяем локально (fallback)
            elif os.path.exists(saved_model_path) and not os.path.isdir(saved_model_path):
                logger.info(f"Восстанавливаю сохраненную модель: {saved_model_path}")
                success = reload_model_by_path(saved_model_path)
                if success:
                    logger.info(f"Модель восстановлена: {saved_model_path}")
                else:
                    logger.warning(f"Не удалось восстановить модель: {saved_model_path}")
            elif os.path.isdir(saved_model_path):
                logger.warning(f"Сохраненный путь является директорией: {saved_model_path}")
            else:
                logger.info(f"Сохраненный путь модели не существует: {saved_model_path}")
        else:
            logger.info("Нет сохраненной модели для восстановления")
    except Exception as e:
        logger.error(f"Ошибка восстановления модели: {e}")
    
    uvicorn.run(
        app,
        host="0.0.0.0",
        port=8000,
        reload=False,
        log_level="info"
    )