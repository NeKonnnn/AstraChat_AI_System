"""
routes/documents.py - загрузка, удаление, запросы к документам, отчеты OCR
"""

import json
import os
from datetime import datetime
from typing import Annotated

from fastapi import APIRouter, Depends, File, HTTPException, Query, Request, UploadFile
from fastapi.responses import FileResponse, Response

import backend.app_state as state
from backend.app_state import ask_agent, get_rag_chat_top_k, get_rag_chunk_index_params, minio_client, rag_client, settings
from backend.auth.jwt_handler import get_current_user
from backend.rag_query.post_generation import maybe_replace_ungrounded
from backend.rag_query.prompts import RAG_STRICT_NOT_FOUND_MESSAGE, merge_strict_rag_system_prompt
from backend.rag_query.semantic_cache import bump_rag_semantic_cache
from backend.realtime.helpers import _is_structure_query
from backend.realtime.rag_evidence import (
    RAG_NO_RELEVANT_CONTEXT_MESSAGE,
    build_rag_id_to_filename,
    filter_rag_hits_by_score,
    format_rag_fragments,
    rag_guard_env,
)
from backend.schemas import DocumentQueryRequest
from backend.settings.cef_logger.cef_logger import log_cef_event
from backend.settings.logging import get_logger
from backend.settings.logging.errors import logged_suppress
from backend.settings.service_toggles import require_service  # FEATURE-FLAG

logger = get_logger(__name__)


def _documents_bucket_name() -> str:
    """Имя bucket для документов/вложений — из settings (YAML + ENV), не hardcoded default."""
    minio_cfg = getattr(settings, "minio", None)
    if minio_cfg is not None:
        docs_bucket = getattr(minio_cfg, "documents_bucket_name", None)
        if docs_bucket:
            return str(docs_bucket)
        main_bucket = getattr(minio_cfg, "bucket_name", None)
        if main_bucket:
            return str(main_bucket)
    return os.getenv("MINIO_DOCUMENTS_BUCKET_NAME") or os.getenv("MINIO_BUCKET_NAME") or "astrachat-documents"


router = APIRouter(prefix="/api/documents", tags=["documents"])
INLINE_ATTACHMENT_MAX_BYTES = 50 * 1024 * 1024
_INLINE_IMAGE_EXTS = {".jpg", ".jpeg", ".png", ".webp", ".gif"}
_INLINE_DOC_EXTS = {".pdf", ".docx", ".xlsx", ".xls", ".txt"}
_INLINE_ATTACH_SUPPORTED_EXTENSIONS = sorted(_INLINE_IMAGE_EXTS | _INLINE_DOC_EXTS)
_CONTENT_TYPE_BY_EXT = {
    ".pdf": "application/pdf",
    ".docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    ".doc": "application/msword",
    ".txt": "text/plain",
    ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    ".xls": "application/vnd.ms-excel",
    ".jpg": "image/jpeg",
    ".jpeg": "image/jpeg",
    ".png": "image/png",
    ".webp": "image/webp",
    ".gif": "image/gif",
}


def _inline_attach_supported_extensions_label() -> str:
    return ", ".join(_INLINE_ATTACH_SUPPORTED_EXTENSIONS)


def _unsupported_inline_attach_extension_detail(filename: str) -> str:
    ext = os.path.splitext(filename or "")[1].lower()
    supported = _inline_attach_supported_extensions_label()
    if ext:
        return f"{ext} файлы не поддерживаются, поддерживаются только следующие расширения файлов: {supported}"
    return f"файлы без расширения не поддерживаются, поддерживаются только следующие расширения файлов: {supported}"


def _emit_attach_info(message: str) -> None:
    logger.info(message)


def _log_inline_attach_upload_success(filename: str) -> None:
    logger.info("файл %s загружен успешно", filename)


def _log_inline_attach_upload_failure(filename: str, detail: str) -> None:
    logger.info("файл %s загружен не успешно. %s", filename, detail)


def _resolve_content_type(filename: str, content_type: str) -> str:
    ext = os.path.splitext(filename or "")[1].lower()
    if content_type:
        return content_type.lower()
    return _CONTENT_TYPE_BY_EXT.get(ext, "application/octet-stream")


def _detect_inline_attachment_kind(filename: str, content_type: str) -> str:
    ext = os.path.splitext(filename or "")[1].lower()
    ct = (content_type or "").lower()
    if ct.startswith("image/") or ext in _INLINE_IMAGE_EXTS:
        return "image"
    if ext in _INLINE_DOC_EXTS:
        return "document"
    doc_mimes = {
        "application/pdf",
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel",
        "text/plain",
    }
    if ct in doc_mimes:
        return "document"
    return "unsupported"


async def _analyze_inline_attachment_for_model(filename: str, content_type: str, file_size: int) -> dict:
    from backend.app_state import get_current_model_path

    kind = _detect_inline_attachment_kind(filename, content_type)
    format_allowed = kind != "unsupported"
    size_allowed = file_size <= INLINE_ATTACHMENT_MAX_BYTES
    model_path = get_current_model_path()
    provider_id = None
    model_vision = None
    model_compatible = True
    incompatibility_reason = None
    if model_path:
        try:
            from backend.llm_providers import get_registry

            registry = await get_registry()
            provider, _model_id = registry.resolve(model_path)
            provider_id = provider.id
            model_vision = bool(provider.capabilities.vision)
            if kind == "image" and (not model_vision):
                model_compatible = False
                incompatibility_reason = "модель не поддерживает vision (изображения)"
        except Exception:
            logger.exception("[ChatAttach] не удалось определить capabilities модели")
    elif kind == "image":
        model_compatible = False
        incompatibility_reason = "модель не выбрана — изображение может не обработаться"
    allowed = format_allowed and size_allowed
    reject_reason = None
    if not format_allowed:
        reject_reason = "Неподдерживаемый формат (PDF, DOCX, XLSX, TXT, JPEG, PNG, WEBP, GIF)"
    elif not size_allowed:
        reject_reason = f"Размер превышает {INLINE_ATTACHMENT_MAX_BYTES // (1024 * 1024)}MB"
    return {
        "filename": filename,
        "content_type": content_type,
        "size": file_size,
        "kind": kind,
        "format_allowed": format_allowed,
        "size_allowed": size_allowed,
        "allowed": allowed,
        "model_path": model_path,
        "provider_id": provider_id,
        "model_vision": model_vision,
        "model_compatible": model_compatible,
        **({"reject_reason": reject_reason} if reject_reason else {}),
        **({"incompatibility_reason": incompatibility_reason} if incompatibility_reason else {}),
    }


def _log_inline_attachment_debug(stage: str, analysis: dict, **extra) -> None:
    payload = {"stage": stage, **analysis, **extra}
    logger.info("[ChatAttach] %s", json.dumps(payload, ensure_ascii=False, default=str))


def _log_chat_attach_request(request: Request, **extra) -> None:
    logger.info(
        "[ChatAttach] %s",
        json.dumps(
            {
                "stage": "request-start",
                "content_length_header": request.headers.get("content-length"),
                "content_type_header": request.headers.get("content-type"),
                "client_host": getattr(request.client, "host", None) if request.client else None,
                **extra,
            },
            ensure_ascii=False,
            default=str,
        ),
    )


def _extract_inline_payload(content: bytes, filename: str, content_type: str) -> dict:
    """
    Извлекает содержимое для inline-передачи в модель (без RAG).
    Возвращает: { type: 'text'|'image', content: str, cs1: str }
    """
    import base64
    import io

    ext = os.path.splitext(filename)[1].lower()
    ct = (content_type or "").lower()
    if ct.startswith("image/") or ext in (".jpg", ".jpeg", ".png", ".webp", ".gif"):
        mime = ct if ct.startswith("image/") else _CONTENT_TYPE_BY_EXT.get(ext, "image/jpeg")
        b64 = base64.b64encode(content).decode("ascii")
        return {"type": "image", "content": f"data:{mime};base64,{b64}", "cs1": "image"}
    if ext == ".pdf":
        from PyPDF2 import PdfReader

        reader = PdfReader(io.BytesIO(content))
        pages_text = []
        for i, page in enumerate(reader.pages, 1):
            page_text = page.extract_text() or ""
            if page_text.strip():
                pages_text.append(f"""[Страница {i}]
{page_text}""")
        text = "\n\n".join(pages_text) if pages_text else "(PDF не содержит извлекаемого текста)"
        return {"type": "text", "content": text, "cs1": "pdf"}
    if ext == ".docx":
        import docx as docx_lib

        doc = docx_lib.Document(io.BytesIO(content))
        paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
        text = "\n".join(paragraphs) if paragraphs else "(Документ не содержит текста)"
        return {"type": "text", "content": text, "cs1": "docx"}
    if ext in (".xlsx", ".xls"):
        import openpyxl

        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
        lines = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            lines.append(f"=== Лист: {sheet_name} ===")
            for row in ws.iter_rows(values_only=True):
                row_str = "\t".join(("" if c is None else str(c) for c in row))
                if row_str.strip():
                    lines.append(row_str)
        text = "\n".join(lines) if lines else "(Таблица не содержит данных)"
        return {"type": "text", "content": text, "cs1": "xlsx"}
    text = content.decode("utf-8", errors="replace")
    return {"type": "text", "content": text, "cs1": "text"}


def _try_upload_bytes_to_minio(
    request: Request, content: bytes, filename: str, content_type: str
) -> tuple[str | None, str | None]:
    """
    Пытается загрузить байты в MinIO. При недоступности MinIO возвращает (None, None)
    — прикрепление к сообщению всё равно работает (inline в модель).
    """
    if not minio_client:
        logger.warning("MinIO недоступен — файл %s не сохранён в объектное хранилище", filename)
        return (None, None)
    documents_bucket = _documents_bucket_name()
    ext = os.path.splitext(filename)[1].lower()
    is_image = ext in (".jpg", ".jpeg", ".png", ".webp", ".gif") or content_type.startswith("image/")
    file_object_name = minio_client.generate_object_name(prefix="img_" if is_image else "doc_", extension=ext or ".bin")
    try:
        minio_client.upload_file(
            content,
            file_object_name,
            content_type=content_type,
            bucket_name=documents_bucket,
            cef_display_name=filename,
        )
        return (file_object_name, documents_bucket)
    except Exception:
        logger.exception("MinIO attach upload error")
        return (None, None)


@router.post("/upload")
async def upload_document(request: Request, file: Annotated[UploadFile, File(...)]):
    require_service("rag")  # FEATURE-FLAG
    if not rag_client:
        raise HTTPException(status_code=503, detail="RAG service недоступен")
    documents_bucket = _documents_bucket_name()
    file_object_name = None
    try:
        content = await file.read()
        file_extension = os.path.splitext(file.filename)[1].lower() if file.filename else ""
        is_image = file_extension in (".jpg", ".jpeg", ".png", ".webp")
        content_type_map = {
            ".pdf": "application/pdf",
            ".docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            ".doc": "application/msword",
            ".txt": "text/plain",
            ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ".xls": "application/vnd.ms-excel",
            ".jpg": "image/jpeg",
            ".jpeg": "image/jpeg",
            ".png": "image/png",
            ".webp": "image/webp",
        }
        content_type = content_type_map.get(file_extension, "application/octet-stream")
        if minio_client:
            try:
                file_object_name = minio_client.generate_object_name(
                    prefix="img_" if is_image else "doc_", extension=file_extension
                )
                minio_client.upload_file(
                    content,
                    file_object_name,
                    content_type=content_type,
                    bucket_name=documents_bucket,
                    cef_display_name=file.filename or file_object_name,
                )
            except Exception:
                logger.exception("MinIO upload")
                file_object_name = None
        try:
            chunk_params = get_rag_chunk_index_params()
            rag_result = await rag_client.upload_document(
                file_bytes=content,
                filename=file.filename or file_object_name or "unknown",
                minio_object=file_object_name,
                minio_bucket=documents_bucket if minio_client and file_object_name else None,
                original_path=None,
                **chunk_params,
            )
        except Exception as e:
            logger.exception("Ошибка операции")
            if minio_client and file_object_name:
                with logged_suppress(logger):
                    minio_client.delete_file(file_object_name, bucket_name=documents_bucket)
            raise HTTPException(status_code=502, detail=f"Ошибка RAG-сервиса: {e}") from e
        if not rag_result.get("ok"):
            if minio_client and file_object_name:
                with logged_suppress(logger):
                    minio_client.delete_file(file_object_name, bucket_name=documents_bucket)
            raise HTTPException(status_code=400, detail=rag_result.get("error", "Ошибка индексации"))
        bump_rag_semantic_cache()
        result = {
            "message": "Документ успешно загружен",
            "filename": file.filename,
            "success": True,
            "rag_document_id": rag_result.get("document_id"),
        }
        _doc_id = rag_result.get("document_id")
        _ex: dict = {"fname": file.filename or "unknown", "fsize": len(content)}
        if _doc_id:
            _ex["cs2"] = str(_doc_id)
            _ex["cs2Label"] = "ObjectId"
        log_cef_event("FS005", request=request, status_code=200, extra=_ex)
        if is_image and minio_client and file_object_name:
            result["minio_object"] = file_object_name
            result["minio_bucket"] = documents_bucket
        return result
    except HTTPException:
        raise
    except Exception as e:
        logger.exception("Ошибка операции")
        raise HTTPException(status_code=500, detail=str(e)) from e


@router.post("/query")
async def query_document(request: DocumentQueryRequest):
    require_service("rag")  # FEATURE-FLAG
    if not rag_client:
        raise HTTPException(status_code=503, detail="RAG service недоступен")
    if not ask_agent:
        raise HTTPException(status_code=503, detail="AI agent не доступен")
    try:
        from backend.app_state import current_rag_strategy

        min_sim, _ = rag_guard_env()
        hits = await rag_client.search(request.query, k=get_rag_chat_top_k(), strategy=current_rag_strategy)
        hits = filter_rag_hits_by_score(hits, min_sim)
        if not hits:
            return {"response": RAG_NO_RELEVANT_CONTEXT_MESSAGE, "query": request.query, "success": True}
        if _is_structure_query(request.query):
            seen = {(d, i) for _, _, d, i in hits}
            for doc_id in {d for _, _, d, _ in hits if d}:
                with logged_suppress(logger):
                    for c, sc, did, idx in await rag_client.get_document_start_chunks(doc_id, max_chunks=2):
                        if (did, idx) not in seen:
                            hits = [(c, sc, did, idx)] + hits
                            seen.add((did, idx))
        id_map = build_rag_id_to_filename(list(await rag_client.list_documents() or []))
        parts, _ = format_rag_fragments(hits, id_map, max_chars=12000, store_label="global/rest-documents-search")
        prompt = f"CONTEXT:\n{chr(10).join(parts)}\nВопрос: {request.query}\n\nОтвет:"
        response_text = ask_agent(
            prompt,
            system_prompt=merge_strict_rag_system_prompt(
                None, rag_override=getattr(state, "rag_system_prompt", None)
            ),
        )
        response_text = await maybe_replace_ungrounded(prompt[:20000], response_text, RAG_STRICT_NOT_FOUND_MESSAGE)
        return {
            "response": response_text,
            "query": request.query,
            "success": True,
            "timestamp": datetime.now().isoformat(),
        }
    except Exception as e:
        logger.exception("Ошибка операции")
        raise HTTPException(status_code=500, detail=str(e)) from e


@router.get("")
@router.get("/")
async def get_documents():
    if not rag_client:
        raise HTTPException(status_code=503, detail="RAG service недоступен")
    try:
        docs = await rag_client.list_documents()
        filenames = [d.get("filename") for d in docs]
        return {"documents": filenames, "count": len(filenames), "success": True}
    except Exception as e:
        logger.exception("Ошибка операции")
        raise HTTPException(status_code=500, detail=str(e)) from e


@router.delete("/{filename}")
async def delete_document(filename: str, request: Request):
    require_service("rag")  # FEATURE-FLAG
    if not rag_client:
        raise HTTPException(status_code=503, detail="RAG service недоступен")
    try:
        docs = await rag_client.list_documents()
        filenames = [d.get("filename") for d in docs]
        if filename not in filenames:
            raise HTTPException(status_code=404, detail=f"Документ {filename} не найден")
        if minio_client:
            try:
                minio_info = await rag_client.get_image_minio_info(filename)
                if minio_info:
                    minio_client.delete_file(minio_info["minio_object"], bucket_name=minio_info["minio_bucket"])
            except Exception:
                logger.exception("MinIO delete")
        try:
            await rag_client.delete_document_by_filename(filename)
        except Exception as e:
            logger.exception("Ошибка операции")
            raise HTTPException(status_code=502, detail=f"Ошибка RAG-сервиса: {e}") from e
        bump_rag_semantic_cache()
        new_docs = await rag_client.list_documents()
        _fsize = 0
        _oid = None
        for d in docs or []:
            if d.get("filename") == filename:
                _fsize = int(d.get("size") or d.get("bytes") or d.get("file_size") or 0)
                _oid = d.get("document_id") or d.get("id")
                break
        _ex2: dict = {"fname": filename, "fsize": _fsize}
        if _oid:
            _ex2["cs2"] = str(_oid)
            _ex2["cs2Label"] = "ObjectId"
        log_cef_event("FS006", request=request, status_code=200, extra=_ex2)
        return {
            "message": f"Документ {filename} удален",
            "success": True,
            "remaining_documents": [d.get("filename") for d in new_docs],
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.exception("Ошибка операции")
        raise HTTPException(status_code=500, detail=str(e)) from e


@router.get("/report/generate")
async def generate_confidence_report():
    require_service("rag")  # FEATURE-FLAG
    if not rag_client:
        raise HTTPException(status_code=503, detail="RAG service недоступен")
    try:
        report_data = await rag_client.get_confidence_report()
        logger.info(f"Получены данные отчета: {report_data['total_documents']} документов")
        report_text = (
            f"ОТЧЕТ О СТЕПЕНИ УВЕРЕННОСТИ МОДЕЛИ В РАСПОЗНАННОМ ТЕКСТЕ\n"
            f"{'=' * 80}\n"
            f"Дата генерации: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n"
            f"{'=' * 80}\n\n"
            f"ОБЩАЯ ИНФОРМАЦИЯ:\n"
            f"- Всего обработано документов: {report_data['total_documents']}\n"
            f"- Средняя уверенность модели: {report_data['average_confidence']:.2f}%\n"
            f"- Всего слов: {report_data.get('total_words', 0)}\n"
            f"{'=' * 80}\n"
            f"ДЕТАЛЬНАЯ ИНФОРМАЦИЯ ПО ДОКУМЕНТАМ:\n"
        )
        for i, doc in enumerate(report_data["documents"], 1):
            report_text += (
                f"\n{i}. {doc['filename']}\n"
                f"   Тип файла: {doc['file_type']}\n"
                f"   Уверенность модели: {doc['confidence']:.2f}%\n"
                f"   Длина распознанного текста: {doc['text_length']} символов\n"
                f"   Количество слов: {doc.get('words_count', 0)}\n"
                f"   {'-' * 80}\n\n"
                f"   РАСПОЗНАННЫЙ ТЕКСТ С УВЕРЕННОСТЬЮ:\n"
            )
            formatted_text_info = next(
                (ft for ft in report_data.get("formatted_texts", []) if ft["filename"] == doc["filename"]), None
            )
            if formatted_text_info and formatted_text_info.get("words"):
                words = formatted_text_info.get("words", [])
                if words:
                    line_words = []
                    current_line = []
                    for word_info in words:
                        word = word_info.get("word", "")
                        conf = word_info.get("confidence", 0.0)
                        current_line.append((word, conf))
                        if len(current_line) >= 8:
                            line_words.append(current_line)
                            current_line = []
                    if current_line:
                        line_words.append(current_line)
                    if line_words:
                        for line in line_words:
                            import re

                            tokens_data = []
                            prev_is_punctuation = False
                            for word, conf in line:
                                is_punctuation = bool(re.match("^[^\\w\\s]+$", word))
                                word_width = len(word)
                                col_width = max(word_width + 2, 10)
                                tokens_data.append(
                                    {
                                        "word": word,
                                        "conf": conf,
                                        "is_punctuation": is_punctuation,
                                        "col_width": col_width,
                                        "needs_space_before": not prev_is_punctuation
                                        and (not is_punctuation)
                                        and tokens_data,
                                    }
                                )
                                prev_is_punctuation = is_punctuation
                            percent_line = "│"
                            word_line = "│"
                            separator_line = "├"
                            for idx, token in enumerate(tokens_data):
                                if token["needs_space_before"]:
                                    word_line += "│"
                                    percent_line += "│"
                                    separator_line += "┼"
                                percent_str = f"{token['conf']:.0f}%"
                                word_str = token["word"]
                                percent_padded = percent_str.center(token["col_width"])
                                word_padded = word_str.ljust(token["col_width"])
                                percent_line += percent_padded + "│"
                                word_line += word_padded + "│"
                                separator_line += "─" * token["col_width"] + (
                                    "┤" if idx == len(tokens_data) - 1 else "┼"
                                )
                            report_text += f"{percent_line}\n"
                            report_text += f"{separator_line}\n"
                            report_text += f"{word_line}\n\n"
                    else:
                        report_text += "[Нет валидных слов для отображения]\n"
                else:
                    report_text += "[Нет данных о словах]\n"
            else:
                report_text += "[Нет отформатированного текста]\n"
            report_text += f"{'-' * 80}\n"
        overall_conf = report_data.get("overall_confidence", report_data.get("average_confidence", 0.0))
        report_text += (
            f"\n{'=' * 80}\n"
            f"ИТОГО:\n"
            f"- Итоговая уверенность по всему распознанному тексту: {overall_conf:.2f}%\n"
            f"- Средняя уверенность по документам: {report_data['average_confidence']:.2f}%\n"
            f"- Всего документов: {report_data['total_documents']}\n"
            f"- Всего слов: {report_data.get('total_words', 0)}\n"
            f"{'=' * 80}\n"
        )
        report_json = {
            "generated_at": datetime.now().isoformat(),
            "summary": {
                "total_documents": report_data["total_documents"],
                "average_confidence": round(report_data["average_confidence"], 2),
                "overall_confidence": round(overall_conf, 2),
                "total_words": report_data.get("total_words", 0),
            },
            "documents": report_data["documents"],
        }
        return {
            "success": True,
            "report_text": report_text,
            "report_json": report_json,
            "summary": {
                "total_documents": report_data["total_documents"],
                "average_confidence": round(report_data["average_confidence"], 2),
                "overall_confidence": round(overall_conf, 2),
                "total_words": report_data.get("total_words", 0),
            },
        }
    except Exception as e:
        logger.exception("Ошибка при генерации отчета")
        raise HTTPException(status_code=500, detail=str(e)) from e


@router.get("/report/download")
async def download_confidence_report():
    require_service("rag")  # FEATURE-FLAG
    if not rag_client:
        raise HTTPException(status_code=503, detail="RAG service недоступен")
    try:
        import tempfile

        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        report_data = await rag_client.get_confidence_report()
        logger.info(f"Получены данные отчета: {report_data['total_documents']} документов")
        wb = Workbook()
        ws = wb.active
        ws.title = "Отчет об уверенности"
        header_font = Font(bold=True, size=14, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        subheader_font = Font(bold=True, size=12)
        subheader_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
        high_confidence_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        medium_confidence_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        low_confidence_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin")
        )
        current_row = 1
        ws.merge_cells(f"A{current_row}:D{current_row}")
        header_cell = ws[f"A{current_row}"]
        header_cell.value = "ОТЧЕТ О СТЕПЕНИ УВЕРЕННОСТИ МОДЕЛИ В РАСПОЗНАННОМ ТЕКСТЕ"
        header_cell.font = header_font
        header_cell.fill = header_fill
        header_cell.alignment = header_alignment
        header_cell.border = thin_border
        current_row += 1
        ws.merge_cells(f"A{current_row}:D{current_row}")
        date_cell = ws[f"A{current_row}"]
        date_cell.value = f"Дата генерации: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        date_cell.alignment = Alignment(horizontal="center")
        current_row += 2
        if report_data["total_documents"] == 0:
            ws.merge_cells(f"A{current_row}:D{current_row}")
            warning_cell = ws[f"A{current_row}"]
            warning_cell.value = "ВНИМАНИЕ: Нет обработанных документов для формирования отчета."
            warning_cell.font = Font(bold=True, color="FF0000")
            warning_cell.alignment = Alignment(horizontal="center")
            current_row += 1
        else:
            info_row = current_row
            ws[f"A{info_row}"] = "ОБЩАЯ ИНФОРМАЦИЯ:"
            ws[f"A{info_row}"].font = subheader_font
            ws[f"A{info_row}"].fill = subheader_fill
            current_row += 1
            ws[f"A{current_row}"] = "Всего обработано документов:"
            ws[f"B{current_row}"] = report_data["total_documents"]
            current_row += 1
            ws[f"A{current_row}"] = "Средняя уверенность модели:"
            ws[f"B{current_row}"] = f"{report_data['average_confidence']:.2f}%"
            current_row += 1
            ws[f"A{current_row}"] = "Всего слов:"
            ws[f"B{current_row}"] = report_data.get("total_words", 0)
            current_row += 2
            for doc_idx, doc in enumerate(report_data.get("documents", []), 1):
                ws.merge_cells(f"A{current_row}:D{current_row}")
                doc_header = ws[f"A{current_row}"]
                doc_header.value = f"{doc_idx}. {doc.get('filename', 'Неизвестный файл')}"
                doc_header.font = subheader_font
                doc_header.fill = subheader_fill
                doc_header.border = thin_border
                current_row += 1
                ws[f"A{current_row}"] = "Тип файла:"
                ws[f"B{current_row}"] = doc.get("file_type", "unknown")
                current_row += 1
                ws[f"A{current_row}"] = "Уверенность модели:"
                conf_value = doc.get("confidence", 0.0)
                ws[f"B{current_row}"] = f"{conf_value:.2f}%"
                if conf_value >= 80:
                    ws[f"B{current_row}"].fill = high_confidence_fill
                elif conf_value >= 50:
                    ws[f"B{current_row}"].fill = medium_confidence_fill
                else:
                    ws[f"B{current_row}"].fill = low_confidence_fill
                current_row += 1
                ws[f"A{current_row}"] = "Длина текста:"
                ws[f"B{current_row}"] = f"{doc.get('text_length', 0)} символов"
                current_row += 1
                ws[f"A{current_row}"] = "Количество слов:"
                ws[f"B{current_row}"] = doc.get("words_count", 0)
                current_row += 2
                formatted_text_info = next(
                    (ft for ft in report_data.get("formatted_texts", []) if ft.get("filename") == doc.get("filename")),
                    None,
                )
                if formatted_text_info and formatted_text_info.get("words"):
                    words = formatted_text_info.get("words", [])
                    if words:
                        ws[f"A{current_row}"] = "Слово"
                        ws[f"B{current_row}"] = "Уверенность"
                        ws[f"A{current_row}"].font = Font(bold=True)
                        ws[f"B{current_row}"].font = Font(bold=True)
                        ws[f"A{current_row}"].fill = subheader_fill
                        ws[f"B{current_row}"].fill = subheader_fill
                        ws[f"A{current_row}"].border = thin_border
                        ws[f"B{current_row}"].border = thin_border
                        current_row += 1
                        for word_info in words:
                            word = word_info.get("word", "")
                            conf = word_info.get("confidence", 0.0)
                            if word:
                                ws[f"A{current_row}"] = word
                                ws[f"B{current_row}"] = f"{conf:.1f}%"
                                ws[f"A{current_row}"].border = thin_border
                                ws[f"B{current_row}"].border = thin_border
                                if conf >= 80:
                                    ws[f"B{current_row}"].fill = high_confidence_fill
                                elif conf >= 50:
                                    ws[f"B{current_row}"].fill = medium_confidence_fill
                                else:
                                    ws[f"B{current_row}"].fill = low_confidence_fill
                                current_row += 1
                current_row += 1
            overall_conf = report_data.get("overall_confidence", report_data.get("average_confidence", 0.0))
            ws.merge_cells(f"A{current_row}:D{current_row}")
            summary_header = ws[f"A{current_row}"]
            summary_header.value = "ИТОГО"
            summary_header.font = subheader_font
            summary_header.fill = subheader_fill
            summary_header.border = thin_border
            current_row += 1
            ws[f"A{current_row}"] = "Итоговая уверенность по всему тексту:"
            ws[f"B{current_row}"] = f"{overall_conf:.2f}%"
            if overall_conf >= 80:
                ws[f"B{current_row}"].fill = high_confidence_fill
            elif overall_conf >= 50:
                ws[f"B{current_row}"].fill = medium_confidence_fill
            else:
                ws[f"B{current_row}"].fill = low_confidence_fill
            current_row += 1
            ws[f"A{current_row}"] = "Средняя уверенность по документам:"
            ws[f"B{current_row}"] = f"{report_data['average_confidence']:.2f}%"
            current_row += 1
            ws[f"A{current_row}"] = "Всего документов:"
            ws[f"B{current_row}"] = report_data["total_documents"]
            current_row += 1
            ws[f"A{current_row}"] = "Всего слов:"
            ws[f"B{current_row}"] = report_data.get("total_words", 0)
        ws.column_dimensions["A"].width = 50
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 15
        temp_dir = tempfile.gettempdir()
        report_filename = f"confidence_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        report_path = os.path.join(temp_dir, report_filename)
        try:
            os.makedirs(temp_dir, exist_ok=True)
            wb.save(report_path)
            logger.info(f"Excel отчет сохранен: {report_path}")
            return FileResponse(
                report_path,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                filename=report_filename,
                headers={"Content-Disposition": f"attachment; filename*=UTF-8''{report_filename}"},
            )
        except Exception as e:
            logger.exception("Ошибка при сохранении Excel")
            raise HTTPException(status_code=500, detail="Ошибка при сохранении Excel") from e
    except HTTPException:
        raise
    except Exception as e:
        logger.exception("Ошибка операции")
        logger.error(f"Ошибка при генерации Excel: {str(e)}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e)) from e


@router.post("/extract")
async def extract_document_inline(request: Request, file: Annotated[UploadFile, File(...)]):
    """
    Устаревший путь: только извлечение в памяти без MinIO.
    Предпочтительно: POST /attach (MinIO + inline для модели).
    """
    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="Файл пустой")
    filename = file.filename or "unknown"
    content_type = _resolve_content_type(filename, file.content_type or "")
    file_size = len(content)
    try:
        payload = _extract_inline_payload(content, filename, content_type)
    except HTTPException:
        raise
    except Exception as e:
        logger.exception("Inline extract error (%s)", filename)
        raise HTTPException(status_code=500, detail=f"Не удалось обработать файл: {e}") from e
    log_cef_event(
        "FS007", request=request, status_code=200, extra={"fname": filename, "fsize": file_size, "cs1": payload["cs1"]}
    )
    return {"type": payload["type"], "content": payload["content"], "filename": filename, "success": True}


@router.post("/attach")
async def attach_document_for_message(request: Request, file: Annotated[UploadFile, File(...)]):
    """
    Прикрепление к сообщению: сохранение в MinIO + извлечение содержимого для модели.
    Без RAG, эмбеддингов и PostgreSQL — только объектное хранилище и inline-контент.
    """
    _log_chat_attach_request(request, upload_filename=file.filename, upload_content_type=file.content_type)
    filename = file.filename or "unknown"
    try:
        content = await file.read()
    except Exception as exc:
        logger.exception("[ChatAttach] read-upload-failed filename=%s", file.filename)
        _log_inline_attach_upload_failure(filename, f"Не удалось прочитать файл: {exc}")
        raise HTTPException(status_code=400, detail=f"Не удалось прочитать файл: {exc}") from exc
    if not content:
        _log_inline_attachment_debug("rejected-empty", {"filename": filename})
        _log_inline_attach_upload_failure(filename, "Файл пустой")
        raise HTTPException(status_code=400, detail="Файл пустой")
    content_type = _resolve_content_type(filename, file.content_type or "")
    file_size = len(content)
    _log_inline_attachment_debug(
        "read-complete",
        {
            "filename": filename,
            "content_type": content_type,
            "size": file_size,
            "size_mb": round(file_size / (1024 * 1024), 3),
            "header_content_length": request.headers.get("content-length"),
        },
    )
    analysis = await _analyze_inline_attachment_for_model(filename, content_type, file_size)
    _log_inline_attachment_debug("validate", analysis)
    if not analysis["allowed"]:
        _log_inline_attachment_debug("rejected", analysis)
        if not analysis["format_allowed"]:
            failure_detail = _unsupported_inline_attach_extension_detail(filename)
        else:
            max_mb = INLINE_ATTACHMENT_MAX_BYTES // (1024 * 1024)
            failure_detail = f"размер файла превышает допустимый лимит {max_mb} МБ, поддерживается максимальный размер до {max_mb} МБ"
        _log_inline_attach_upload_failure(filename, failure_detail)
        raise HTTPException(status_code=400, detail=analysis.get("reject_reason") or "Файл не может быть прикреплён")
    if not analysis["model_compatible"]:
        _log_inline_attachment_debug("model-incompatible", analysis)
    _log_inline_attach_upload_success(filename)
    minio_object, minio_bucket = _try_upload_bytes_to_minio(request, content, filename, content_type)
    _log_inline_attachment_debug(
        "minio-upload", analysis, minio_saved=bool(minio_object), minio_object=minio_object, minio_bucket=minio_bucket
    )
    try:
        payload = _extract_inline_payload(content, filename, content_type)
    except Exception as e:
        logger.exception("[ChatAttach] extract-error filename=%s size=%s", filename, file_size)
        _log_inline_attachment_debug("extract-error", analysis, error=str(e), error_type=type(e).__name__)
        _log_inline_attach_upload_failure(filename, f"Не удалось извлечь содержимое: {e}")
        if minio_object and minio_bucket and minio_client:
            with logged_suppress(logger):
                minio_client.delete_file(minio_object, bucket_name=minio_bucket)
        raise HTTPException(status_code=500, detail=f"Не удалось извлечь содержимое: {e}") from e
    inline_content = payload.get("content") or ""
    inline_content_chars = len(inline_content)
    _log_inline_attachment_debug(
        "extract-complete",
        analysis,
        extracted_type=payload["type"],
        extracted_cs1=payload.get("cs1"),
        inline_content_chars=inline_content_chars,
        minio_saved=bool(minio_object),
    )
    _cef_extra: dict = {"fname": filename, "fsize": file_size, "cs1": payload["cs1"]}
    if minio_object:
        _cef_extra["cs2"] = minio_object
        _cef_extra["cs2Label"] = "MinioObject"
    log_cef_event("FS007", request=request, status_code=200, extra=_cef_extra)
    _log_inline_attachment_debug(
        "success",
        analysis,
        extracted_type=payload["type"],
        minio_saved=bool(minio_object),
        response_content_chars=inline_content_chars,
    )
    result = {
        "success": True,
        "type": payload["type"],
        "content": payload["content"],
        "filename": filename,
        "minio_saved": bool(minio_object),
    }
    if minio_object:
        result["minio_object"] = minio_object
        result["minio_bucket"] = minio_bucket
    if not minio_object:
        result["warning"] = "MinIO недоступен — файл прикреплён к сообщению, но не сохранён в хранилище"
    return result


@router.get("/inline-file")
async def get_inline_attachment_file(
    bucket: Annotated[str, Query(..., min_length=1)],
    object_name: Annotated[str, Query(..., min_length=1, alias="object")],
    _current_user: Annotated[dict, Depends(get_current_user)],
):
    """Отдаёт файл вложения из MinIO для превью в UI (после перезагрузки страницы)."""
    if not minio_client:
        raise HTTPException(status_code=503, detail="MinIO недоступен")
    try:
        data = minio_client.download_file(object_name, bucket_name=bucket)
    except Exception as e:
        logger.exception("inline-file download /")
        raise HTTPException(status_code=404, detail="Файл не найден") from e
    ext = os.path.splitext(object_name)[1].lower()
    mime = _CONTENT_TYPE_BY_EXT.get(ext, "application/octet-stream")
    return Response(content=data, media_type=mime)
